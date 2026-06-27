import sys
import os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
"""
URL Accountability Diagnostic - reads from existing completed job outputs.
Answers all 10 questions with actual runtime values.
"""
import os
import sys
import json
import zipfile
import tempfile
import openpyxl

# Force ASCII output to avoid Windows cp1252 encoding issues
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

BASE = os.path.join("backend", "outputs")

# Find the most recently completed job that has a zip
jobs = []
for fname in os.listdir(BASE):
    if fname.endswith(".zip"):
        full = os.path.join(BASE, fname)
        jobs.append((os.path.getmtime(full), full, fname))
jobs.sort(reverse=True)

if not jobs:
    print("No ZIP files found in backend/outputs")
    sys.exit(1)

# Use most recent zip
_, zip_path, zip_name = jobs[0]
print(f"\n{'='*70}")
print(f"Inspecting most recent job ZIP: {zip_name}")
print(f"Full path: {os.path.abspath(zip_path)}")
print(f"ZIP size: {os.path.getsize(zip_path):,} bytes")
print("="*70)

with zipfile.ZipFile(zip_path, "r") as z:
    namelist = z.namelist()
    print(f"\nQ7. diagnostic_report.xlsx in ZIP: {'diagnostic_report.xlsx' in namelist}")
    print(f"Q8. ZIP contents: {namelist}")
    print(f"\nQ8. ZIP packaging code path:")
    print(f"    batch_processor.py line 906-931:")
    print(f"    zip_filepath = os.path.join(BASE_OUTPUT_DIR, f'job_{{job_id}}.zip')")
    print(f"    with zipfile.ZipFile(zip_filepath, 'w', zipfile.ZIP_DEFLATED) as zipf:")
    print(f"      zipf.write(summary_path, 'collection_summary.txt')")
    print(f"      zipf.write(not_found_path, 'not_found_report.txt')")
    print(f"      zipf.write(diagnostic_csv_path, 'diagnostic_report.csv')")
    print(f"      zipf.write(diagnostic_xlsx_path, 'diagnostic_report.xlsx')  <- line 928")
    print(f"      zipf.write(diagnostic_txt_path, 'diagnostic_report.txt')")

    if "diagnostic_report.xlsx" not in namelist:
        print("ERROR: diagnostic_report.xlsx not in ZIP!")
        sys.exit(1)

    td = tempfile.mkdtemp()
    local_xlsx = z.extract("diagnostic_report.xlsx", td)
    print(f"\nQ3. diagnostic_report.xlsx EXISTS: YES")
    print(f"Q4. Exact path inside ZIP: diagnostic_report.xlsx")
    print(f"    Extracted to: {local_xlsx}")
    print(f"    File size: {os.path.getsize(local_xlsx):,} bytes")

    wb = openpyxl.load_workbook(local_xlsx)
    sheets = wb.sheetnames
    print(f"\nQ5. Worksheet names: {sheets}")

    print(f"\nQ6. Row counts per sheet:")
    for sname in sheets:
        ws = wb[sname]
        all_rows = list(ws.iter_rows(values_only=True))
        header = all_rows[0] if all_rows else ()
        data = all_rows[1:] if len(all_rows) > 1 else []
        print(f"\n  [{sname}]")
        print(f"    Total rows (including header): {len(all_rows)}")
        print(f"    Data rows: {len(data)}")
        print(f"    Header: {header}")
        for i, row in enumerate(data[:3], 1):
            print(f"    Row {i}: {row}")

print(f"\n{'='*70}")
print("QUESTIONS 1-2: url_info_map - reading from batch_processor source")
print("="*70)
print("""
The url_info_map is built inside collect_logs_from_web() in crawler.py:

  url_info_map = {}  # initialized at line 581

  For each search result URL (line 612-630):
    url_info_map[url] = {
      "Platform": platform,
      "Product": product_name,
      "Log Type": service,
      "URL": url,
      "Title": res.get("title", ""),
      "Source Rank": r_rank,
      "Search Rank": len(search_results_ranked),
      "Search Query Used": query_used,
      "Search Position": pos,
      "Source Type": source_type,
      "Crawled": "No",
      "Logs Extracted": 0,
      "Logs Validated": 0,
      "Logs Rejected": 0,
      "Status": "no_logs_found",
      "Reason": "Search result not visited ...",
      "Failure Stage": "search_filtering"
    }

  After crawling each URL, status is updated (lines 641-662).
  After extraction/validation, Logs Extracted/Validated/Rejected are updated (lines 775-802).

  raw_results_count = len(url_info_map)  <- line 835
  scraped_text.url_info_map = url_info_map  <- line 839

  In batch_processor.py line 442-443:
    all_urls_info.extend(scraped_text.url_info_map.values())

  Then line 762-794:
    for info in all_urls_info:
      ws_audit.append([...])  -> SOURCE_AUDIT row
      if info.get("Logs Validated", 0) == 0:
        ws_unprocessed.append([...])  -> UNPROCESSED_SOURCES row
""")

print(f"\n{'='*70}")
print("LIVE PROBE: Running collect_logs_from_web for 1 row to get actual counts")
print("="*70)

import logging
logging.disable(logging.CRITICAL)  # suppress log spam
os.environ["SQLITE_DB_PATH"] = ":memory:"

from backend import db_manager
from backend.crawler import collect_logs_from_web

db_manager.init_db()

print("\nCalling collect_logs_from_web('Docker', '25', 'daemon', 3)...")
result = collect_logs_from_web("Docker", "25", "daemon", 3)

umap = getattr(result, "url_info_map", {})
raw_count = getattr(result, "raw_results_count", 0)
visited = getattr(result, "visited_urls", [])
queries = getattr(result, "queries", [])
extracted = getattr(result, "extracted_logs", [])

print(f"\nQ1. url_info_map size: {len(umap)} URLs")
print(f"    raw_results_count attribute: {raw_count}")
print(f"    visited_urls count: {len(visited)}")
print(f"    queries used: {len(queries)}")
print(f"    extracted_logs count: {len(extracted)}")

print(f"\nQ2. First 5 entries in url_info_map:")
for i, (url, info) in enumerate(list(umap.items())[:5], 1):
    print(f"\n  [{i}] URL: {url}")
    for k, v in info.items():
        print(f"       {k}: {v!r}")

# Count statuses
status_counts = {}
crawled_count = 0
for url, info in umap.items():
    st = info.get("Status", "unknown")
    status_counts[st] = status_counts.get(st, 0) + 1
    if info.get("Crawled") == "Yes":
        crawled_count += 1

print(f"\nStatus breakdown across all {len(umap)} URLs:")
for st, cnt in sorted(status_counts.items()):
    print(f"  {st}: {cnt}")
print(f"Crawled=Yes: {crawled_count}")

validated_urls = sum(1 for info in umap.values() if info.get("Logs Validated", 0) > 0)
unprocessed_urls = sum(1 for info in umap.values() if info.get("Logs Validated", 0) == 0)
print(f"\nURLs with Logs Validated > 0 (-> SOURCE_AUDIT only):         {validated_urls}")
print(f"URLs with Logs Validated == 0 (-> UNPROCESSED_SOURCES + SOURCE_AUDIT): {unprocessed_urls}")

print(f"\nQ9/10. Explanation of sheet row counts:")
print(f"  SOURCE_AUDIT row count = len(all_urls_info) = {len(umap)} (one row per search result URL)")
print(f"  UNPROCESSED_SOURCES row count = URLs where Logs Validated == 0 = {unprocessed_urls}")
print(f"\n  If SOURCE_AUDIT = 0 rows: url_info_map is empty because search returned 0 results")
print(f"  If UNPROCESSED_SOURCES = 0 rows: every URL produced at least 1 validated log (unlikely)")

print(f"\n{'='*70}")
print("DIAGNOSTIC COMPLETE")
print("="*70)
