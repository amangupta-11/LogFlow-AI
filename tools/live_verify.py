import sys
import os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
"""
Live batch verification script.
Runs a 2-row batch job with real crawling (no mocks) and validates
that diagnostic_report.xlsx contains all 4 required sheets.
"""
import os
import csv
import tempfile
import zipfile
import openpyxl

os.environ["SQLITE_DB_PATH"] = ":memory:"

from backend import db_manager, batch_processor

db_manager.init_db()

# Create CSV
fd, path = tempfile.mkstemp(suffix=".csv")
os.close(fd)
with open(path, "w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerow(["Platform", "Version", "Service", "Max Logs"])
    writer.writerow(["Docker", "25", "daemon", "3"])
    writer.writerow(["Nginx", "1.25", "upstream", "3"])

rows = batch_processor.parse_uploaded_file(path)
print(f"Parsed {len(rows)} rows from CSV")

job_id = "live_verification_job"
db_manager.create_job(job_id, len(rows))
for r in rows:
    db_manager.add_job_row(job_id, r["platform"], r["version"], r["service"], r["max_logs"])

print("Job created. Starting batch processing...")
batch_processor.process_batch_job(job_id)

job = db_manager.get_job(job_id)
print(f"Job status: {job['status']}")
print(f"Completed rows: {job['completed_rows']}")
print(f"Failed rows: {job['failed_rows']}")
print(f"ZIP path: {job['zip_path']}")

if job["zip_path"] and os.path.exists(job["zip_path"]):
    with zipfile.ZipFile(job["zip_path"], "r") as z:
        namelist = z.namelist()
        print(f"ZIP contents: {namelist}")

        has_xlsx = "diagnostic_report.xlsx" in namelist
        print(f"diagnostic_report.xlsx present: {has_xlsx}")

        if has_xlsx:
            td = tempfile.mkdtemp()
            xlsx_path = z.extract("diagnostic_report.xlsx", td)
            wb = openpyxl.load_workbook(xlsx_path)
            print(f"Excel sheets: {wb.sheetnames}")

            # SUMMARY
            ws_sum = wb["SUMMARY"]
            sum_rows = list(ws_sum.iter_rows(values_only=True))
            print(f"\nSUMMARY rows (header+data): {len(sum_rows)}")
            for r in sum_rows:
                print(f"  {r}")

            # SOURCE_AUDIT
            ws_audit = wb["SOURCE_AUDIT"]
            audit_rows = list(ws_audit.iter_rows(values_only=True))
            print(f"\nSOURCE_AUDIT rows (header+data): {len(audit_rows)}")
            for r in audit_rows[:6]:
                print(f"  {r}")

            # UNPROCESSED_SOURCES
            ws_unp = wb["UNPROCESSED_SOURCES"]
            unp_rows = list(ws_unp.iter_rows(values_only=True))
            print(f"\nUNPROCESSED_SOURCES rows (header+data): {len(unp_rows)}")
            for r in unp_rows[:6]:
                print(f"  {r}")

            # SEARCH_DIAGNOSTICS
            ws_diag = wb["SEARCH_DIAGNOSTICS"]
            diag_rows = list(ws_diag.iter_rows(values_only=True))
            print(f"\nSEARCH_DIAGNOSTICS rows (header+data): {len(diag_rows)}")
            for r in diag_rows[:6]:
                print(f"  {r}")

            print(f"\nWORKBOOK SAVED AT: {xlsx_path}")
else:
    print("ERROR: ZIP or XLSX not found!")

os.remove(path)
print("\nDONE")
