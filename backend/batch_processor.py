import os
import re
import csv
import zipfile
import tempfile
import logging
import threading
import hashlib
from datetime import datetime, timedelta
import openpyxl

from backend import db_manager
from backend.crawler import collect_logs_from_web
from backend.extractor import parse_logs_with_llm, parse_logs_with_regex, map_source_urls, safe_to_text
from backend.validator import validate_logs_with_claude
from backend.vector_store import vector_store


logger = logging.getLogger(__name__)

# Determine a Vercel-compatible output directory
if os.getenv("VERCEL") or not os.access(os.path.dirname(os.path.abspath(__file__)), os.W_OK):
    BASE_OUTPUT_DIR = os.path.join(tempfile.gettempdir(), "log_collector_outputs")
else:
    BASE_OUTPUT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "outputs"))

# Ensure the output directory exists
os.makedirs(BASE_OUTPUT_DIR, exist_ok=True)

# Global lock to prevent concurrent processing of the same job rows
_processing_lock = threading.Lock()

def send_slack_notification(message: str):
    """
    Hook for future Slack notifications.
    Currently disabled but prepared with environment check and logging.
    """
    webhook_url = os.getenv("SLACK_WEBHOOK_URL")
    logger.info(f"[Slack Hook] Notification prepared: {message}")
    if webhook_url:
        # Future implementation:
        # import requests
        # requests.post(webhook_url, json={"text": message})
        pass

def merge_and_deduplicate_logs(llm_logs: list, regex_logs: list) -> list:
    combined = []
    seen_messages = set()
    for log in llm_logs:
        msg = safe_to_text(log.get("message")).strip()
        if not msg:
            continue
        msg_lower = msg.lower()
        if msg_lower not in seen_messages:
            seen_messages.add(msg_lower)
            combined.append(log)
    for log in regex_logs:
        msg = safe_to_text(log.get("message")).strip()
        if not msg:
            continue
        msg_lower = msg.lower()
        if msg_lower not in seen_messages:
            seen_messages.add(msg_lower)
            combined.append(log)
    return combined

def sanitize_filename(platform: str, service: str, product_name: str = "") -> str:
    p = platform.lower().strip()
    prod = product_name.lower().strip() if product_name else ""
    s = service.lower().strip() if service else ""
    
    parts = []
    if p:
        parts.append(p)
    if prod:
        parts.append(prod)
    if s:
        parts.append(s)
        
    combined = "_".join(parts)
    clean = re.sub(r'[^a-z0-9]+', '_', combined).strip('_')
    clean = re.sub(r'_+', '_', clean)
    
    return f"{clean}.txt" if clean else "log.txt"

def extract_product_and_version(model_version_str):
    if not model_version_str:
        return "", ""
    # Look for a parenthesized year range pattern like (2010-present) or (2006-2012) or (2004) at the end of the string
    match = re.search(r'\s*\(([^)]*\b\d{4}\b[^)]*)\)\s*$', model_version_str)
    if match:
        version = match.group(1).strip()
        product_name = model_version_str[:match.start()].strip()
        return product_name, version
    return "", model_version_str.strip()

def detect_log_format(text: str) -> str:
    text_strip = text.strip()
    
    # 1. CloudWatch
    has_cloudwatch = any(sig in text_strip for sig in ["START RequestId:", "REPORT RequestId:", "END RequestId:"])
    if has_cloudwatch:
        return "CloudWatch"
        
    # 2. Syslog
    has_syslog_header = re.search(r'\b[a-zA-Z0-9_-]+\s+[a-zA-Z0-9_./-]+(?:\[\d+\])?:\s+', text_strip) is not None
    if has_syslog_header:
        return "Syslog"
        
    # 3. Apache/Nginx
    has_http_status = re.search(r'\bHTTP/\d\.\d\s*"?\s*[1-5]\d{2}\b|"(?:GET|POST|PUT|DELETE|PATCH|HEAD|OPTIONS|CONNECT)\s+', text_strip, re.IGNORECASE) is not None
    has_client = re.search(r'\b\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\b', text_strip) is not None
    has_request = re.search(r'\bHTTP/\d\.\d\b|"(?:GET|POST|PUT|DELETE|PATCH|HEAD|OPTIONS|CONNECT)\s+', text_strip, re.IGNORECASE) is not None
    if has_http_status or (has_client and has_request):
        return "Apache"
        
    # 4. Application
    has_severity = re.search(r'\b(INFO|WARN|WARNING|ERROR|FATAL|DEBUG|CRITICAL|TRACE|crit|warn|err|info|dbg|stdout|stderr)\b', text_strip, re.IGNORECASE) is not None
    has_stacktrace = text_strip.startswith("at ") or re.search(r'\bat\s+[a-zA-Z0-9_]+\.[a-zA-Z0-9_]+', text_strip) is not None
    ex_match = re.search(r'\b([A-Z][a-zA-Z0-9_]*Exception|[A-Z][a-zA-Z0-9_]*Error|Exception|Traceback|Caused\s+by:)\b', text_strip)
    has_exception = False
    if ex_match:
        matched_word = ex_match.group(1)
        if matched_word.upper() != "ERROR":
            has_exception = True
    if has_severity or has_stacktrace or has_exception:
        return "Application"
        
    return "Unknown"

def map_headers_to_indices(headers):
    mapped = {}
    for idx, h in enumerate(headers):
        if h is None:
            continue
        h_norm = str(h).strip().lower()
        h_clean = re.sub(r'\s+', ' ', h_norm)
        # Remove slash spaces: "model / version" -> "model/version"
        h_clean = h_clean.replace(" / ", "/").replace(" /", "/").replace("/ ", "/")
        
        if h_clean == "platform":
            mapped["platform"] = idx
        elif h_clean in ("version", "model/version"):
            mapped["version"] = idx
        elif h_clean in ("log type", "service"):
            mapped["service"] = idx
        elif h_clean == "error code":
            mapped["error_code"] = idx
        elif h_clean == "error message":
            mapped["excel_error_message"] = idx
        elif h_clean == "reason":
            mapped["reason"] = idx
        elif h_clean == "source":
            mapped["source"] = idx
        elif h_clean == "error message long":
            mapped["error_message_long"] = idx
        elif h_clean == "category":
            mapped["category"] = idx
        elif h_clean in ("max logs", "maxlogs", "count", "limit"):
            mapped["max_logs"] = idx
    return mapped

def idx_in_row(row, idx):
    return idx is not None and 0 <= idx < len(row) and row[idx] is not None

def build_row_dict(row, mapped):
    row_dict = {}
    
    # Platform
    platform_val = str(row[mapped["platform"]]).strip() if "platform" in mapped and idx_in_row(row, mapped["platform"]) else ""
    row_dict["platform"] = platform_val
    
    # Model / Version -> Version and Product Name extraction
    model_version_val = str(row[mapped["version"]]).strip() if "version" in mapped and idx_in_row(row, mapped["version"]) else ""
    product_name, version_val = extract_product_and_version(model_version_val)
    row_dict["product_name"] = product_name
    row_dict["version"] = version_val
    
    # Log Type / Service
    row_dict["service"] = str(row[mapped["service"]]).strip() if "service" in mapped and idx_in_row(row, mapped["service"]) else ""
    
    # Optional fields
    row_dict["error_code"] = str(row[mapped["error_code"]]).strip() if "error_code" in mapped and idx_in_row(row, mapped["error_code"]) else ""
    row_dict["excel_error_message"] = str(row[mapped["excel_error_message"]]).strip() if "excel_error_message" in mapped and idx_in_row(row, mapped["excel_error_message"]) else ""
    row_dict["reason"] = str(row[mapped["reason"]]).strip() if "reason" in mapped and idx_in_row(row, mapped["reason"]) else ""
    row_dict["source"] = str(row[mapped["source"]]).strip() if "source" in mapped and idx_in_row(row, mapped["source"]) else ""
    row_dict["error_message_long"] = str(row[mapped["error_message_long"]]).strip() if "error_message_long" in mapped and idx_in_row(row, mapped["error_message_long"]) else ""
    row_dict["category"] = str(row[mapped["category"]]).strip() if "category" in mapped and idx_in_row(row, mapped["category"]) else ""
    
    # Max Logs
    max_logs_val = str(row[mapped["max_logs"]]).strip() if "max_logs" in mapped and idx_in_row(row, mapped["max_logs"]) else ""
    try:
        row_dict["max_logs"] = int(float(max_logs_val)) if max_logs_val else 10
    except:
        row_dict["max_logs"] = 10
        
    return row_dict

def parse_uploaded_file(file_path: str) -> list:
    """
    Parses a CSV or Excel file and returns a list of dictionaries.
    """
    results = []
    _, ext = os.path.splitext(file_path.lower())
    
    if ext == ".csv":
        with open(file_path, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            headers = next(reader, None)
            if not headers:
                return []
            
            mapped = map_headers_to_indices(headers)
            if "platform" not in mapped or "version" not in mapped:
                logger.warning("Uploaded CSV file is missing required headers: Platform and/or Version/Model / Version")
                return []
                
            for row in reader:
                if not row or not any(row):
                    continue
                row_dict = build_row_dict(row, mapped)
                results.append(row_dict)
                
    elif ext == ".xlsx":
        wb = openpyxl.load_workbook(file_path, read_only=True)
        try:
            sheet = wb.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return []
            
            headers = rows[0]
            mapped = map_headers_to_indices(headers)
            if "platform" not in mapped or "version" not in mapped:
                logger.warning("Uploaded Excel file is missing required headers: Platform and/or Version/Model / Version")
                return []
                
            for row in rows[1:]:
                if not row or not any(row):  # skip empty rows
                    continue
                row_dict = build_row_dict(row, mapped)
                results.append(row_dict)
        finally:
            wb.close()
            
    return results

# Removed get_file_hash.

def process_batch_job(job_id: str):
    """
    Background worker function that processes a batch job row by row.
    """
    with _processing_lock:
        try:
            job = db_manager.get_job(job_id)
            if not job:
                logger.error(f"Job {job_id} not found in DB.")
                return
            
            logger.info(f"Starting processing of batch job {job_id}")
            db_manager.update_job_status(job_id, status="processing")
            
            rows = db_manager.get_job_rows(job_id)
            
            # Create a temporary working directory for this job's file outputs
            job_work_dir = os.path.join(BASE_OUTPUT_DIR, f"job_{job_id}")
            os.makedirs(job_work_dir, exist_ok=True)
            
            # Keep track of outputs
            summary_data = []
            skipped_rows_data = []
            all_validated_logs = []
            all_urls_info = []
            all_search_diagnostics = []
            
            row_outputs = []
            not_found_report_entries = []
            used_filenames = set()
            row_diagnostics = []
            
            for row in rows:
                row_id = row["id"]
                platform = row["platform"]
                version = row["version"]
                service = row["service"]
                product_name = row.get("product_name") or ""
                max_logs = row["max_logs"] or 10
                
                # Check row status - skip if already processed
                if row["status"] in ("completed", "failed", "skipped"):
                    continue
                
                logger.info(f"Processing row {row_id}: platform={platform}, product_name={product_name}, version={version}")
                db_manager.update_row_status(row_id, status="processing")
                
                # Initialize diagnostic data for this row
                diag = {
                    "Platform": platform,
                    "Product Name": product_name,
                    "Log Type": service,
                    "Generated Query": [],
                    "Search Results Count": 0,
                    "Extracted Logs Count": 0,
                    "Validated Logs Count": 0,
                    "Rejected Logs Count": 0,
                    "Coverage %": 0,
                    "Rejection Reasons": [],
                    "Sample Rejected Logs": [],
                    "Search Result Ranking": [],
                    "Top 10 Source URLs": [],
                    "Validation Format Detected": {
                        "Syslog": 0,
                        "CloudWatch": 0,
                        "Apache": 0,
                        "Application": 0,
                        "Unknown": 0
                    },
                    "Failure Reason": ""
                }
                
                # Validation rules: Platform and Version are required
                if not platform or not version:
                    reason = "Missing required Platform or Version"
                    logger.warning(f"Row {row_id} skipped: {reason}")
                    db_manager.update_row_status(row_id, status="skipped", error_message=reason)
                    db_manager.increment_job_stats(job_id, skipped=1)
                    
                    skipped_rows_data.append({
                        "Platform": platform,
                        "Version": version,
                        "Service": service,
                        "Max Logs": max_logs,
                        "Reason": reason
                    })
                    diag["Failure Reason"] = reason
                    diag["Coverage %"] = 0
                    row_diagnostics.append(diag)
                    continue
                
                # Run normal validation pipeline (No cache checks)
                try:
                    scraped_text = collect_logs_from_web(platform, version, service, max_logs, product_name=product_name)
                    
                    logs = []
                    validated_count = 0
                    non_validated_count = 0
                    sources_found = set()
                    row_validated_logs = []
                    
                    # Store generated queries, search count, rankings, top 10 URLs in diag
                    queries = getattr(scraped_text, "queries", [])
                    raw_results = getattr(scraped_text, "raw_results_count", 0)
                    ranking = getattr(scraped_text, "search_results_ranked", [])
                    visited = getattr(scraped_text, "visited_urls", [])
                    diag["Generated Query"] = queries
                    diag["Search Results Count"] = raw_results
                    diag["Search Result Ranking"] = ranking
                    diag["Top 10 Source URLs"] = visited[:10]
                    
                    search_diags = getattr(scraped_text, "search_diagnostics", [])
                    diag["Search Diagnostics"] = search_diags

                    # Parse/extract logs (even if scraped_text is empty/whitespace)
                    if hasattr(scraped_text, "extracted_logs") and scraped_text.extracted_logs:
                        logs = scraped_text.extracted_logs
                    else:
                        if scraped_text and str(scraped_text).strip():
                            llm_logs = parse_logs_with_llm(scraped_text, platform, version, service, max_logs)
                            regex_logs = parse_logs_with_regex(scraped_text, platform, version, service, max_logs)
                        else:
                            llm_logs = []
                            regex_logs = []

                        if llm_logs:
                            llm_logs = map_source_urls(llm_logs, scraped_text)
                        if regex_logs:
                            regex_logs = map_source_urls(regex_logs, scraped_text)

                        logs = merge_and_deduplicate_logs(llm_logs, regex_logs)

                        if logs:
                            logs = validate_logs_with_claude(logs, scraped_text, platform, version)

                    if logs:
                        # Add to semantic vector store
                        vector_store.add_logs(logs, platform, version, service)

                        # Ingest to central validated logs repository
                        db_manager.insert_validated_logs(
                            logs,
                            job_platform=platform,
                            job_product_name=product_name,
                            job_log_type=service,
                            job_error_code=row.get("error_code")
                        )

                        # Count validated vs non-validated
                        for log in logs:
                            url = log.get("source_url")
                            if url and url.startswith("http"):
                                sources_found.add(url)

                            if log.get("validation", {}).get("valid") is True:
                                validated_count += 1

                                # Collect validated log text (with timestamp requirement)
                                original_log = str(log.get("original_log") or "").strip()
                                timestamp = str(log.get("timestamp") or "").strip()
                                if timestamp:
                                    # Prefix timestamp if not already present
                                    if original_log and timestamp not in original_log.splitlines()[0]:
                                        original_log = f"{timestamp} {original_log}".strip()
                                    elif not original_log:
                                        original_log = timestamp
                                if original_log:
                                    all_validated_logs.append(original_log)
                                    row_validated_logs.append(original_log)
                            else:
                                non_validated_count += 1
                    
                    # Fill validation and rejection details
                    diag["Extracted Logs Count"] = len(logs)
                    diag["Validated Logs Count"] = validated_count
                    diag["Rejected Logs Count"] = non_validated_count
                    
                    rejection_reasons = []
                    sample_rejected = []
                    for log in logs:
                        is_valid = log.get("validation", {}).get("valid") is True
                        original_text = safe_to_text(log.get("original_log")).strip() or safe_to_text(log.get("message")).strip()
                        if is_valid:
                            fmt = detect_log_format(original_text)
                            diag["Validation Format Detected"][fmt] = diag["Validation Format Detected"].get(fmt, 0) + 1
                        else:
                            reason = safe_to_text(log.get("validation", {}).get("reason", "Unknown validation failure reason")).strip()
                            if reason not in rejection_reasons:
                                rejection_reasons.append(reason)
                            if len(sample_rejected) < 5:
                                sample_rejected.append({
                                    "log": original_text,
                                    "reason": reason
                                })
                    
                    diag["Rejection Reasons"] = rejection_reasons
                    diag["Sample Rejected Logs"] = sample_rejected
                    
                    from backend.main import calculate_coverage_score
                    diag["Coverage %"] = calculate_coverage_score(logs) if logs else 0
                    
                    if hasattr(scraped_text, "url_info_map") and scraped_text.url_info_map:
                        all_urls_info.extend(scraped_text.url_info_map.values())
                    if hasattr(scraped_text, "search_diagnostics") and scraped_text.search_diagnostics:
                        all_search_diagnostics.extend(scraped_text.search_diagnostics)
                    
                    if getattr(scraped_text, "extraction_error", None):
                        raise ValueError(f"Extraction/validation error: {scraped_text.extraction_error}")
                    
                    # Classify failure if validated count = 0
                    if validated_count == 0:
                        if raw_results == 0:
                            diag["Failure Reason"] = "No search results"
                        elif not str(scraped_text).strip():
                            visited_set = getattr(scraped_text, "visited_urls", [])
                            blocklist = ["medium.com", "linkedin.com", "twitter.com", "superuser.com"]
                            if visited_set and all(any(b in url.lower() for b in blocklist) for url in visited_set):
                                diag["Failure Reason"] = "Source filtering removed all candidates"
                            else:
                                diag["Failure Reason"] = "Search results found but no logs extracted"
                        else:
                            if not logs:
                                diag["Failure Reason"] = "Search results found but no logs extracted"
                            else:
                                 diag["Failure Reason"] = "Logs extracted but validation rejected all"
                                 

                    # Reconciliation validation
                    validated_sources = 0
                    rejected_sources = 0
                    failed_sources = 0
                    blocked_sources = 0
                    lowvalue_sources = 0
                    nolog_sources = 0

                    row_urls_info = scraped_text.url_info_map.values() if hasattr(scraped_text, "url_info_map") and scraped_text.url_info_map else []
                    for info in row_urls_info:
                        if info.get("Logs Validated", 0) > 0:
                            validated_sources += 1
                        elif info.get("Logs Extracted", 0) > 0 and info.get("Logs Validated", 0) == 0:
                            rejected_sources += 1
                        elif info.get("Status") == "low_value_url" or info.get("Failure Stage") == "pre_crawl_filter":
                            lowvalue_sources += 1
                        elif info.get("Status") == "crawl_failed" or info.get("Failure Stage") == "crawling":
                            failed_sources += 1
                        elif info.get("Status") == "access_denied" or info.get("Failure Stage") == "search_filtering":
                            blocked_sources += 1
                        else:
                            nolog_sources += 1

                    total_reconciled = validated_sources + rejected_sources + failed_sources + blocked_sources + lowvalue_sources + nolog_sources
                    expected_count = len(row_urls_info)

                    if total_reconciled != expected_count or total_reconciled != diag["Search Results Count"]:
                        err_msg = (
                            f"Reconciliation validation failed for platform={platform}, service={service}: "
                            f"Total unique URLs in map is {expected_count}, but reconciled sum is {total_reconciled} "
                            f"(Search Results Count={diag['Search Results Count']}, Validated={validated_sources}, Rejected={rejected_sources}, Failed={failed_sources}, Blocked={blocked_sources}, Low-Value={lowvalue_sources}, No-Log={nolog_sources})"
                        )
                        logger.error(err_msg)
                        db_manager.update_row_status(row_id, status="reporting_error", error_message=err_msg)
                        db_manager.increment_job_stats(job_id, failed=1)
                        diag["Failure Reason"] = f"reporting_error: {err_msg}"
                        row_diagnostics.append(diag)
                        not_found_report_entries.append(
                            f"Platform: {platform} | Product: {product_name} | Service: {service} | Version: {version} (Reporting Error: Reconciliation Mismatch)"
                        )
                        continue
                    
                    # Debug printing as per requirements 1 and 4
                    extracted_count = len(logs)
                    originally_valid_count = sum(1 for log in logs if log.get("originally_valid") is True)
                    
                    print(f"ORIGINAL_ROW: platform={platform}, version={version}, service={service}, product_name={product_name}")
                    print(f"PARSED_PLATFORM: {platform}")
                    print(f"PARSED_PRODUCT_NAME: {product_name}")
                    print(f"PARSED_VERSION: {version}")
                    print(f"PARSED_LOG_TYPE: {service}")
                    print(f"GENERATED_SEARCH_QUERY: {queries}")
                    print(f"SEARCH_RESULTS_FOUND: {raw_results}")
                    print(f"EXTRACTED_LOGS_COUNT: {extracted_count}")
                    print(f"VALIDATED_LOGS_COUNT: {originally_valid_count}")
                    
                    print(f"RAW_SEARCH_RESULTS_COUNT: {raw_results}")
                    print(f"EXTRACTED_LOGS_COUNT: {extracted_count}")
                    print(f"VALIDATED_LOGS_COUNT: {originally_valid_count}")
                    
                    # Update row status
                    db_manager.update_row_status(
                        row_id,
                        status="completed",
                        validated_count=validated_count,
                        non_validated_count=non_validated_count,
                        sources_found=len(sources_found)
                    )
                    db_manager.increment_job_stats(job_id, completed=1)
                    
                    summary_data.append({
                        "Platform": platform,
                        "Version": version,
                        "Validated Logs": validated_count,
                        "Non-Validated Logs": non_validated_count,
                        "Sources Found": len(sources_found)
                    })
                    
                    if row_validated_logs:
                        # Generate unique filename for this row
                        base_fn = sanitize_filename(platform, service, product_name=product_name)
                        fn = base_fn
                        counter = 2
                        while fn in used_filenames:
                            base_without_ext = base_fn[:-4] if base_fn.endswith(".txt") else base_fn
                            fn = f"{base_without_ext}_{counter}.txt"
                            counter += 1
                        used_filenames.add(fn)
                        
                        row_outputs.append({
                            "fn": fn,
                            "logs": row_validated_logs,
                            "sources": sources_found
                        })
                    else:
                        not_found_report_entries.append(
                            f"Platform: {platform} | Product: {product_name} | Service: {service} | Version: {version} (No validated logs found)"
                        )
                        
                    row_diagnostics.append(diag)
                    
                except Exception as row_err:
                    logger.error(f"Error processing row {row_id}: {row_err}", exc_info=True)
                    db_manager.update_row_status(row_id, status="failed", error_message=str(row_err))
                    db_manager.increment_job_stats(job_id, failed=1)
                    
                    diag["Failure Reason"] = f"Failed with error: {row_err}"
                    diag["Coverage %"] = 0
                    row_diagnostics.append(diag)
                    
                    print(f"ORIGINAL_ROW: platform={platform}, version={version}, service={service}, product_name={product_name}")
                    print(f"PARSED_PLATFORM: {platform}")
                    print(f"PARSED_PRODUCT_NAME: {product_name}")
                    print(f"PARSED_VERSION: {version}")
                    print(f"PARSED_LOG_TYPE: {service}")
                    print(f"GENERATED_SEARCH_QUERY: []")
                    print(f"SEARCH_RESULTS_FOUND: 0")
                    print(f"EXTRACTED_LOGS_COUNT: 0")
                    print(f"VALIDATED_LOGS_COUNT: 0")
                    
                    print(f"RAW_SEARCH_RESULTS_COUNT: 0")
                    print(f"EXTRACTED_LOGS_COUNT: 0")
                    print(f"VALIDATED_LOGS_COUNT: 0")
                    
                    summary_data.append({
                        "Platform": platform,
                        "Version": version,
                        "Validated Logs": 0,
                        "Non-Validated Logs": 0,
                        "Sources Found": 0
                    })
                    
                    not_found_report_entries.append(
                        f"Platform: {platform} | Product: {product_name} | Service: {service} | Version: {version} (Failed with error: {row_err})"
                    )
            
            # Create a "Sources" sub-directory in job_work_dir
            sources_dir = os.path.join(job_work_dir, "Sources")
            os.makedirs(sources_dir, exist_ok=True)
            
            # Generate platform specific files and sources files
            for ro in row_outputs:
                fn = ro["fn"]
                fp = os.path.join(job_work_dir, fn)
                with open(fp, "w", encoding="utf-8") as f:
                    f.write("\n".join(ro["logs"]))
                    
                fn_sources = fn[:-4] + "_sources.txt" if fn.endswith(".txt") else fn + "_sources.txt"
                fp_sources = os.path.join(sources_dir, fn_sources)
                with open(fp_sources, "w", encoding="utf-8") as f:
                    f.write("\n".join(sorted(list(ro["sources"]))))
            
            # Generate collection_summary.txt content
            summary_lines = []
            summary_lines.append("=== LOG COLLECTION RUN SUMMARY ===")
            summary_lines.append(f"Job ID: {job_id}")
            summary_lines.append(f"Time: {datetime.utcnow().isoformat()}")
            summary_lines.append(f"Total rows in spreadsheet: {len(rows)}")
            summary_lines.append(f"Processed valid rows: {len(rows) - len(skipped_rows_data)}")
            summary_lines.append(f"Skipped invalid rows: {len(skipped_rows_data)}")
            summary_lines.append(f"Rows with logs successfully generated: {len(row_outputs)}")
            summary_lines.append(f"Rows with no logs found / failed: {len(not_found_report_entries)}")
            summary_lines.append("")
            
            summary_lines.append("=== GENERATED LOG FILES ===")
            if row_outputs:
                for idx, ro in enumerate(row_outputs, 1):
                    summary_lines.append(f"{idx}. {ro['fn']} ({len(ro['logs'])} validated logs, {len(ro['sources'])} sources)")
            else:
                summary_lines.append("(None)")
            summary_lines.append("")
            
            summary_lines.append("=== SKIPPED INVALID ROWS ===")
            if skipped_rows_data:
                for idx, sk in enumerate(skipped_rows_data, 1):
                    summary_lines.append(f"{idx}. Platform={sk['Platform']}, Version={sk['Version']}, Service={sk['Service']} -> Reason: {sk['Reason']}")
            else:
                summary_lines.append("(None)")
                
            summary_path = os.path.join(job_work_dir, "collection_summary.txt")
            with open(summary_path, "w", encoding="utf-8") as f:
                f.write("\n".join(summary_lines))
                
            # Generate not_found_report.txt
            not_found_lines = []
            not_found_lines.append("=== NOT FOUND / FAILED ROWS REPORT ===")
            not_found_lines.append(f"Job ID: {job_id}")
            not_found_lines.append("")
            if not_found_report_entries:
                for idx, entry in enumerate(not_found_report_entries, 1):
                    not_found_lines.append(f"{idx}. {entry}")
            else:
                not_found_lines.append("All processed rows successfully found logs.")
                
            not_found_path = os.path.join(job_work_dir, "not_found_report.txt")
            with open(not_found_path, "w", encoding="utf-8") as f:
                f.write("\n".join(not_found_lines))
                
            # Generate diagnostic_report.csv
            diagnostic_csv_path = os.path.join(job_work_dir, "diagnostic_report.csv")
            with open(diagnostic_csv_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow([
                    "Platform", 
                    "Product Name", 
                    "Log Type", 
                    "Generated Query", 
                    "Search Results Count", 
                    "Extracted Logs Count", 
                    "Validated Logs Count", 
                    "Rejected Logs Count", 
                    "Failure Reason",
                    "Coverage %"
                ])
                for d in row_diagnostics:
                    q_str = "; ".join(d["Generated Query"])
                    writer.writerow([
                        d["Platform"],
                        d["Product Name"],
                        d["Log Type"],
                        q_str,
                        d["Search Results Count"],
                        d["Extracted Logs Count"],
                        d["Validated Logs Count"],
                        d["Rejected Logs Count"],
                        d["Failure Reason"],
                        f"{d.get('Coverage %', 0)}%"
                    ])
                    
            # Generate diagnostic_report.xlsx
            diagnostic_xlsx_path = os.path.join(job_work_dir, "diagnostic_report.xlsx")
            
            wb = openpyxl.Workbook()
            
            # 1. SUMMARY sheet
            ws_summary = wb.active
            ws_summary.title = "SUMMARY"
            ws_summary.append([
                "Platform", 
                "Product Name", 
                "Log Type", 
                "Generated Query", 
                "Search Results Count", 
                "Extracted Logs Count", 
                "Validated Logs Count", 
                "Rejected Logs Count", 
                "Failure Reason",
                "Coverage %"
            ])
            for d in row_diagnostics:
                q_str = "; ".join(d["Generated Query"])
                ws_summary.append([
                    d["Platform"],
                    d["Product Name"],
                    d["Log Type"],
                    q_str,
                    d["Search Results Count"],
                    d["Extracted Logs Count"],
                    d["Validated Logs Count"],
                    d["Rejected Logs Count"],
                    d["Failure Reason"],
                    f"{d.get('Coverage %', 0)}%"
                ])
                
            # 2. UNPROCESSED_SOURCES sheet
            ws_unprocessed = wb.create_sheet(title="UNPROCESSED_SOURCES")
            ws_unprocessed.append([
                "Platform",
                "Product",
                "Log Type",
                "Search Query Used",
                "URL",
                "Title",
                "Source Type",
                "Source Rank",
                "Status",
                "Failure Stage",
                "Reason"
            ])
            
            # 3. SOURCE_AUDIT sheet
            ws_audit = wb.create_sheet(title="SOURCE_AUDIT")
            ws_audit.append([
                "Platform",
                "Product",
                "Log Type",
                "Search Query Used",
                "URL",
                "Search Position",
                "Source Type",
                "Search Rank",
                "Crawled",
                "Logs Extracted",
                "Logs Validated",
                "Logs Rejected",
                "Final Status"
            ])
            
            for info in all_urls_info:
                # Add to SOURCE_AUDIT
                ws_audit.append([
                    info.get("Platform", ""),
                    info.get("Product", ""),
                    info.get("Log Type", ""),
                    info.get("Search Query Used", ""),
                    info.get("URL", ""),
                    info.get("Search Position", ""),
                    info.get("Source Type", ""),
                    info.get("Search Rank", ""),
                    info.get("Crawled", ""),
                    info.get("Logs Extracted", 0),
                    info.get("Logs Validated", 0),
                    info.get("Logs Rejected", 0),
                    info.get("Status", "")
                ])
                
                # Add to UNPROCESSED_SOURCES if validation count == 0
                if info.get("Logs Validated", 0) == 0:
                    ws_unprocessed.append([
                        info.get("Platform", ""),
                        info.get("Product", ""),
                        info.get("Log Type", ""),
                        info.get("Search Query Used", ""),
                        info.get("URL", ""),
                        info.get("Title", ""),
                        info.get("Source Type", ""),
                        info.get("Source Rank", ""),
                        info.get("Status", ""),
                        info.get("Failure Stage", ""),
                        info.get("Reason", "")
                    ])
                    
            # 4. SEARCH_DIAGNOSTICS sheet
            ws_search_diag = wb.create_sheet(title="SEARCH_DIAGNOSTICS")
            ws_search_diag.append([
                "Platform",
                "Product",
                "Log Type",
                "Query",
                "Provider",
                "HTTP Status",
                "Results Parsed",
                "URLs Returned",
                "URLs Crawled",
                "URLs Rejected",
                "Duration",
                "Failure Reason"
            ])
            for sd in all_search_diagnostics:
                ws_search_diag.append([
                    sd.get("Platform", ""),
                    sd.get("Product", ""),
                    sd.get("Log Type", ""),
                    sd.get("Query", ""),
                    sd.get("Provider", ""),
                    sd.get("HTTP Status", 0),
                    sd.get("Results Parsed", 0),
                    sd.get("URLs Returned", 0),
                    sd.get("URLs Crawled", 0),
                    sd.get("URLs Rejected", 0),
                    sd.get("Duration", 0.0),
                    sd.get("Failure Reason", "")
                ])
                
            # 5. REPOSITORY_STATS sheet
            ws_repo_stats = wb.create_sheet(title="REPOSITORY_STATS")
            ws_repo_stats.append([
                "Platform",
                "Product",
                "Log Type",
                "Total Logs",
                "Unique Sources",
                "Last Discovery Date"
            ])
            repo_stats = db_manager.get_repository_stats_for_sheet()
            for r in repo_stats:
                ws_repo_stats.append([
                    r.get("platform", ""),
                    r.get("product_name", ""),
                    r.get("log_type", ""),
                    r.get("total_logs", 0),
                    r.get("unique_sources", 0),
                    r.get("last_discovery_date", "")
                ])
                
            # 6. REPOSITORY_HEALTH sheet
            ws_repo_health = wb.create_sheet(title="REPOSITORY_HEALTH")
            ws_repo_health.append([
                "Total Validated Logs",
                "Total Duplicates Skipped",
                "Unique Sources",
                "Unique Platforms",
                "Unique Products",
                "Last Repository Update",
                "Repository Size (MB)",
                "Total Technologies",
                "Total Categories",
                "Total Vendors"
            ])
            health_data = db_manager.get_repository_health_data()
            ws_repo_health.append([
                health_data["total_logs"],
                health_data["duplicates_skipped"],
                health_data["unique_sources"],
                health_data["unique_platforms"],
                health_data["unique_products"],
                health_data["last_update"],
                health_data["size_mb"],
                health_data.get("total_technologies", 0),
                health_data.get("total_categories", 0),
                health_data.get("total_vendors", 0)
            ])
            
            # 7. PLATFORM_COVERAGE sheet
            ws_platform_coverage = wb.create_sheet(title="PLATFORM_COVERAGE")
            ws_platform_coverage.append([
                "Platform",
                "Search Results",
                "URLs Crawled",
                "Extracted Logs",
                "Validated Logs",
                "Validation Rate %",
                "Repository Count",
                "Coverage Status"
            ])
            
            job_plat_stats = {}
            for d in row_diagnostics:
                plat = d["Platform"].strip() if d.get("Platform") else "Unknown"
                plat_key = plat.lower()
                if plat_key not in job_plat_stats:
                    job_plat_stats[plat_key] = {
                        "name": plat,
                        "search_results": 0,
                        "urls_crawled": 0,
                        "extracted": 0,
                        "validated": 0
                    }
                stat = job_plat_stats[plat_key]
                stat["search_results"] += d.get("Search Results Count", 0)
                stat["extracted"] += d.get("Extracted Logs Count", 0)
                stat["validated"] += d.get("Validated Logs Count", 0)
                
            for info in all_urls_info:
                plat = info.get("Platform", "").strip() if info.get("Platform") else "Unknown"
                plat_key = plat.lower()
                if plat_key in job_plat_stats and info.get("Crawled") == "Yes":
                    job_plat_stats[plat_key]["urls_crawled"] += 1
                    
            for key, stat in job_plat_stats.items():
                # Query repository count
                repo_count = db_manager.get_validated_logs_count_for_platform(key)
                
                # Calculate Validation Rate %
                val_rate = (stat["validated"] / stat["extracted"] * 100.0) if stat["extracted"] > 0 else 0.0
                
                # Determine Coverage Status
                if stat["validated"] >= 10:
                    status = "GOOD"
                elif stat["validated"] > 0:
                    status = "PARTIAL"
                elif stat["extracted"] > 0 and stat["validated"] == 0:
                    status = "WEAK"
                else:
                    status = "NO_COVERAGE"
                    
                ws_platform_coverage.append([
                    stat["name"],
                    stat["search_results"],
                    stat["urls_crawled"],
                    stat["extracted"],
                    stat["validated"],
                    f"{val_rate:.2f}%",
                    repo_count,
                    status
                ])
                
            # 8. TECHNOLOGY_CATALOG sheet
            ws_tech_catalog = wb.create_sheet(title="TECHNOLOGY_CATALOG")
            ws_tech_catalog.append([
                "Technology",
                "Category",
                "Vendor",
                "Discovery Source",
                "First Seen",
                "Last Seen"
            ])
            
            tech_rows = db_manager.get_accepted_technologies_catalog()
            for row in tech_rows:
                ws_tech_catalog.append([
                    row[0],
                    row[1],
                    row[2],
                    row[3],
                    row[4],
                    row[5]
                ])
                
            # 9. TECHNOLOGY_AUDIT sheet
            ws_tech_audit = wb.create_sheet(title="TECHNOLOGY_AUDIT")
            ws_tech_audit.append([
                "Technology",
                "Classification",
                "Confidence",
                "Discovery Source",
                "Accepted",
                "Rejection Reason"
            ])
            
            audit_rows = db_manager.get_technology_audit_catalog()
            for row in audit_rows:
                acc_val = row[4]
                acc_str = "Yes" if (acc_val == 1 or acc_val is True or str(acc_val).lower() == 'true') else "No"
                ws_tech_audit.append([
                    row[0],
                    row[1],
                    row[2],
                    row[3],
                    acc_str,
                    row[5] or ""
                ])
                
            # Recalculate technology coverage & intelligence
            readiness_data = db_manager.recalculate_technology_coverage()
            
            # 10. TECHNOLOGY_COVERAGE sheet
            ws_tech_cov = wb.create_sheet(title="TECHNOLOGY_COVERAGE")
            ws_tech_cov.append([
                "Technology", "Category", "Vendor", "Search Results", "URLs Crawled",
                "Logs Extracted", "Logs Validated", "Repository Logs", "Unique Sources",
                "Coverage Score", "Status", "Last Discovery Date"
            ])
            
            cov_rows = db_manager.get_technology_coverage_details()
            for r in cov_rows:
                rep_logs = r["repository_logs"]
                uniq_srcs = r["unique_sources"]
                ed_score = r["error_diversity_score"]
                cov_score = min(100.0, float(rep_logs * 1.0 + uniq_srcs * 5.0))
                
                if rep_logs == 0:
                    status = "NO_LOGS"
                elif rep_logs < 10:
                    status = "WEAK"
                elif rep_logs >= 50 and uniq_srcs >= 5 and ed_score >= 4:
                    status = "EXCELLENT"
                elif rep_logs >= 20 and uniq_srcs >= 3 and ed_score >= 2:
                    status = "GOOD"
                else:
                    status = "PARTIAL"
                    
                ws_tech_cov.append([
                    r["technology_name"], r["category"], r["vendor"], r["search_results"], r["urls_crawled"],
                    r["logs_extracted"], r["logs_validated"], rep_logs, uniq_srcs, cov_score, status, r["last_discovery_date"]
                ])
                
            # 11. TECHNOLOGY_INTELLIGENCE sheet
            ws_tech_intel = wb.create_sheet(title="TECHNOLOGY_INTELLIGENCE")
            ws_tech_intel.append([
                "Technology", "Total Logs", "Unique Sources", 
                "Unique Error Codes", "Unique Event Types", "Unique Components", "Error Diversity Score"
            ])
            
            intel_rows = db_manager.get_technology_log_profiles()
            for r in intel_rows:
                ws_tech_intel.append([
                    r[0], r[1], r[2], r[3], r[4], r[5], r[6]
                ])
                
            # 12. AUTONOMOUS_READINESS sheet
            ws_ready = wb.create_sheet(title="AUTONOMOUS_READINESS")
            ws_ready.append(["Metric", "Value"])
            ws_ready.append(["Technologies Cataloged", readiness_data["technologies_cataloged"]])
            ws_ready.append(["Technologies Producing Logs", readiness_data["technologies_producing_logs"]])
            ws_ready.append(["Technologies Without Logs", readiness_data["technologies_without_logs"]])
            ws_ready.append(["Average Validation Rate", f"{readiness_data['average_validation_rate']:.2f}%"])
            ws_ready.append(["Average Sources Per Technology", f"{readiness_data['average_sources_per_technology']:.2f}"])
            ws_ready.append(["Repository Growth Rate", f"{readiness_data['repository_growth_rate']:.4f}%"])
            
            wb.save(diagnostic_xlsx_path)

            
            # Generate diagnostic_report.txt
            diagnostic_txt_path = os.path.join(job_work_dir, "diagnostic_report.txt")
            with open(diagnostic_txt_path, "w", encoding="utf-8") as f:
                f.write("=== LOG COLLECTION DIAGNOSTIC REPORT ===\n")
                f.write(f"Job ID: {job_id}\n")
                f.write(f"Time: {datetime.utcnow().isoformat()}\n\n")
                
                for idx, d in enumerate(row_diagnostics, 1):
                    f.write(f"======================================================================\n")
                    f.write(f"ROW {idx}: Platform={d['Platform']} | Product={d['Product Name']} | Log Type={d['Log Type']}\n")
                    f.write(f"======================================================================\n")
                    f.write(f"Platform: {d['Platform']}\n")
                    f.write(f"Product Name: {d['Product Name']}\n")
                    f.write(f"Log Type: {d['Log Type']}\n")
                    f.write(f"Generated Query:\n")
                    if d["Generated Query"]:
                        for q in d["Generated Query"]:
                            f.write(f"  * {q}\n")
                    else:
                        f.write(f"  (None)\n")
                    f.write(f"Search Results Found: {d['Search Results Count']}\n")
                    f.write(f"Candidate Logs Extracted: {d['Extracted Logs Count']}\n")
                    f.write(f"Validated Logs: {d['Validated Logs Count']}\n")
                    f.write(f"Rejected Logs: {d['Rejected Logs Count']}\n")
                    
                    f.write(f"\nRejection Reasons:\n")
                    if d["Rejection Reasons"]:
                        for r in d["Rejection Reasons"]:
                            f.write(f"  * {r}\n")
                    else:
                        f.write(f"  (None)\n")
                        
                    f.write(f"\nSample Rejected Logs:\n")
                    if d["Sample Rejected Logs"]:
                        for s_idx, s in enumerate(d["Sample Rejected Logs"], 1):
                            log_line = s["log"].replace("\n", " ")
                            f.write(f"  {s_idx}. Log: {log_line}\n")
                            f.write(f"     Reason: {s['reason']}\n")
                    else:
                        f.write(f"  (None)\n")
                        
                    f.write(f"\nSearch Result Ranking:\n")
                    if d["Search Result Ranking"]:
                        for r in d["Search Result Ranking"]:
                            f.write(f"  Rank {r['rank']} | {r['domain']} | {r['url']}\n")
                    else:
                        f.write(f"  (None)\n")
                        
                    f.write(f"\nTop 10 Source URLs:\n")
                    if d["Top 10 Source URLs"]:
                        for url in d["Top 10 Source URLs"]:
                            f.write(f"  * {url}\n")
                    else:
                        f.write(f"  (None)\n")
                        
                    f.write(f"\nSearch Diagnostics:\n")
                    if d.get("Search Diagnostics"):
                        f.write(f"  Provider | HTTP Status | Results Parsed | Duration (s) | Error\n")
                        f.write(f"  ---------|-------------|----------------|--------------|------\n")
                        for sd in d["Search Diagnostics"]:
                            f.write(f"  {sd['Provider']} | {sd['HTTP Status']} | {sd['Results Parsed']} | {sd['Duration']} | {sd['Failure Reason']}\n")
                    else:
                        f.write(f"  (None)\n")
                        
                    f.write(f"\nValidation Format Detected:\n")
                    f.write(f"  * Syslog: {d['Validation Format Detected'].get('Syslog', 0)}\n")
                    f.write(f"  * CloudWatch: {d['Validation Format Detected'].get('CloudWatch', 0)}\n")
                    f.write(f"  * Apache: {d['Validation Format Detected'].get('Apache', 0)}\n")
                    f.write(f"  * Application: {d['Validation Format Detected'].get('Application', 0)}\n")
                    f.write(f"  * Unknown: {d['Validation Format Detected'].get('Unknown', 0)}\n")
                    
                    if d["Validated Logs Count"] == 0:
                        f.write(f"\nFailure Reason: {d['Failure Reason']}\n")
                    f.write("\n\n")
            
            # Package everything into a ZIP file (simplified format)
            zip_filename = f"job_{job_id}.zip"
            zip_filepath = os.path.join(BASE_OUTPUT_DIR, zip_filename)
            
            with zipfile.ZipFile(zip_filepath, "w", zipfile.ZIP_DEFLATED) as zipf:
                # Add platform files and sources files
                for ro in row_outputs:
                    fn = ro["fn"]
                    zipf.write(os.path.join(job_work_dir, fn), fn)
                    
                    fn_sources = fn[:-4] + "_sources.txt" if fn.endswith(".txt") else fn + "_sources.txt"
                    zipf.write(os.path.join(sources_dir, fn_sources), f"Sources/{fn_sources}")
                    
                # Add diagnostic_report.xlsx
                zipf.write(diagnostic_xlsx_path, "diagnostic_report.xlsx")
                    
            logger.info(f"Batch job {job_id} processing completed. Zip created at {zip_filepath}")
            db_manager.update_job_status(job_id, status="completed", zip_path=zip_filepath)
            
            # Trigger Slack notification
            send_slack_notification(f"Batch Job {job_id} processing completed successfully. Total: {len(rows)} rows processed.")
            
        except Exception as job_err:
            logger.error(f"Global error in batch job {job_id}: {job_err}", exc_info=True)
            db_manager.update_job_status(job_id, status="failed")
            send_slack_notification(f"Batch Job {job_id} failed with error: {job_err}")

def resume_unfinished_jobs():
    """
    Scans the database for any jobs that were in 'pending' or 'processing' states
    and resumes processing them in a background thread.
    """
    unfinished_jobs = db_manager.get_unfinished_jobs()
    if not unfinished_jobs:
        return
    
    logger.info(f"Found {len(unfinished_jobs)} unfinished jobs during recovery. Resuming processing...")
    for job in unfinished_jobs:
        job_id = job["id"]
        # Start a new thread for each job to process it
        t = threading.Thread(target=process_batch_job, args=(job_id,))
        t.daemon = True
        t.start()
