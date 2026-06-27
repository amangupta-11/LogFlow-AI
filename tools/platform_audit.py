import sys
import os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
import os
import glob
import openpyxl
import sqlite3
from urllib.parse import urlparse

def get_db_stats(db_path):
    stats = {}
    if not os.path.exists(db_path):
        return stats
        
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    
    # Get repository count per platform
    try:
        cur.execute("SELECT platform, COUNT(*) FROM validated_logs GROUP BY platform")
        for row in cur.fetchall():
            plat = row[0].strip()
            stats[plat.lower()] = {
                "inserts": row[1],
                "duplicate_skips": 0
            }
    except Exception as e:
        print(f"Error querying validated_logs: {e}")
        
    # Get duplicate skips per platform from discovery history
    try:
        cur.execute("SELECT platform, COUNT(*) FROM repository_discovery_history WHERE status = 'duplicate_skipped' GROUP BY platform")
        for row in cur.fetchall():
            plat = row[0].strip().lower()
            if plat not in stats:
                stats[plat] = {"inserts": 0, "duplicate_skips": 0}
            stats[plat]["duplicate_skips"] = row[1]
    except Exception as e:
        print(f"Error querying repository_discovery_history: {e}")
        
    conn.close()
    return stats

def run_audit():
    # Find all Excel files
    xlsx_files = glob.glob('backend/outputs/**/*.xlsx', recursive=True)
    
    platform_data = {}
    
    def get_plat_entry(plat):
        p_clean = plat.strip()
        # standard mapping
        p_lower = p_clean.lower()
        if p_lower == "nginx":
            name = "Nginx"
        elif p_lower == "nginx/docker":
            name = "Nginx/Docker"
        elif p_lower == "ubuntu":
            name = "Ubuntu"
        elif p_lower == "apache":
            name = "Apache"
        elif p_lower == "aws":
            name = "AWS"
        elif p_lower == "docker":
            name = "Docker"
        else:
            name = p_clean
            
        key = name.lower()
        if key not in platform_data:
            platform_data[key] = {
                "name": name,
                "search_results": 0,
                "urls_crawled": 0,
                "logs_extracted": 0,
                "logs_validated": 0
            }
        return platform_data[key]

    for f in xlsx_files:
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
            
            # 1. Read SUMMARY sheet
            if "SUMMARY" in wb.sheetnames:
                ws = wb["SUMMARY"]
                rows = list(ws.iter_rows(values_only=True))
                if rows and len(rows) > 1:
                    for r in rows[1:]:
                        if not r or len(r) < 8:
                            continue
                        plat = r[0]
                        if not plat:
                            continue
                        entry = get_plat_entry(plat)
                        
                        try:
                            entry["search_results"] += int(r[4]) if r[4] is not None else 0
                        except: pass
                        try:
                            entry["logs_extracted"] += int(r[5]) if r[5] is not None else 0
                        except: pass
                        try:
                            entry["logs_validated"] += int(r[6]) if r[6] is not None else 0
                        except: pass
                        
            # 2. Read SOURCE_AUDIT sheet
            if "SOURCE_AUDIT" in wb.sheetnames:
                ws = wb["SOURCE_AUDIT"]
                rows = list(ws.iter_rows(values_only=True))
                if rows and len(rows) > 1:
                    for r in rows[1:]:
                        if not r or len(r) < 9:
                            continue
                        plat = r[0]
                        if not plat:
                            continue
                        entry = get_plat_entry(plat)
                        
                        crawled = r[8]
                        if crawled == "Yes":
                            entry["urls_crawled"] += 1
                            
            wb.close()
        except Exception as e:
            pass

    # Read DB stats
    db_stats = get_db_stats("backend/validated_logs.db")
    
    # Print the table header
    print(f"{'Platform':<15} | {'Search Results':<14} | {'URLs Crawled':<12} | {'Extracted Logs':<14} | {'Validated Logs':<14} | {'Val Rate %':<10} | {'Repo Inserts':<12} | {'Dup Skips':<9}")
    print("-" * 115)
    
    weak_val_platforms = []
    no_extract_platforms = []
    high_reject_platforms = []
    low_validated_platforms = []
    
    for key, data in sorted(platform_data.items(), key=lambda x: x[1]["logs_validated"], reverse=True):
        name = data["name"]
        sr = data["search_results"]
        uc = data["urls_crawled"]
        le = data["logs_extracted"]
        lv = data["logs_validated"]
        
        val_rate = (lv / le * 100.0) if le > 0 else 0.0
        
        db_info = db_stats.get(key, {"inserts": 0, "duplicate_skips": 0})
        inserts = db_info["inserts"]
        dups = db_info["duplicate_skips"]
        
        print(f"{name:<15} | {sr:<14} | {uc:<12} | {le:<14} | {lv:<14} | {val_rate:<9.2f}% | {inserts:<12} | {dups:<9}")
        
        # 1. Platforms where discovery is working but validation is weak (e.g. discovery finds results, validation rate is low)
        if sr > 0 and val_rate < 20.0:
            weak_val_platforms.append(name)
            
        # 2. Platforms where URLs are found but no logs are extracted
        if sr > 0 and le == 0:
            no_extract_platforms.append(name)
            
        # 3. Platforms where extraction works but validation rejects most logs (Extracted > 0, Val Rate < 10%)
        if le > 0 and val_rate < 10.0:
            high_reject_platforms.append(name)
            
        # 4. Platforms with less than 10 validated logs
        if lv < 10:
            low_validated_platforms.append(name)
            
    print("\n" + "="*50)
    print("AUDIT FINDINGS IDENTIFICATION")
    print("="*50)
    print(f"1. Discovery working but validation is weak: {', '.join(weak_val_platforms) if weak_val_platforms else 'None'}")
    print(f"2. URLs found but no logs extracted:         {', '.join(no_extract_platforms) if no_extract_platforms else 'None'}")
    print(f"3. Extraction works but validation rejects:  {', '.join(high_reject_platforms) if high_reject_platforms else 'None'}")
    print(f"4. Less than 10 validated logs in batch run:  {', '.join(low_validated_platforms) if low_validated_platforms else 'None'}")

if __name__ == '__main__':
    run_audit()
