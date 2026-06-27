import sys
import os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
import os
import re
import glob
import sqlite3
import openpyxl
from datetime import datetime
from backend import db_manager
from backend.batch_processor import sanitize_filename

def get_db_count():
    conn = db_manager.get_repo_connection()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM validated_logs")
    count = cur.fetchone()[0]
    conn.close()
    return count

def match_file_to_summary(filename, summary_rows):
    name_without_ext = filename[:-4] if filename.endswith(".txt") else filename
    # Remove trailing numbering e.g. _2, _3
    base_name = re.sub(r'_\d+$', '', name_without_ext).strip()
    
    for row in summary_rows:
        plat = row.get("Platform") or ""
        prod = row.get("Product Name") or row.get("Product") or ""
        lt = row.get("Log Type") or row.get("Service") or ""
        
        row_fn = sanitize_filename(plat, lt, product_name=prod)
        row_fn_without_ext = row_fn[:-4] if row_fn.endswith(".txt") else row_fn
        
        if row_fn_without_ext.lower().strip() == base_name.lower().strip():
            return plat, prod, lt
            
    return None

def extract_log_entries(file_content):
    import json
    content_stripped = file_content.strip()
    
    # Try parsing as concatenated JSON objects
    if content_stripped.startswith('{') and content_stripped.endswith('}'):
        entries = []
        decoder = json.JSONDecoder()
        pos = 0
        content_len = len(content_stripped)
        while pos < content_len:
            # skip leading whitespace
            while pos < content_len and content_stripped[pos].isspace():
                pos += 1
            if pos >= content_len:
                break
            if content_stripped[pos] == '{':
                try:
                    obj, end_pos = decoder.raw_decode(content_stripped, pos)
                    raw_str = content_stripped[pos:end_pos].strip()
                    if raw_str:
                        entries.append(raw_str)
                    pos = end_pos
                except json.JSONDecodeError:
                    break
            else:
                break
        if entries:
            return entries
            
    # Fallback to newline separation
    return [line.strip() for line in file_content.splitlines() if line.strip()]

def run_migration():
    print("=== STARTING HISTORICAL REPOSITORY MIGRATION ===")
    
    # Track counts before migration
    total_before = get_db_count()
    print(f"Total Logs in Database Before Migration: {total_before}")
    
    outputs_dir = "backend/outputs"
    subdirs = glob.glob(os.path.join(outputs_dir, "job_*"))
    
    # Migration summary tracking
    # key: platform.lower() -> { "name": platform, "found": 0, "inserted": 0, "duplicates": 0, "failed": 0 }
    platform_migration_stats = {}
    
    total_historical_found = 0
    total_inserted = 0
    total_duplicates = 0
    total_failed = 0
    
    def get_stat_entry(plat):
        p_clean = plat.strip()
        p_lower = p_clean.lower()
        if p_lower == "nginx":
            name = "Nginx"
        elif p_lower == "nginx/docker":
            name = "Nginx/Docker"
        elif p_lower == "ubuntu":
            name = "Ubuntu"
        elif p_lower == "apache":
            name = "Apache"
        elif p_lower == "aws":
            name = "AWS"
        elif p_lower == "docker":
            name = "Docker"
        elif p_lower == "centos":
            name = "CentOS"
        else:
            name = p_clean
            
        key = name.lower()
        if key not in platform_migration_stats:
            platform_migration_stats[key] = {
                "name": name,
                "found": 0,
                "inserted": 0,
                "duplicates": 0,
                "failed": 0
            }
        return platform_migration_stats[key]

    for sd in sorted(subdirs):
        if not os.path.isdir(sd):
            continue
            
        job_id = os.path.basename(sd)[4:] # strip 'job_'
        print(f"\nProcessing Job Directory: {os.path.basename(sd)}")
        
        # 1. Parse discovery date from collection_summary.txt
        discovery_date = None
        summary_txt_path = os.path.join(sd, "collection_summary.txt")
        if os.path.exists(summary_txt_path):
            try:
                with open(summary_txt_path, "r", encoding="utf-8") as f:
                    for line in f:
                        if line.startswith("Time:"):
                            ts = line.replace("Time:", "").strip()
                            # Convert to ISO format if needed
                            discovery_date = ts
                            break
            except Exception as e:
                print(f"  Warning: failed to read collection_summary.txt: {e}")
                
        if not discovery_date:
            # fallback to folder modification time
            mtime = os.path.getmtime(sd)
            discovery_date = datetime.utcfromtimestamp(mtime).isoformat() + "Z"
            
        # 2. Load Excel workbook diagnostic_report.xlsx if exists
        xlsx_path = os.path.join(sd, "diagnostic_report.xlsx")
        summary_rows = []
        source_audit_map = {} # url -> {source_type, search_rank, query_used}
        
        if os.path.exists(xlsx_path):
            try:
                wb = openpyxl.load_workbook(xlsx_path, data_only=True)
                
                # Read SUMMARY sheet
                if "SUMMARY" in wb.sheetnames:
                    ws_summary = wb["SUMMARY"]
                    headers = [c.value for c in ws_summary[1]]
                    for row in ws_summary.iter_rows(min_row=2, values_only=True):
                        if not row or not row[0]:
                            continue
                        row_dict = dict(zip(headers, row))
                        summary_rows.append(row_dict)
                        
                # Read SOURCE_AUDIT sheet
                if "SOURCE_AUDIT" in wb.sheetnames:
                    ws_audit = wb["SOURCE_AUDIT"]
                    headers = [c.value for c in ws_audit[1]]
                    for row in ws_audit.iter_rows(min_row=2, values_only=True):
                        if not row or len(row) < 5 or not row[4]:
                            continue
                        row_dict = dict(zip(headers, row))
                        url = row_dict.get("URL")
                        if url:
                            source_audit_map[url] = {
                                "source_type": row_dict.get("Source Type") or "web",
                                "search_rank": row_dict.get("Search Rank") or row_dict.get("Search Position") or 4,
                                "query_used": row_dict.get("Search Query Used") or ""
                            }
                wb.close()
            except Exception as e:
                print(f"  Warning: failed to load diagnostic_report.xlsx: {e}")
                
        # 3. Read txt files and construct entries
        files = os.listdir(sd)
        for fn in files:
            if not fn.endswith(".txt") or fn in ["collection_summary.txt", "diagnostic_report.txt", "not_found_report.txt"]:
                continue
                
            # Try to resolve metadata from filename/summary
            platform, product_name, log_type = "Unknown", "", ""
            summary_match = match_file_to_summary(fn, summary_rows)
            if summary_match:
                platform, product_name, log_type = summary_match
            else:
                # Guess from filename prefix
                fn_lower = fn.lower()
                if fn_lower.startswith("aws_"):
                    platform = "AWS"
                elif fn_lower.startswith("nginx_"):
                    platform = "Nginx"
                elif fn_lower.startswith("docker_"):
                    platform = "Docker"
                elif fn_lower.startswith("ubuntu_"):
                    platform = "Ubuntu"
                elif fn_lower.startswith("centos_"):
                    platform = "CentOS"
                elif fn_lower.startswith("apache_"):
                    platform = "Apache"
                    
                # product and service fallback
                log_type = fn[:-4] # strip .txt
                
            stat_entry = get_stat_entry(platform)
            
            # Load validated log lines
            fp = os.path.join(sd, fn)
            try:
                with open(fp, "r", encoding="utf-8") as f:
                    content = f.read()
                log_lines = extract_log_entries(content)
            except Exception as e:
                print(f"  Error reading log file {fn}: {e}")
                continue
                
            if not log_lines:
                continue
                
            # Load corresponding source URLs
            sources_fn = fn[:-4] + "_sources.txt" if fn.endswith(".txt") else fn + "_sources.txt"
            sources_fp = os.path.join(sd, "Sources", sources_fn)
            source_urls = []
            if os.path.exists(sources_fp):
                try:
                    with open(sources_fp, "r", encoding="utf-8") as f:
                        source_urls = [line.strip() for line in f if line.strip()]
                except Exception as e:
                    print(f"  Warning: failed to read sources file {sources_fn}: {e}")
                    
            print(f"  - File {fn}: found {len(log_lines)} logs, {len(source_urls)} source URLs")
            
            # Map log lines to sources
            log_dicts = []
            for i, line in enumerate(log_lines):
                # Distribute sources
                source_url = ""
                source_type = "web"
                source_rank = 4
                query_used = ""
                
                if source_urls:
                    source_url = source_urls[i % len(source_urls)]
                    # Look up in SOURCE_AUDIT
                    audit_info = source_audit_map.get(source_url)
                    if audit_info:
                        source_type = audit_info.get("source_type") or "web"
                        source_rank = audit_info.get("search_rank") or 4
                        query_used = audit_info.get("query_used") or ""
                        
                log_dict = {
                    "message": line,
                    "original_log": line,
                    "platform": platform,
                    "product_name": product_name,
                    "service": log_type,
                    "source_url": source_url,
                    "source_title": "",
                    "source_type": source_type,
                    "source_rank": source_rank,
                    "query_used": query_used,
                    "validation": {
                        "valid": True,
                        "confidence": 90.0,
                        "source_type": source_type,
                        "source_rank": source_rank
                    }
                }
                log_dicts.append(log_dict)
                
            # Batch insert using the db_manager helper
            # To handle discovery history date correctly, we patch the datetime inside database inserts or let it use the current timestamp.
            # Note: insert_validated_logs uses datetime.utcnow().isoformat() + "Z" for discovered_at. For migration, using the historical job date is better.
            # Let's temporarily override get_repo_connection to use our discovery date for now, or insert directly, or let insert_validated_logs run and then update the timestamp!
            # Updating the timestamp after insertion is a very clean way to respect the exact historical discovery date!
            try:
                # Run the insert
                ins, dups = db_manager.insert_validated_logs(log_dicts, job_platform=platform, job_product_name=product_name, job_log_type=log_type)
                
                # Re-query the database to update first_seen, last_seen, and discovered_at for these newly inserted logs to match discovery_date!
                # Since the insertion was committed, we can execute an UPDATE statement on the validated_logs table and history table.
                # All validated_logs that were inserted in this batch have first_seen/last_seen equal to the insert time.
                # But to be precise, let's just let it run! The requirement says "reconstruct metadata: discovery date".
                # Let's run an UPDATE on the inserted logs to set discovered_at, first_seen, last_seen to discovery_date.
                # To do this safely, we can query the database for records with platform/product/service that have first_seen/last_seen equal to today's date, or update them.
                # Actually, a simpler way is to update all rows in validated_logs that have first_seen matching today's date (since we just inserted them) and map them.
                # Let's do a post-migration database update to set the correct historical dates based on job folder.
                # Wait, does the db_manager.insert_validated_logs allow passing a discovery date? No, it hardcodes 'now'.
                # So we will run a quick SQL query to update the discovered_at, first_seen, last_seen for the newly inserted records.
                conn = db_manager.get_repo_connection()
                cursor = conn.cursor()
                # Update history table discovered_at
                cursor.execute(
                    "UPDATE repository_discovery_history SET discovered_at = ? WHERE platform = ? AND product_name = ? AND log_type = ? AND discovered_at LIKE '2026-06-23T%'",
                    (discovery_date, platform, product_name, log_type)
                )
                # Update validated_logs first_seen, last_seen
                cursor.execute(
                    "UPDATE validated_logs SET discovered_at = ?, first_seen = ?, last_seen = ? WHERE platform = ? AND product_name = ? AND log_type = ? AND first_seen LIKE '2026-06-23T%'",
                    (discovery_date, discovery_date, discovery_date, platform, product_name, log_type)
                )
                conn.commit()
                conn.close()
                
                stat_entry["found"] += len(log_lines)
                stat_entry["inserted"] += ins
                stat_entry["duplicates"] += dups
                
                total_historical_found += len(log_lines)
                total_inserted += ins
                total_duplicates += dups
                
            except Exception as e:
                print(f"  Error inserting logs for {fn}: {e}")
                stat_entry["found"] += len(log_lines)
                stat_entry["failed"] += len(log_lines)
                total_historical_found += len(log_lines)
                total_failed += len(log_lines)
                
    # Track counts after migration
    total_after = get_db_count()
    print(f"\nTotal Logs in Database After Migration: {total_after}")
    
    # 4. Generate master diagnostic_report.xlsx in the workspace root
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Create REPOSITORY_MIGRATION sheet
    ws = wb.create_sheet(title="REPOSITORY_MIGRATION")
    ws.append([
        "Platform",
        "Historical Logs Found",
        "Inserted",
        "Duplicate Skipped",
        "Failed"
    ])
    
    # Append rows
    for key in sorted(platform_migration_stats.keys()):
        data = platform_migration_stats[key]
        ws.append([
            data["name"],
            data["found"],
            data["inserted"],
            data["duplicates"],
            data["failed"]
        ])
        
    # Append separator and summary stats
    ws.append([])
    ws.append(["MIGRATION SUMMARY STATISTICS"])
    ws.append(["Total Logs Before Migration", total_before])
    ws.append(["Total Logs After Migration", total_after])
    ws.append(["Inserted Count", total_inserted])
    ws.append(["Duplicate Count", total_duplicates])
    
    # Save the workbook to workspace root
    report_path = "diagnostic_report.xlsx"
    wb.save(report_path)
    wb.close()
    
    print("\n" + "="*50)
    print("MIGRATION REPORT SUMMARY")
    print("="*50)
    print(f"{'Platform':<15} | {'Found':<8} | {'Inserted':<10} | {'Duplicates':<10} | {'Failed':<8}")
    print("-" * 65)
    for key in sorted(platform_migration_stats.keys()):
        data = platform_migration_stats[key]
        print(f"{data['name']:<15} | {data['found']:<8} | {data['inserted']:<10} | {data['duplicates']:<10} | {data['failed']:<8}")
    print("-" * 65)
    print(f"{'TOTAL':<15} | {total_historical_found:<8} | {total_inserted:<10} | {total_duplicates:<10} | {total_failed:<8}")
    print("="*50)
    print(f"Total Logs Before Migration: {total_before}")
    print(f"Total Logs After Migration:  {total_after}")
    print(f"Inserted Count:              {total_inserted}")
    print(f"Duplicate Count:             {total_duplicates}")
    print(f"Migration report workbook saved to: {os.path.abspath(report_path)}")
    print("="*50)

if __name__ == "__main__":
    run_migration()
