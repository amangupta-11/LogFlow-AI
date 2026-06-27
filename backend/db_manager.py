import os
import sqlite3
import json
from datetime import datetime

DATABASE_URL = os.getenv("DATABASE_URL")

_memory_conn = None

class MemoryConnectionWrapper:
    def __init__(self, conn):
        self._conn = conn
    def cursor(self, *args, **kwargs):
        return self._conn.cursor(*args, **kwargs)
    def commit(self):
        self._conn.commit()
    def rollback(self):
        self._conn.rollback()
    def close(self):
        # Keep open to prevent losing memory database
        pass

class PostgresRow(dict):
    def __init__(self, description, values):
        self._values = values
        d = dict(zip([desc[0] for desc in description], values))
        super().__init__(d)
    def __getitem__(self, key):
        if isinstance(key, int):
            return self._values[key]
        return super().__getitem__(key)

def translate_query(query):
    if not query:
        return query
    
    # Replace placeholder ? with %s
    query = query.replace("?", "%s")
    
    # Translate INSERT OR IGNORE
    if "INSERT OR IGNORE" in query.upper():
        q_upper = query.upper()
        if "AGENT_STATUS" in q_upper:
            query = query.replace("INSERT OR IGNORE", "INSERT").replace("insert or ignore", "insert") + " ON CONFLICT (id) DO NOTHING"
        elif "TECHNOLOGY_ALIASES" in q_upper:
            query = query.replace("INSERT OR IGNORE", "INSERT").replace("insert or ignore", "insert") + " ON CONFLICT (alias) DO NOTHING"
        elif "REPOSITORY_METRICS" in q_upper:
            query = query.replace("INSERT OR IGNORE", "INSERT").replace("insert or ignore", "insert") + " ON CONFLICT (key) DO NOTHING"
            
    # Translate INSERT OR REPLACE
    if "INSERT OR REPLACE" in query.upper():
        q_upper = query.upper()
        if "DOMAIN_PERFORMANCE" in q_upper:
            query = """
                INSERT INTO domain_performance (
                    domain, urls_crawled, logs_extracted, logs_validated, logs_inserted, yield_score
                ) VALUES (%s, %s, %s, %s, %s, %s)
                ON CONFLICT (domain) DO UPDATE SET
                    urls_crawled = EXCLUDED.urls_crawled,
                    logs_extracted = EXCLUDED.logs_extracted,
                    logs_validated = EXCLUDED.logs_validated,
                    logs_inserted = EXCLUDED.logs_inserted,
                    yield_score = EXCLUDED.yield_score
            """
            
    # Translate datetime('now')
    query = query.replace("datetime('now')", "CURRENT_TIMESTAMP")
    query = query.replace("datetime('now', 'localtime')", "CURRENT_TIMESTAMP")
    
    return query

class PostgresCursorWrapper:
    def __init__(self, cursor):
        self._cursor = cursor
    def execute(self, query, params=()):
        cleaned_query = translate_query(query)
        self._cursor.execute(cleaned_query, params)
    def fetchone(self):
        row = self._cursor.fetchone()
        if row is None:
            return None
        if isinstance(row, dict):
            return row
        return PostgresRow(self._cursor.description, row)
    def fetchall(self):
        rows = self._cursor.fetchall()
        if not rows:
            return []
        if isinstance(rows[0], dict):
            return rows
        desc = self._cursor.description
        return [PostgresRow(desc, r) for r in rows]
    def close(self):
        self._cursor.close()
    def __iter__(self):
        while True:
            row = self.fetchone()
            if row is None:
                break
            yield row
    def __getattr__(self, name):
        return getattr(self._cursor, name)

class PostgresConnectionWrapper:
    def __init__(self, conn):
        self._conn = conn
    def cursor(self, *args, **kwargs):
        cursor = self._conn.cursor(*args, **kwargs)
        return PostgresCursorWrapper(cursor)
    def commit(self):
        self._conn.commit()
    def rollback(self):
        self._conn.rollback()
    def close(self):
        self._conn.close()
    def execute(self, query, params=()):
        cleaned_query = translate_query(query)
        cursor = self._conn.cursor()
        cursor.execute(cleaned_query, params)
        return PostgresCursorWrapper(cursor)
    def __getattr__(self, name):
        return getattr(self._conn, name)

def get_connection():
    global _memory_conn
    if DATABASE_URL and (DATABASE_URL.startswith("postgres://") or DATABASE_URL.startswith("postgresql://")):
        try:
            import psycopg2
            url = DATABASE_URL
            if url.startswith("postgres://"):
                url = url.replace("postgres://", "postgresql://", 1)
            conn = psycopg2.connect(url)
            return PostgresConnectionWrapper(conn), True
        except ImportError:
            print("psycopg2 not found, falling back to SQLite.")
    
    # SQLite path
    db_path = os.getenv("SQLITE_DB_PATH", "jobs.db")
    if os.getenv("VERCEL"):
        db_path = "/tmp/jobs.db"

    if db_path == ":memory:":
        if _memory_conn is None:
            _memory_conn = sqlite3.connect(":memory:")
            _memory_conn.row_factory = sqlite3.Row
        return MemoryConnectionWrapper(_memory_conn), False

    if not os.path.isabs(db_path):
        base_dir = os.path.dirname(os.path.abspath(__file__))
        db_path = os.path.join(base_dir, db_path)
    
    conn = sqlite3.connect(db_path, timeout=30.0, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    try:
        conn.execute("PRAGMA journal_mode=WAL;")
        conn.execute("PRAGMA synchronous=NORMAL;")
    except Exception as e:
        print(f"Could not configure SQLite PRAGMAs: {e}")
    return conn, False


def init_db():
    conn, is_postgres = get_connection()
    try:
        cursor = conn.cursor()
        try:
            if is_postgres:
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS jobs (
                        id VARCHAR(255) PRIMARY KEY,
                        status VARCHAR(50),
                        total_rows INTEGER,
                        completed_rows INTEGER,
                        failed_rows INTEGER,
                        skipped_rows INTEGER,
                        remaining_rows INTEGER,
                        zip_path TEXT,
                        created_at VARCHAR(100),
                        updated_at VARCHAR(100)
                    )
                """)
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS job_rows (
                        id SERIAL PRIMARY KEY,
                        job_id VARCHAR(255),
                        platform VARCHAR(255),
                        product_name VARCHAR(255),
                        version VARCHAR(100),
                        service VARCHAR(255),
                        max_logs INTEGER,
                        status VARCHAR(50),
                        error_message TEXT,
                        validated_count INTEGER,
                        non_validated_count INTEGER,
                        sources_found INTEGER,
                        error_code VARCHAR(255),
                        excel_error_message TEXT,
                        reason TEXT,
                        source TEXT,
                        error_message_long TEXT,
                        category VARCHAR(255),
                        created_at VARCHAR(100),
                        updated_at VARCHAR(100)
                    )
                """)
            else:
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS jobs (
                        id TEXT PRIMARY KEY,
                        status TEXT,
                        total_rows INTEGER,
                        completed_rows INTEGER,
                        failed_rows INTEGER,
                        skipped_rows INTEGER,
                        remaining_rows INTEGER,
                        zip_path TEXT,
                        created_at TEXT,
                        updated_at TEXT
                    )
                """)
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS job_rows (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        job_id TEXT,
                        platform TEXT,
                        product_name TEXT,
                        version TEXT,
                        service TEXT,
                        max_logs INTEGER,
                        status TEXT,
                        error_message TEXT,
                        validated_count INTEGER,
                        non_validated_count INTEGER,
                        sources_found INTEGER,
                        error_code TEXT,
                        excel_error_message TEXT,
                        reason TEXT,
                        source TEXT,
                        error_message_long TEXT,
                        category TEXT,
                        created_at TEXT,
                        updated_at TEXT
                    )
                """)
            conn.commit()
        finally:
            cursor.close()
    finally:
        conn.close()
    
    migrate_db()

def migrate_db():
    conn, is_postgres = get_connection()
    try:
        cursor = conn.cursor()
        try:
            # Migrate job_rows
            new_cols = {
                "product_name": "VARCHAR(255)" if is_postgres else "TEXT",
                "error_code": "VARCHAR(255)" if is_postgres else "TEXT",
                "excel_error_message": "TEXT",
                "reason": "TEXT",
                "source": "TEXT",
                "error_message_long": "TEXT",
                "category": "VARCHAR(255)" if is_postgres else "TEXT"
            }
            
            if is_postgres:
                for col, col_type in new_cols.items():
                    cursor.execute(f"SELECT column_name FROM information_schema.columns WHERE table_name='job_rows' AND column_name='{col}'")
                    if not cursor.fetchone():
                        cursor.execute(f"ALTER TABLE job_rows ADD COLUMN {col} {col_type}")
            else:
                cursor.execute("PRAGMA table_info(job_rows)")
                columns = [row[1] for row in cursor.fetchall()]
                for col, col_type in new_cols.items():
                    if col not in columns:
                        cursor.execute(f"ALTER TABLE job_rows ADD COLUMN {col} {col_type}")
            conn.commit()
        finally:
            cursor.close()
    finally:
        conn.close()


def execute_write(query, params=()):
    conn, is_postgres = get_connection()
    try:
        cursor = conn.cursor()
        try:
            # In PostgreSQL, we replace ? placeholders with %s
            if is_postgres:
                query = query.replace("?", "%s")
            cursor.execute(query, params)
            conn.commit()
        finally:
            cursor.close()
    finally:
        conn.close()

def execute_read(query, params=()):
    conn, is_postgres = get_connection()
    try:
        # For PostgreSQL we want to get dict-like results
        if is_postgres:
            from psycopg2.extras import RealDictCursor
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            try:
                query = query.replace("?", "%s")
                cursor.execute(query, params)
                results = [dict(row) for row in cursor.fetchall()]
            finally:
                cursor.close()
        else:
            cursor = conn.cursor()
            try:
                cursor.execute(query, params)
                results = [dict(row) for row in cursor.fetchall()]
            finally:
                cursor.close()
    finally:
        conn.close()
    return results

def execute_read_one(query, params=()):
    results = execute_read(query, params)
    return results[0] if results else None

# Helper DB Methods
def create_job(job_id, total_rows):
    now = datetime.utcnow().isoformat()
    execute_write(
        "INSERT INTO jobs (id, status, total_rows, completed_rows, failed_rows, skipped_rows, remaining_rows, zip_path, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (job_id, "pending", total_rows, 0, 0, 0, total_rows, "", now, now)
    )

def add_job_row(job_id, platform, version, service, max_logs, status="pending", error_message="",
                product_name="", error_code="", excel_error_message="", reason="", source="", error_message_long="", category=""):
    now = datetime.utcnow().isoformat()
    execute_write(
        """INSERT INTO job_rows (
            job_id, platform, version, service, max_logs, status, error_message, validated_count, non_validated_count, sources_found,
            product_name, error_code, excel_error_message, reason, source, error_message_long, category, created_at, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (
            job_id, platform, version, service, max_logs, status, error_message, 0, 0, 0,
            product_name, error_code, excel_error_message, reason, source, error_message_long, category, now, now
        )
    )

def get_job(job_id):
    return execute_read_one("SELECT * FROM jobs WHERE id = ?", (job_id,))

def get_job_rows(job_id):
    return execute_read("SELECT * FROM job_rows WHERE job_id = ?", (job_id,))

def update_job_status(job_id, status=None, zip_path=None):
    now = datetime.utcnow().isoformat()
    if status and zip_path is not None:
        execute_write("UPDATE jobs SET status = ?, zip_path = ?, updated_at = ? WHERE id = ?", (status, zip_path, now, job_id))
    elif status:
        execute_write("UPDATE jobs SET status = ?, updated_at = ? WHERE id = ?", (status, now, job_id))
    elif zip_path is not None:
        execute_write("UPDATE jobs SET zip_path = ?, updated_at = ? WHERE id = ?", (zip_path, now, job_id))

def increment_job_stats(job_id, completed=0, failed=0, skipped=0):
    now = datetime.utcnow().isoformat()
    job = get_job(job_id)
    if not job:
        return
    
    new_completed = job["completed_rows"] + completed
    new_failed = job["failed_rows"] + failed
    new_skipped = job["skipped_rows"] + skipped
    new_remaining = max(0, job["total_rows"] - (new_completed + new_failed + new_skipped))
    
    execute_write(
        "UPDATE jobs SET completed_rows = ?, failed_rows = ?, skipped_rows = ?, remaining_rows = ?, updated_at = ? WHERE id = ?",
        (new_completed, new_failed, new_skipped, new_remaining, now, job_id)
    )

def update_row_status(row_id, status, error_message="", validated_count=0, non_validated_count=0, sources_found=0):
    now = datetime.utcnow().isoformat()
    execute_write(
        "UPDATE job_rows SET status = ?, error_message = ?, validated_count = ?, non_validated_count = ?, sources_found = ?, updated_at = ? WHERE id = ?",
        (status, error_message, validated_count, non_validated_count, sources_found, now, row_id)
    )

def get_unfinished_jobs():
    return execute_read("SELECT * FROM jobs WHERE status IN ('pending', 'processing')")

def get_unfinished_rows(job_id):
    return execute_read("SELECT * FROM job_rows WHERE job_id = ? AND status IN ('pending', 'processing')", (job_id,))

# Cache functions removed entirely.


def get_fallback_source(platform, product_name):
    plat_lower = (platform or "").lower().strip()
    prod_lower = (product_name or "").lower().strip()
    
    if "aws" in plat_lower or "amazon" in plat_lower:
        if "s3" in prod_lower or "s3" in plat_lower:
            return "https://docs.aws.amazon.com/s3/", "Amazon Simple Storage Service Documentation"
        elif "ec2" in prod_lower or "ec2" in plat_lower:
            return "https://docs.aws.amazon.com/ec2/", "Amazon Elastic Compute Cloud Documentation"
        elif "rds" in prod_lower or "rds" in plat_lower:
            return "https://docs.aws.amazon.com/rds/", "Amazon Relational Database Service Documentation"
        elif "lambda" in prod_lower or "lambda" in plat_lower:
            return "https://docs.aws.amazon.com/lambda/", "AWS Lambda Documentation"
        elif "dynamodb" in prod_lower or "dynamodb" in plat_lower:
            return "https://docs.aws.amazon.com/amazondynamodb/", "Amazon DynamoDB Documentation"
        return "https://docs.aws.amazon.com/", "Amazon Web Services Documentation"
    elif "azure" in plat_lower:
        if "blob" in prod_lower:
            return "https://learn.microsoft.com/en-us/azure/storage/blobs/", "Azure Blob Storage Documentation"
        elif "vm" in prod_lower or "virtual machine" in prod_lower:
            return "https://learn.microsoft.com/en-us/azure/virtual-machines/", "Azure Virtual Machines Documentation"
        elif "sql" in prod_lower:
            return "https://learn.microsoft.com/en-us/azure/azure-sql/", "Azure SQL Documentation"
        return "https://learn.microsoft.com/en-us/azure/", "Microsoft Azure Documentation"
    elif "google" in plat_lower or "gcp" in plat_lower:
        return "https://cloud.google.com/docs/", "Google Cloud Documentation"
    elif "nginx" in plat_lower:
        return "https://nginx.org/en/docs/", "nginx documentation"
    elif "docker" in plat_lower:
        return "https://docs.docker.com/", "Docker Docs"
    elif "kubernetes" in plat_lower or "k8s" in plat_lower:
        return "https://kubernetes.io/docs/", "Kubernetes Documentation"
    elif "oracle" in plat_lower:
        return "https://docs.oracle.com/", "Oracle Help Center"
    else:
        return "https://docs.google.com/", "General Documentation Fallback"


def get_columns(cursor, table_name, is_postgres):
    if is_postgres:
        cursor.execute("SELECT column_name FROM information_schema.columns WHERE table_name = %s", (table_name.lower(),))
        return [r[0] for r in cursor.fetchall()]
    else:
        cursor.execute(f"PRAGMA table_info({table_name})")
        return [row[1] for row in cursor.fetchall()]

def migrate_repo_db(conn):
    is_postgres = "sqlite3" not in str(type(conn))
    cursor = conn.cursor()
    try:
        _migrate_repo_db_internal(cursor, conn, is_postgres)
    finally:
        cursor.close()

def _migrate_repo_db_internal(cursor, conn, is_postgres):
    
    # 1. Add columns to validated_logs if they don't exist
    repo_migrations = [
        ("source_status", "VARCHAR(50)" if is_postgres else "TEXT"),
        ("fallback_source_url", "TEXT"),
        ("fallback_source_title", "TEXT")
    ]
    columns = get_columns(cursor, "validated_logs", is_postgres)
    
    migrated = False
    for col_name, col_type in repo_migrations:
        if col_name not in columns:
            cursor.execute(f"ALTER TABLE validated_logs ADD COLUMN {col_name} {col_type}")
            migrated = True
            
    if migrated:
        conn.commit()
        
    # 2. Backfill existing records
    cursor.execute("SELECT id, platform, product_name, source_url, source_status FROM validated_logs")
    rows = cursor.fetchall()
    updated = 0
    for r in rows:
        log_id = r[0]
        platform = r[1]
        product_name = r[2]
        source_url = r[3]
        current_status = r[4]
        
        # If source_status is not set yet
        if not current_status:
            if source_url and source_url.strip():
                new_status = 'VERIFIED'
                fallback_url = None
                fallback_title = None
            else:
                new_status = 'MISSING'
                # Keep source_url and source_title unchanged
                fallback_url, fallback_title = get_fallback_source(platform, product_name)
                
            cursor.execute(
                """UPDATE validated_logs 
                   SET source_status = ?, fallback_source_url = ?, fallback_source_title = ? 
                   WHERE id = ?""",
                (new_status, fallback_url, fallback_title, log_id)
            )
            updated += 1
            
    if updated > 0:
        conn.commit()
        
    # 3. Create domain_performance table
    if is_postgres:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS domain_performance (
                domain VARCHAR(255) PRIMARY KEY,
                urls_crawled INTEGER DEFAULT 0,
                logs_extracted INTEGER DEFAULT 0,
                logs_validated INTEGER DEFAULT 0,
                logs_inserted INTEGER DEFAULT 0,
                yield_score REAL DEFAULT 0.0
            )
        """)
    else:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS domain_performance (
                domain TEXT PRIMARY KEY,
                urls_crawled INTEGER DEFAULT 0,
                logs_extracted INTEGER DEFAULT 0,
                logs_validated INTEGER DEFAULT 0,
                logs_inserted INTEGER DEFAULT 0,
                yield_score REAL DEFAULT 0.0
            )
        """)
    conn.commit()

    # 4. Create agent_status table
    if is_postgres:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS agent_status (
                id INTEGER PRIMARY KEY CHECK (id = 1),
                status VARCHAR(50) DEFAULT 'idle',
                current_job VARCHAR(255) DEFAULT NULL,
                current_tech VARCHAR(255) DEFAULT NULL,
                last_active VARCHAR(100)
            )
        """)
    else:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS agent_status (
                id INTEGER PRIMARY KEY CHECK (id = 1),
                status TEXT DEFAULT 'idle',
                current_job TEXT DEFAULT NULL,
                current_tech TEXT DEFAULT NULL,
                last_active TEXT
            )
        """)
    
    if is_postgres:
        cursor.execute("INSERT INTO agent_status (id, status, last_active) VALUES (1, 'idle', CURRENT_TIMESTAMP) ON CONFLICT (id) DO NOTHING")
    else:
        cursor.execute("INSERT OR IGNORE INTO agent_status (id, status, last_active) VALUES (1, 'idle', datetime('now'))")
    conn.commit()
    
    # 5. Create agent_control_queue table
    if is_postgres:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS agent_control_queue (
                id SERIAL PRIMARY KEY,
                command VARCHAR(255),
                status VARCHAR(50) DEFAULT 'pending',
                created_at VARCHAR(100)
            )
        """)
    else:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS agent_control_queue (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                command TEXT,
                status TEXT DEFAULT 'pending',
                created_at TEXT
            )
        """)
    conn.commit()
    
    # 6. Create agent_event_feed table
    if is_postgres:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS agent_event_feed (
                id SERIAL PRIMARY KEY,
                timestamp VARCHAR(100),
                event_type VARCHAR(100),
                message TEXT
            )
        """)
    else:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS agent_event_feed (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp TEXT,
                event_type TEXT,
                message TEXT
            )
        """)
    conn.commit()
    
    # 7. Create agent_runtime_metrics table if not exists
    if is_postgres:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS agent_runtime_metrics (
                id SERIAL PRIMARY KEY,
                timestamp VARCHAR(100),
                technologies_processed INTEGER,
                urls_crawled INTEGER,
                logs_extracted INTEGER,
                logs_validated INTEGER,
                logs_inserted INTEGER,
                failures INTEGER,
                urls_skipped INTEGER DEFAULT 0,
                pages_log_rich INTEGER DEFAULT 0,
                pages_low_value INTEGER DEFAULT 0,
                insert_yield_pct REAL DEFAULT 0.0
            )
        """)
    else:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS agent_runtime_metrics (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp TEXT,
                technologies_processed INTEGER,
                urls_crawled INTEGER,
                logs_extracted INTEGER,
                logs_validated INTEGER,
                logs_inserted INTEGER,
                failures INTEGER,
                urls_skipped INTEGER DEFAULT 0,
                pages_log_rich INTEGER DEFAULT 0,
                pages_low_value INTEGER DEFAULT 0,
                insert_yield_pct REAL DEFAULT 0.0
            )
        """)
    conn.commit()

    # 7b. Migrate agent_runtime_metrics columns
    metrics_migrations = [
        ("urls_skipped", "INTEGER DEFAULT 0"),
        ("pages_log_rich", "INTEGER DEFAULT 0"),
        ("pages_low_value", "INTEGER DEFAULT 0"),
        ("insert_yield_pct", "REAL DEFAULT 0.0")
    ]
    metrics_columns = get_columns(cursor, "agent_runtime_metrics", is_postgres)
    
    metrics_migrated = False
    for col_name, col_type in metrics_migrations:
        if col_name not in metrics_columns:
            cursor.execute(f"ALTER TABLE agent_runtime_metrics ADD COLUMN {col_name} {col_type}")
            metrics_migrated = True
            
    if metrics_migrated:
        conn.commit()

    # 8. Migrate notification_queue for retry history fields
    q_cols = get_columns(cursor, "notification_queue", is_postgres)
    q_migrations = [
        ("retry_count", "INTEGER DEFAULT 0"),
        ("attempt_number", "INTEGER DEFAULT 0"),
        ("last_attempt", "VARCHAR(100)" if is_postgres else "TEXT"),
        ("next_retry", "VARCHAR(100)" if is_postgres else "TEXT"),
        ("delivered_at", "VARCHAR(100)" if is_postgres else "TEXT")
    ]
    q_migrated = False
    for col_name, col_type in q_migrations:
        if col_name not in q_cols:
            cursor.execute(f"ALTER TABLE notification_queue ADD COLUMN {col_name} {col_type}")
            q_migrated = True
    if q_migrated:
        conn.commit()

    # 9. Migrate notification_history for event metrics fields
    h_cols = get_columns(cursor, "notification_history", is_postgres)
    h_migrations = [
        ("severity", "VARCHAR(50)" if is_postgres else "TEXT"),
        ("job_id", "VARCHAR(255)" if is_postgres else "TEXT"),
        ("technology", "VARCHAR(255)" if is_postgres else "TEXT")
    ]
    h_migrated = False
    for col_name, col_type in h_migrations:
        if col_name not in h_cols:
            cursor.execute(f"ALTER TABLE notification_history ADD COLUMN {col_name} {col_type}")
            h_migrated = True
    if h_migrated:
        conn.commit()

    # 10. Migrate agent_status for monitoring fields
    status_cols = get_columns(cursor, "agent_status", is_postgres)
    status_migrations = [
        ("technologies_processed", "INTEGER DEFAULT 0"),
        ("technologies_total", "INTEGER DEFAULT 0"),
        ("current_query", "INTEGER DEFAULT 0"),
        ("total_queries", "INTEGER DEFAULT 0"),
        ("current_url", "INTEGER DEFAULT 0"),
        ("total_urls", "INTEGER DEFAULT 0"),
        ("cycle_start_time", "VARCHAR(100)" if is_postgres else "TEXT"),
        ("cycle_urls_crawled", "INTEGER DEFAULT 0"),
        ("cycle_pages_classified", "INTEGER DEFAULT 0"),
        ("cycle_logs_extracted", "INTEGER DEFAULT 0"),
        ("cycle_logs_validated", "INTEGER DEFAULT 0"),
        ("cycle_logs_inserted", "INTEGER DEFAULT 0"),
        ("cycle_duplicates_skipped", "INTEGER DEFAULT 0"),
        ("current_phase", "VARCHAR(100) DEFAULT 'Idle'" if is_postgres else "TEXT DEFAULT 'Idle'"),
        ("next_technology_discovery", "VARCHAR(100)" if is_postgres else "TEXT"),
        ("next_log_discovery", "VARCHAR(100)" if is_postgres else "TEXT"),
        ("next_health_check", "VARCHAR(100)" if is_postgres else "TEXT"),
        ("next_daily_report", "VARCHAR(100)" if is_postgres else "TEXT")
    ]
    status_migrated = False
    for col_name, col_type in status_migrations:
        if col_name not in status_cols:
            cursor.execute(f"ALTER TABLE agent_status ADD COLUMN {col_name} {col_type}")
            status_migrated = True
    if status_migrated:
        conn.commit()

    # 11. Create agent_health_alert_states table to track health state machine
    if is_postgres:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS agent_health_alert_states (
                job_type VARCHAR(255) PRIMARY KEY,
                last_state VARCHAR(50),
                last_notified_id INTEGER,
                downtime_start VARCHAR(100)
            )
        """)
    else:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS agent_health_alert_states (
                job_type TEXT PRIMARY KEY,
                last_state TEXT,
                last_notified_id INTEGER,
                downtime_start TEXT
            )
        """)
    conn.commit()
    
    # 12. Create agent_runtime_metrics table to track daily runs
    if is_postgres:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS agent_runtime_metrics (
                id SERIAL PRIMARY KEY,
                timestamp VARCHAR(100),
                technologies_processed INTEGER,
                urls_crawled INTEGER,
                logs_extracted INTEGER,
                logs_validated INTEGER,
                logs_inserted INTEGER,
                failures INTEGER,
                urls_skipped INTEGER DEFAULT 0,
                pages_log_rich INTEGER DEFAULT 0,
                pages_low_value INTEGER DEFAULT 0,
                insert_yield_pct REAL DEFAULT 0.0
            )
        """)
    else:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS agent_runtime_metrics (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp TEXT,
                technologies_processed INTEGER,
                urls_crawled INTEGER,
                logs_extracted INTEGER,
                logs_validated INTEGER,
                logs_inserted INTEGER,
                failures INTEGER,
                urls_skipped INTEGER DEFAULT 0,
                pages_log_rich INTEGER DEFAULT 0,
                pages_low_value INTEGER DEFAULT 0,
                insert_yield_pct REAL DEFAULT 0.0
            )
        """)
    conn.commit()



def get_repo_connection():
    if DATABASE_URL and (DATABASE_URL.startswith("postgres://") or DATABASE_URL.startswith("postgresql://")):
        try:
            import psycopg2
            url = DATABASE_URL
            if url.startswith("postgres://"):
                url = url.replace("postgres://", "postgresql://", 1)
            conn = psycopg2.connect(url)
            return PostgresConnectionWrapper(conn)
        except ImportError:
            print("psycopg2 not found, falling back to SQLite for repository.")

    db_path = os.getenv("REPO_DB_PATH", "validated_logs.db")
    if os.getenv("VERCEL"):
        db_path = "/tmp/validated_logs.db"

    if not os.path.isabs(db_path):
        base_dir = os.path.dirname(os.path.abspath(__file__))
        db_path = os.path.join(base_dir, db_path)
    conn = sqlite3.connect(db_path, timeout=30.0, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    try:
        conn.execute("PRAGMA journal_mode=WAL;")
        conn.execute("PRAGMA synchronous=NORMAL;")
    except Exception as e:
        print(f"Could not configure repository SQLite PRAGMAs: {e}")
    return conn


def init_repo_db():
    conn = get_repo_connection()
    try:
        _init_repo_db_internal(conn)
    finally:
        conn.close()

def _init_repo_db_internal(conn):
    is_postgres = "sqlite3" not in str(type(conn))
    cursor = conn.cursor()
    
    try:
        if is_postgres:
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS validated_logs (
                    id SERIAL PRIMARY KEY,
                    platform VARCHAR(255),
                    product_name VARCHAR(255),
                    log_type VARCHAR(255),
                    log_message TEXT,
                    normalized_log TEXT,
                    source_url TEXT,
                    source_title TEXT,
                    source_type VARCHAR(100),
                    source_rank INTEGER,
                    validation_score REAL,
                    discovery_query TEXT,
                    discovered_at VARCHAR(100),
                    raw_hash VARCHAR(64),
                    normalized_hash VARCHAR(64) UNIQUE,
                    platform_category VARCHAR(255),
                    vendor VARCHAR(255),
                    technology_version VARCHAR(100),
                    source_domain VARCHAR(255),
                    log_severity VARCHAR(50),
                    event_type VARCHAR(100),
                    component VARCHAR(255),
                    hostname VARCHAR(255),
                    process_name VARCHAR(255),
                    error_code VARCHAR(100),
                    first_seen VARCHAR(100),
                    last_seen VARCHAR(100),
                    source_status VARCHAR(50),
                    fallback_source_url TEXT,
                    fallback_source_title TEXT
                )
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS repository_metrics (
                    key VARCHAR(255) PRIMARY KEY,
                    value INTEGER
                )
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS repository_discovery_history (
                    id SERIAL PRIMARY KEY,
                    platform VARCHAR(255),
                    product_name VARCHAR(255),
                    log_type VARCHAR(255),
                    source_url TEXT,
                    discovered_at VARCHAR(100),
                    status VARCHAR(50),
                    validation_result TEXT
                )
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS technology_catalog (
                    id SERIAL PRIMARY KEY,
                    technology_name VARCHAR(255) UNIQUE,
                    category VARCHAR(255),
                    vendor VARCHAR(255),
                    discovery_source VARCHAR(255),
                    first_seen VARCHAR(100),
                    last_seen VARCHAR(100),
                    status VARCHAR(50),
                    log_queries TEXT,
                    classification TEXT,
                    technology_confidence REAL,
                    accepted INTEGER,
                    rejection_reason TEXT
                )
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS technology_coverage (
                    technology_name VARCHAR(255) PRIMARY KEY,
                    category VARCHAR(255),
                    vendor VARCHAR(255),
                    search_results INTEGER,
                    urls_crawled INTEGER,
                    logs_extracted INTEGER,
                    logs_validated INTEGER,
                    repository_logs INTEGER,
                    unique_sources INTEGER,
                    coverage_score REAL,
                    status VARCHAR(50),
                    last_discovery_date VARCHAR(100)
                )
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS technology_aliases (
                    technology_name VARCHAR(255),
                    alias VARCHAR(255) UNIQUE
                )
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS technology_log_profile (
                    technology_name VARCHAR(255) PRIMARY KEY,
                    total_logs INTEGER,
                    unique_sources INTEGER,
                    unique_error_codes INTEGER,
                    unique_event_types INTEGER,
                    unique_components INTEGER,
                    first_seen VARCHAR(100),
                    last_seen VARCHAR(100),
                    error_diversity_score INTEGER
                )
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS notification_queue (
                    id SERIAL PRIMARY KEY,
                    created_at VARCHAR(100),
                    notification_type VARCHAR(100),
                    status VARCHAR(50),
                    content TEXT,
                    retry_count INTEGER DEFAULT 0,
                    attempt_number INTEGER DEFAULT 0,
                    last_attempt VARCHAR(100),
                    next_retry VARCHAR(100),
                    delivered_at VARCHAR(100)
                )
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS notification_history (
                    id SERIAL PRIMARY KEY,
                    timestamp VARCHAR(100),
                    notification_type VARCHAR(100),
                    recipient VARCHAR(255),
                    subject TEXT,
                    status VARCHAR(50),
                    error_message TEXT,
                    severity VARCHAR(50),
                    job_id VARCHAR(255),
                    technology VARCHAR(255)
                )
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS agent_job_history (
                    id SERIAL PRIMARY KEY,
                    job_type VARCHAR(255),
                    start_time VARCHAR(100),
                    end_time VARCHAR(100),
                    status VARCHAR(50),
                    records_processed INTEGER,
                    errors TEXT
                )
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS agent_health_history (
                    id SERIAL PRIMARY KEY,
                    timestamp VARCHAR(100),
                    component VARCHAR(255),
                    status VARCHAR(50),
                    details TEXT
                )
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS agent_runtime_metrics (
                    id SERIAL PRIMARY KEY,
                    timestamp VARCHAR(100),
                    technologies_processed INTEGER,
                    urls_crawled INTEGER,
                    logs_extracted INTEGER,
                    logs_validated INTEGER,
                    logs_inserted INTEGER,
                    failures INTEGER,
                    urls_skipped INTEGER DEFAULT 0,
                    pages_log_rich INTEGER DEFAULT 0,
                    pages_low_value INTEGER DEFAULT 0,
                    insert_yield_pct REAL DEFAULT 0.0
                )
            """)
        else:
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS validated_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    platform TEXT,
                    product_name TEXT,
                    log_type TEXT,
                    log_message TEXT,
                    normalized_log TEXT,
                    source_url TEXT,
                    source_title TEXT,
                    source_type TEXT,
                    source_rank INTEGER,
                    validation_score REAL,
                    discovery_query TEXT,
                    discovered_at TEXT,
                    raw_hash TEXT,
                    normalized_hash TEXT UNIQUE,
                    platform_category TEXT,
                    vendor TEXT,
                    technology_version TEXT,
                    source_domain TEXT,
                    log_severity TEXT,
                    event_type TEXT,
                    component TEXT,
                    hostname TEXT,
                    process_name TEXT,
                    error_code TEXT,
                    first_seen TEXT,
                    last_seen TEXT,
                    source_status TEXT,
                    fallback_source_url TEXT,
                    fallback_source_title TEXT
                )
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS repository_metrics (
                    key TEXT PRIMARY KEY,
                    value INTEGER
                )
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS repository_discovery_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    platform TEXT,
                    product_name TEXT,
                    log_type TEXT,
                    source_url TEXT,
                    discovered_at TEXT,
                    status TEXT,
                    validation_result TEXT
                )
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS technology_catalog (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    technology_name TEXT UNIQUE,
                    category TEXT,
                    vendor TEXT,
                    discovery_source TEXT,
                    first_seen TEXT,
                    last_seen TEXT,
                    status TEXT,
                    log_queries TEXT,
                    classification TEXT,
                    technology_confidence REAL,
                    accepted INTEGER,
                    rejection_reason TEXT
                )
            """)
            
            # Run migrations for technology_catalog columns if they don't exist
            migrations = [
                ("classification", "TEXT"),
                ("technology_confidence", "REAL"),
                ("accepted", "INTEGER"),
                ("rejection_reason", "TEXT")
            ]
            for col_name, col_type in migrations:
                try:
                    cursor.execute(f"ALTER TABLE technology_catalog ADD COLUMN {col_name} {col_type}")
                    conn.commit()
                except Exception:
                    pass

            cursor.execute("""
                CREATE TABLE IF NOT EXISTS technology_coverage (
                    technology_name TEXT PRIMARY KEY,
                    category TEXT,
                    vendor TEXT,
                    search_results INTEGER,
                    urls_crawled INTEGER,
                    logs_extracted INTEGER,
                    logs_validated INTEGER,
                    repository_logs INTEGER,
                    unique_sources INTEGER,
                    coverage_score REAL,
                    status TEXT,
                    last_discovery_date TEXT
                )
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS technology_aliases (
                    technology_name TEXT,
                    alias TEXT UNIQUE
                )
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS technology_log_profile (
                    technology_name TEXT PRIMARY KEY,
                    total_logs INTEGER,
                    unique_sources INTEGER,
                    unique_error_codes INTEGER,
                    unique_event_types INTEGER,
                    unique_components INTEGER,
                    first_seen TEXT,
                    last_seen TEXT,
                    error_diversity_score INTEGER
                )
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS notification_queue (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    created_at TEXT,
                    notification_type TEXT,
                    status TEXT,
                    content TEXT,
                    retry_count INTEGER DEFAULT 0,
                    attempt_number INTEGER DEFAULT 0,
                    last_attempt TEXT,
                    next_retry TEXT,
                    delivered_at TEXT
                )
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS notification_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    timestamp TEXT,
                    notification_type TEXT,
                    recipient TEXT,
                    subject TEXT,
                    status TEXT,
                    error_message TEXT,
                    severity TEXT,
                    job_id TEXT,
                    technology TEXT
                )
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS agent_job_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    job_type TEXT,
                    start_time TEXT,
                    end_time TEXT,
                    status TEXT,
                    records_processed INTEGER,
                    errors TEXT
                )
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS agent_health_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    timestamp TEXT,
                    component TEXT,
                    status TEXT,
                    details TEXT
                )
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS agent_runtime_metrics (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    timestamp TEXT,
                    technologies_processed INTEGER,
                    urls_crawled INTEGER,
                    logs_extracted INTEGER,
                    logs_validated INTEGER,
                    logs_inserted INTEGER,
                    failures INTEGER,
                    urls_skipped INTEGER DEFAULT 0,
                    pages_log_rich INTEGER DEFAULT 0,
                    pages_low_value INTEGER DEFAULT 0,
                    insert_yield_pct REAL DEFAULT 0.0
                )
            """)

        DEFAULT_ALIASES = {
            "Amazon S3": ["s3", "amazon s3"],
            "Amazon EC2": ["ec2", "amazon ec2"],
            "Amazon RDS": ["rds", "amazon rds"],
            "AWS Lambda": ["lambda", "aws lambda"],
            "Amazon DynamoDB": ["dynamodb", "amazon dynamodb"],
            "Azure Blob Storage": ["azure blob storage", "blob storage"],
            "Azure Virtual Machines": ["azure virtual machines", "azure vm"],
            "Azure SQL Database": ["azure sql", "azure sql database"],
            "Azure Functions": ["azure functions"],
            "Google Cloud Storage": ["google cloud storage", "gcs"],
            "Google Compute Engine": ["google compute engine", "gce"],
            "Google Cloud SQL": ["google cloud sql"],
            "Google Cloud Functions": ["google cloud functions"],
            "Prometheus": ["prometheus"],
            "Envoy": ["envoy"],
            "CoreDNS": ["coredns"],
            "containerd": ["containerd"],
            "Docker Engine": ["docker", "docker engine"],
            "Docker Compose": ["docker compose"],
            "Docker Registry": ["docker registry"],
            "Kubelet": ["kubelet"],
            "Kube-proxy": ["kube-proxy"],
            "Kubectl": ["kubectl"],
            "Oracle Database": ["oracle database", "oracle db", "oracle"],
            "WebLogic Server": ["weblogic", "weblogic server"],
            "VirtualBox": ["virtualbox"],
            "VMware ESXi": ["vmware esxi", "esxi"],
            "VMware vCenter": ["vcenter", "vmware vcenter"],
            "VMware NSX": ["nsx", "vmware nsx"],
            "Red Hat Enterprise Linux": ["rhel", "red hat enterprise linux"],
            "OpenShift": ["openshift"],
            "Ansible": ["ansible"]
        }
        
        cursor.execute("SELECT COUNT(*) FROM technology_aliases")
        if cursor.fetchone()[0] == 0:
            for tech_name, aliases in DEFAULT_ALIASES.items():
                for alias in aliases:
                    cursor.execute(
                        "INSERT OR IGNORE INTO technology_aliases (technology_name, alias) VALUES (?, ?)",
                        (tech_name, alias)
                    )
            conn.commit()

        cursor.execute("INSERT OR IGNORE INTO repository_metrics (key, value) VALUES ('duplicates_skipped', 0)")
        conn.commit()
    finally:
        cursor.close()
    
    migrate_repo_db(conn)



def insert_validated_logs(logs, job_platform=None, job_product_name=None, job_log_type=None, job_error_code=None):
    """
    Inserts validated logs into the validated_logs table in validated_logs.db.
    Handles raw and normalized hashes, duplicates_skipped counting, discovery history logging,
    and extraction of RCA/observability fields.
    """
    import hashlib
    import re
    from urllib.parse import urlparse
    from backend.validator import replace_timestamps_with_marker, normalize_text_for_match
    
    conn = get_repo_connection()
    cursor = conn.cursor()
    
    inserted = 0
    duplicates = 0
    
    now = datetime.utcnow().isoformat() + "Z"
    
    # We import discover_and_classify_platform dynamically to avoid circular import issues
    from backend.crawler import discover_and_classify_platform
    
    for log in logs:
        validation_dict = log.get("validation", {})
        if not validation_dict.get("valid"):
            continue
            
        # Get raw log text
        log_message = log.get("message") or ""
        original_log = log.get("original_log") or log_message or ""
        original_log_str = str(original_log).strip()
        
        # Calculate raw_hash
        raw_hash = hashlib.sha256(original_log_str.encode('utf-8')).hexdigest()
        
        # Calculate normalized_log
        ts_marker = replace_timestamps_with_marker(original_log_str)
        normalized_log = normalize_text_for_match(ts_marker)
        
        # Calculate normalized_hash
        normalized_hash = hashlib.sha256(normalized_log.encode('utf-8')).hexdigest()
        
        # Get basic fields
        platform = log.get("platform") or job_platform or ""
        product_name = log.get("product_name") or job_product_name or ""
        log_type = log.get("source_service") or log.get("service") or job_log_type or ""
        source_url = log.get("source_url") or ""
        source_title = log.get("source_title") or ""
        source_type = log.get("source_type") or validation_dict.get("source_type") or ""
        source_rank = log.get("source_rank") or validation_dict.get("source_rank") or 4
        validation_score = validation_dict.get("confidence") or 0
        discovery_query = log.get("query_used") or validation_dict.get("query_used") or ""
        
        # Extract RCA/observability fields
        # 1. discover_and_classify_platform
        plat_category, vendor, tech_ver = discover_and_classify_platform(platform)
        if not tech_ver:
            # fallback to version in log
            tech_ver = log.get("source_version") or log.get("version") or ""
            
        # 2. source_domain
        source_domain = ""
        if source_url:
            try:
                source_domain = urlparse(source_url).netloc.lower()
                if source_domain.startswith("www."):
                    source_domain = source_domain[4:]
            except:
                pass
                
        # 3. log_severity
        log_severity = log.get("severity") or ""
        if not log_severity:
            # extract from log
            from backend.validator import extract_severity
            log_severity = extract_severity(original_log_str) or "INFO"
            
        # 4. event_type
        event_type = ""
        if "exception" in original_log_str.lower():
            match = re.search(r'\b([A-Za-z0-9_]+(?:Exception|Error))\b', original_log_str)
            if match:
                event_type = match.group(1)
        if not event_type:
            event_type = "LOG_EVENT"
            
        # 5. component
        component = log_type or "unknown"
        
        # 6. hostname
        hostname = ""
        ip_match = re.search(r'\b\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\b', original_log_str)
        if ip_match:
            hostname = ip_match.group(0)
            
        # 7. process_name
        process_name = ""
        proc_match = re.search(r'\b([a-zA-Z0-9_\-]+)\[\d+\]\:', original_log_str)
        if proc_match:
            process_name = proc_match.group(1)
        else:
            proc_match2 = re.search(r'\b\S+\s+([a-zA-Z0-9_\-\.\/]+)(?:\[\d+\])?\:\s*', original_log_str)
            if proc_match2:
                process_name = proc_match2.group(1)
                
        # 8. error_code
        error_code = log.get("error_code") or job_error_code or ""
        if not error_code:
            # search for error code pattern in log (e.g. exit code 1, HTTP 502, error code x)
            err_match = re.search(r'\b(?:exit\s+code\s+|error\s+code\s+|status\s+code\s+|HTTP\s+)([0-9a-zA-Z\-]+)\b', original_log_str, re.IGNORECASE)
            if err_match:
                error_code = err_match.group(1)
                
        # Check if normalized_hash exists
        cursor.execute("SELECT id, first_seen FROM validated_logs WHERE normalized_hash = ?", (normalized_hash,))
        existing_log = cursor.fetchone()
        
        if existing_log:
            # Duplicate detection - skip insert, update last_seen, increment counter
            duplicates += 1
            log_id = existing_log[0]
            cursor.execute(
                "UPDATE validated_logs SET last_seen = ? WHERE id = ?",
                (now, log_id)
            )
            # Record to discovery history
            cursor.execute(
                """INSERT INTO repository_discovery_history (
                    platform, product_name, log_type, source_url, discovered_at, status, validation_result
                ) VALUES (?, ?, ?, ?, ?, ?, ?)""",
                (platform, product_name, log_type, source_url, now, "duplicate_skipped", validation_dict.get("reason", "valid"))
            )
        else:
            # Determine source_status and fallback documentation URL/Title
            if source_url and source_url.strip():
                source_status = 'VERIFIED'
                fallback_url = None
                fallback_title = None
            else:
                source_status = 'MISSING'
                fallback_url, fallback_title = get_fallback_source(platform, product_name)

            # Insert log record (first_seen and last_seen set to now)
            cursor.execute(
                """INSERT INTO validated_logs (
                    platform, product_name, log_type, log_message, normalized_log,
                    source_url, source_title, source_type, source_rank, validation_score,
                    discovery_query, discovered_at, raw_hash, normalized_hash,
                    platform_category, vendor, technology_version, source_domain,
                    log_severity, event_type, component, hostname, process_name,
                    error_code, first_seen, last_seen,
                    source_status, fallback_source_url, fallback_source_title
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    platform, product_name, log_type, log_message, normalized_log,
                    source_url, source_title, source_type, source_rank, validation_score,
                    discovery_query, now, raw_hash, normalized_hash,
                    plat_category, vendor, tech_ver, source_domain,
                    log_severity, event_type, component, hostname, process_name,
                    error_code, now, now,
                    source_status, fallback_url, fallback_title
                )
            )
            inserted += 1
            # Record to discovery history
            cursor.execute(
                """INSERT INTO repository_discovery_history (
                    platform, product_name, log_type, source_url, discovered_at, status, validation_result
                ) VALUES (?, ?, ?, ?, ?, ?, ?)""",
                (platform, product_name, log_type, source_url, now, "inserted", validation_dict.get("reason", "valid"))
            )
            
    if duplicates > 0:
        cursor.execute("UPDATE repository_metrics SET value = value + ? WHERE key = 'duplicates_skipped'", (duplicates,))
        
    conn.commit()
    conn.close()
    return inserted, duplicates

def get_repository_stats_for_sheet():
    conn = get_repo_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT 
            platform, 
            product_name, 
            log_type, 
            COUNT(*) as total_logs, 
            COUNT(DISTINCT source_url) as unique_sources, 
            MAX(last_seen) as last_discovery_date
        FROM validated_logs
        GROUP BY platform, product_name, log_type
    """)
    rows = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return rows


def get_repository_health_data():
    conn = get_repo_connection()
    is_postgres = "sqlite3" not in str(type(conn))
    cursor = conn.cursor()
    
    # 1. Total Validated Logs
    cursor.execute("SELECT COUNT(*) FROM validated_logs")
    total_logs = cursor.fetchone()[0]
    
    # 2. Total Duplicates Skipped
    cursor.execute("SELECT value FROM repository_metrics WHERE key = 'duplicates_skipped'")
    row = cursor.fetchone()
    duplicates_skipped = row[0] if row else 0
    
    # 3. Unique Sources
    cursor.execute("SELECT COUNT(DISTINCT source_url) FROM validated_logs")
    unique_sources = cursor.fetchone()[0]
    
    # 4. Unique Platforms
    cursor.execute("SELECT COUNT(DISTINCT platform) FROM validated_logs")
    unique_platforms = cursor.fetchone()[0]
    
    # 5. Unique Products
    cursor.execute("SELECT COUNT(DISTINCT product_name) FROM validated_logs WHERE product_name IS NOT NULL AND product_name != ''")
    unique_products = cursor.fetchone()[0]
    
    # 6. Last Repository Update
    cursor.execute("SELECT MAX(last_seen) FROM validated_logs")
    last_update = cursor.fetchone()[0] or "N/A"
    
    # 7. Total Technologies
    cursor.execute("SELECT COUNT(*) FROM technology_catalog")
    total_technologies = cursor.fetchone()[0]
    
    # 8. Total Categories
    cursor.execute("SELECT COUNT(DISTINCT category) FROM technology_catalog WHERE category IS NOT NULL AND category != ''")
    total_categories = cursor.fetchone()[0]
    
    # 9. Total Vendors
    cursor.execute("SELECT COUNT(DISTINCT vendor) FROM technology_catalog WHERE vendor IS NOT NULL AND vendor != ''")
    total_vendors = cursor.fetchone()[0]
    
    # 10. Repository Size (MB)
    size_mb = 0.0
    if is_postgres:
        try:
            cursor.execute("SELECT pg_database_size(current_database())")
            size_bytes = cursor.fetchone()[0]
            size_mb = round(size_bytes / (1024 * 1024), 3)
        except Exception:
            size_mb = 0.0
    else:
        db_path = os.getenv("REPO_DB_PATH", "validated_logs.db")
        if not os.path.isabs(db_path):
            base_dir = os.path.dirname(os.path.abspath(__file__))
            db_path = os.path.join(base_dir, db_path)
        if os.path.exists(db_path):
            size_mb = round(os.path.getsize(db_path) / (1024 * 1024), 3)
            
    conn.close()
    
    return {
        "total_logs": total_logs,
        "duplicates_skipped": duplicates_skipped,
        "unique_sources": unique_sources,
        "unique_platforms": unique_platforms,
        "unique_products": unique_products,
        "last_update": last_update,
        "size_mb": size_mb,
        "total_technologies": total_technologies,
        "total_categories": total_categories,
        "total_vendors": total_vendors
    }


def recalculate_technology_coverage():
    """
    Recalculates global stats for all accepted technologies in the catalog
    and populates technology_coverage and technology_log_profile tables.
    Returns autonomous discovery readiness summary metrics.
    """
    conn = get_repo_connection()
    try:
        cursor = conn.cursor()
        try:
            # 1. Fetch all accepted technologies
            cursor.execute("SELECT technology_name, category, vendor FROM technology_catalog WHERE accepted = 1")
            accepted_techs = [dict(row) for row in cursor.fetchall()]
            
            # 2. Fetch all aliases
            cursor.execute("SELECT technology_name, alias FROM technology_aliases")
            aliases_rows = cursor.fetchall()
            aliases_dict = {}
            for row in aliases_rows:
                tname = row[0]
                alias = row[1].lower().strip()
                if tname not in aliases_dict:
                    aliases_dict[tname] = set()
                aliases_dict[tname].add(alias)
                
            # 3. Product mapping static setup
            PRODUCT_MAPPING = {
                "amazon simple storage service": ["Amazon S3", "s3"],
                "amazon elastic compute cloud": ["Amazon EC2", "ec2"],
                "amazon relational database service": ["Amazon RDS", "rds"],
                "amazon aurora mysql": ["MySQL", "Amazon RDS"],
            }
            
            # 4. Fetch all validated logs
            cursor.execute("""
                SELECT platform, product_name, source_url, first_seen, last_seen, 
                       error_code, event_type, component 
                FROM validated_logs
            """)
            all_logs = [dict(row) for row in cursor.fetchall()]
            
            # We clear the existing technology_coverage and technology_log_profile tables
            cursor.execute("DELETE FROM technology_coverage")
            cursor.execute("DELETE FROM technology_log_profile")
            conn.commit()
            
            # Get historical job row metrics from jobs database
            jobs_rows = []
            try:
                jconn, j_is_postgres = get_connection()
                try:
                    jcur = jconn.cursor()
                    try:
                        jcur.execute("SELECT platform, product_name, validated_count, non_validated_count, sources_found FROM job_rows")
                        jobs_rows = [dict(row) for row in jcur.fetchall()]
                    finally:
                        jcur.close()
                finally:
                    jconn.close()
            except Exception as je:
                print(f"Warning: could not read jobs database for metrics: {je}")
                    
            now = datetime.utcnow().isoformat() + "Z"
            
            # Helper to check match (Exact -> Alias -> Product Map)
            def check_match(tech_name, platform, product_name):
                tech_lower = tech_name.lower().strip()
                plat_lower = platform.lower().strip() if platform else ""
                prod_lower = product_name.lower().strip() if product_name else ""
                
                # Priority 1: Exact Technology Match
                if tech_lower == prod_lower or tech_lower == plat_lower:
                    return True
                    
                # Priority 2: Alias Match
                aliases = aliases_dict.get(tech_name, set())
                if (prod_lower in aliases) or (plat_lower in aliases):
                    return True
                    
                # Priority 3: Product Mapping
                mapped = PRODUCT_MAPPING.get(prod_lower)
                if mapped and tech_name in mapped:
                    return True
                mapped_plat = PRODUCT_MAPPING.get(plat_lower)
                if mapped_plat and tech_name in mapped_plat:
                    return True
                    
                return False
                
            for tech in accepted_techs:
                tech_name = tech["technology_name"]
                
                # Gather matching validated logs
                matching_logs = []
                for log in all_logs:
                    if check_match(tech_name, log["platform"], log["product_name"]):
                        matching_logs.append(log)
                        
                # Gather matching job rows for historical crawl metrics
                matching_jobs = []
                for jrow in jobs_rows:
                    if check_match(tech_name, jrow["platform"], jrow["product_name"]):
                        matching_jobs.append(jrow)
                        
                # Aggregate coverage fields
                search_results = sum(jr.get("sources_found", 0) or 0 for jr in matching_jobs)
                logs_validated = sum(jr.get("validated_count", 0) or 0 for jr in matching_jobs)
                logs_extracted = sum((jr.get("validated_count", 0) or 0) + (jr.get("non_validated_count", 0) or 0) for jr in matching_jobs)
                
                repository_logs = len(matching_logs)
                unique_srcs = len(set(log["source_url"] for log in matching_logs if log["source_url"]))
                
                if repository_logs > 0:
                    if logs_validated == 0:
                        logs_validated = repository_logs
                    if logs_extracted == 0:
                        logs_extracted = repository_logs
                    if search_results == 0:
                        search_results = unique_srcs
                        
                urls_crawled = unique_srcs
                
                # Get last discovery date
                last_discovery_date = ""
                first_seen_date = ""
                if matching_logs:
                    dates = []
                    for log in matching_logs:
                        d = log.get("last_seen") or log.get("first_seen") or log.get("discovered_at")
                        if d:
                            dates.append(d)
                    if dates:
                        last_discovery_date = max(dates)
                        first_seen_date = min(dates)
                        
                if not last_discovery_date:
                    last_discovery_date = "N/A"
                if not first_seen_date:
                    first_seen_date = now
                    
                # RCA / intelligence profile details
                unique_error_codes = len(set(log["error_code"] for log in matching_logs if log["error_code"]))
                unique_event_types = len(set(log["event_type"] for log in matching_logs if log["event_type"]))
                unique_components = len(set(log["component"] for log in matching_logs if log["component"]))
                
                error_diversity_score = unique_error_codes + unique_event_types
                
                # Compute status and coverage score
                coverage_score = min(100.0, float(repository_logs * 1.0 + unique_srcs * 5.0))
                
                # Status thresholds (Source and Error Diversity Aware)
                if repository_logs == 0:
                    status = "NO_LOGS"
                elif repository_logs < 10:
                    status = "WEAK"
                elif repository_logs >= 50 and unique_srcs >= 5 and error_diversity_score >= 4:
                    status = "EXCELLENT"
                elif repository_logs >= 20 and unique_srcs >= 3 and error_diversity_score >= 2:
                    status = "GOOD"
                else:
                    status = "PARTIAL"
                    
                # Write to technology_coverage table
                cursor.execute(
                    """INSERT INTO technology_coverage (
                        technology_name, category, vendor, search_results, urls_crawled, 
                        logs_extracted, logs_validated, repository_logs, unique_sources, 
                        coverage_score, status, last_discovery_date
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (tech_name, tech["category"], tech["vendor"], search_results, urls_crawled,
                     logs_extracted, logs_validated, repository_logs, unique_srcs,
                     coverage_score, status, last_discovery_date)
                )
                
                # Write to technology_log_profile table
                cursor.execute(
                    """INSERT INTO technology_log_profile (
                        technology_name, total_logs, unique_sources, unique_error_codes, 
                        unique_event_types, unique_components, first_seen, last_seen, error_diversity_score
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (tech_name, repository_logs, unique_srcs, unique_error_codes,
                     unique_event_types, unique_components, first_seen_date, last_discovery_date, error_diversity_score)
                )
                
            conn.commit()
            
            # Calculate readiness summary metrics
            cursor.execute("SELECT COUNT(*) FROM technology_catalog WHERE accepted = 1")
            techs_cataloged = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) FROM technology_coverage WHERE repository_logs > 0")
            techs_with_logs = cursor.fetchone()[0]
            techs_without_logs = techs_cataloged - techs_with_logs
            
            cursor.execute("SELECT logs_validated, logs_extracted FROM technology_coverage WHERE logs_extracted > 0")
            rates = []
            for r in cursor.fetchall():
                rates.append(r[0] / r[1])
            avg_val_rate = (sum(rates) / len(rates) * 100.0) if rates else 0.0
            
            cursor.execute("SELECT AVG(unique_sources) FROM technology_coverage")
            avg_srcs = cursor.fetchone()[0] or 0.0
            
            cursor.execute("SELECT COUNT(*) FROM validated_logs")
            total_logs = cursor.fetchone()[0]
            
            growth_rate = 0.0
            if total_logs > 0:
                import datetime as dt
                day_ago = (dt.datetime.utcnow() - dt.timedelta(days=1)).isoformat()
                cursor.execute("SELECT COUNT(*) FROM validated_logs WHERE discovered_at >= ?", (day_ago,))
                new_logs = cursor.fetchone()[0]
                growth_rate = (new_logs / total_logs) * 100.0
                
            return {
                "technologies_cataloged": techs_cataloged,
                "technologies_producing_logs": techs_with_logs,
                "technologies_without_logs": techs_without_logs,
                "average_validation_rate": avg_val_rate,
                "average_sources_per_technology": avg_srcs,
                "repository_growth_rate": growth_rate
            }
        finally:
            cursor.close()
    finally:
        conn.close()


def is_downgraded_domain(domain: str) -> bool:
    """
    Check if a domain has been downgraded due to poor yield.
    A domain is downgraded if it has >= 20 URLs crawled and 0 validated logs.
    """
    if not domain:
        return False
    conn = get_repo_connection()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT urls_crawled, logs_validated FROM domain_performance WHERE domain = ?",
        (domain,)
    )
    row = cursor.fetchone()
    conn.close()
    if row:
        return row["urls_crawled"] >= 20 and row["logs_validated"] == 0
    return False


def get_gap_technologies():
    """
    Return accepted technologies that have ZERO validated logs in the repository.
    These are the technologies that need focused discovery.
    """
    conn = get_repo_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT tc.technology_name, tc.category
        FROM technology_catalog tc
        LEFT JOIN validated_logs vl 
            ON LOWER(tc.technology_name) = LOWER(vl.product_name)
        WHERE tc.accepted = 1
        GROUP BY tc.technology_name, tc.category
        HAVING COUNT(vl.id) = 0
        ORDER BY tc.technology_name
    """)
    rows = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return rows


def update_domain_performance(urls_info):
    import urllib.parse
    conn = get_repo_connection()
    cursor = conn.cursor()
    
    # 1. Group urls_info by domain
    domain_stats = {}
    for info in urls_info:
        url = info.get("URL") or info.get("url") or ""
        if not url:
            continue
        try:
            domain = urllib.parse.urlparse(url).netloc.lower()
            if domain.startswith("www."):
                domain = domain[4:]
            if not domain:
                continue
                
            domain_stats.setdefault(domain, {
                "urls_crawled": 0,
                "logs_extracted": 0,
                "logs_validated": 0
            })
            ds = domain_stats[domain]
            if info.get("Crawled") == "Yes" or info.get("crawled") == "Yes":
                ds["urls_crawled"] += 1
            ds["logs_extracted"] += int(info.get("Logs Extracted", 0) or info.get("logs_extracted", 0) or 0)
            ds["logs_validated"] += int(info.get("Logs Validated", 0) or info.get("logs_validated", 0) or 0)
        except Exception:
            pass
            
    # 2. Update domain_performance table
    for domain, ds in domain_stats.items():
        cursor.execute("SELECT urls_crawled, logs_extracted, logs_validated FROM domain_performance WHERE domain = ?", (domain,))
        row = cursor.fetchone()
        was_downgraded = False
        if row:
            new_crawled = row[0] + ds["urls_crawled"]
            new_extracted = row[1] + ds["logs_extracted"]
            new_validated = row[2] + ds["logs_validated"]
            was_downgraded = (row[0] >= 20 and row[2] == 0)
        else:
            new_crawled = ds["urls_crawled"]
            new_extracted = ds["logs_extracted"]
            new_validated = ds["logs_validated"]
            
        # Count actual logs inserted in database from this domain
        cursor.execute("SELECT COUNT(*) FROM validated_logs WHERE source_domain = ?", (domain,))
        logs_inserted = cursor.fetchone()[0]
        
        yield_score = (new_validated / new_crawled) if new_crawled > 0 else 0.0
        
        cursor.execute(
            """INSERT OR REPLACE INTO domain_performance (
                domain, urls_crawled, logs_extracted, logs_validated, logs_inserted, yield_score
            ) VALUES (?, ?, ?, ?, ?, ?)""",
            (domain, new_crawled, new_extracted, new_validated, logs_inserted, yield_score)
        )

        is_now_downgraded = (new_crawled >= 20 and new_validated == 0)
        if is_now_downgraded and not was_downgraded:
            try:
                from backend.notifications import send_event_notification
                subject = f"Domain Downgraded - {domain}"
                body = (
                    f"A domain has been automatically downgraded due to poor yield (0 validated logs after at least 20 crawled URLs).\n\n"
                    f"Domain: {domain}\n"
                    f"URLs Crawled: {new_crawled}\n"
                    f"Logs Validated: {new_validated}\n"
                    f"Yield Score: {yield_score:.2f} ({yield_score * 100:.1f}%)\n"
                    f"Reason: 0 validated logs out of {new_crawled} crawled URLs\n"
                    f"Future Crawl Status: Downgraded / Skipped"
                )
                send_event_notification(
                    event_type="domain_downgraded",
                    severity="WARNING",
                    subject=subject,
                    content_body=body,
                    technology=None
                )
            except Exception as notify_err:
                print(f"Failed to send domain downgrade notification: {notify_err}")
        
    conn.commit()
    conn.close()


def repopulate_domain_performance_from_history():
    import glob
    import urllib.parse
    import openpyxl
    
    conn = get_repo_connection()
    cursor = conn.cursor()
    
    cursor.execute("DELETE FROM domain_performance")
    conn.commit()
    
    base_dir = os.path.dirname(os.path.abspath(__file__))
    outputs_dir = os.path.join(base_dir, "outputs")
    
    domain_stats = {}
    
    # 1. Fetch inserts from live validated_logs
    cursor.execute("SELECT source_url FROM validated_logs WHERE source_url IS NOT NULL AND source_url != ''")
    for r in cursor.fetchall():
        url = r[0]
        try:
            domain = urllib.parse.urlparse(url).netloc.lower()
            if domain.startswith("www."):
                domain = domain[4:]
            if domain:
                domain_stats.setdefault(domain, {
                    "urls_crawled": 0,
                    "logs_extracted": 0,
                    "logs_validated": 0,
                    "logs_inserted": 0
                })
                domain_stats[domain]["logs_inserted"] += 1
        except Exception:
            pass
            
    # 2. Parse Excel logs from job outputs
    xlsx_files = glob.glob(os.path.join(outputs_dir, "job_*", "diagnostic_report.xlsx"))
    for f in xlsx_files:
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
            if "SOURCE_AUDIT" in wb.sheetnames:
                ws = wb["SOURCE_AUDIT"]
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) > 1:
                    headers = rows[0]
                    for r in rows[1:]:
                        if not r or len(r) < 13:
                            continue
                        row_dict = dict(zip(headers, r))
                        url = row_dict.get("URL")
                        if not url:
                            continue
                        try:
                            domain = urllib.parse.urlparse(url).netloc.lower()
                            if domain.startswith("www."):
                                domain = domain[4:]
                            if not domain:
                                continue
                                
                            domain_stats.setdefault(domain, {
                                "urls_crawled": 0,
                                "logs_extracted": 0,
                                "logs_validated": 0,
                                "logs_inserted": 0
                            })
                            ds = domain_stats[domain]
                            if row_dict.get("Crawled") == "Yes":
                                ds["urls_crawled"] += 1
                            ds["logs_extracted"] += int(row_dict.get("Logs Extracted", 0) or 0)
                            ds["logs_validated"] += int(row_dict.get("Logs Validated", 0) or 0)
                            if row_dict.get("Final Status") == "inserted":
                                ds["logs_inserted"] += 1
                        except Exception:
                            pass
            wb.close()
        except Exception:
            pass
            
    # 3. Calculate yields and write back
    for domain, ds in domain_stats.items():
        urls = ds["urls_crawled"]
        validated = ds["logs_validated"]
        yield_score = (validated / urls) if urls > 0 else 0.0
        
        cursor.execute(
            """INSERT OR REPLACE INTO domain_performance (
                domain, urls_crawled, logs_extracted, logs_validated, logs_inserted, yield_score
            ) VALUES (?, ?, ?, ?, ?, ?)""",
            (domain, urls, ds["logs_extracted"], validated, ds["logs_inserted"], yield_score)
        )
        
    conn.commit()
    conn.close()


def log_agent_event(event_type: str, message: str):
    """
    Inserts a record into the agent_event_feed table.
    """
    import datetime as dt
    conn = get_repo_connection()
    cursor = conn.cursor()
    now_str = dt.datetime.utcnow().isoformat() + "Z"
    cursor.execute(
        "INSERT INTO agent_event_feed (timestamp, event_type, message) VALUES (?, ?, ?)",
        (now_str, event_type, message)
    )
    conn.commit()
    conn.close()


def cleanup_event_logs():
    """
    Retention rules for agent_event_feed: Keep latest 1000 logs or 7 days retention.
    """
    from datetime import datetime, timedelta
    conn = None
    try:
        conn = get_repo_connection()
        cursor = conn.cursor()
        
        # 1. Delete logs older than 7 days
        seven_days_ago = (datetime.utcnow() - timedelta(days=7)).isoformat() + "Z"
        cursor.execute("DELETE FROM agent_event_feed WHERE timestamp < ?", (seven_days_ago,))
        
        # 2. Keep only latest 1000 logs
        cursor.execute("SELECT COUNT(*) FROM agent_event_feed")
        count = cursor.fetchone()[0]
        if count > 1000:
            limit = count - 1000
            cursor.execute("""
                DELETE FROM agent_event_feed
                WHERE id IN (
                    SELECT id FROM agent_event_feed
                    ORDER BY id ASC LIMIT ?
                )
            """, (limit,))
            
        conn.commit()
    except Exception as e:
        print(f"Error cleaning up event logs: {e}")
    finally:
        if conn:
            conn.close()


def get_domain_metrics(domain: str) -> dict:
    """
    Get metrics (urls_crawled, logs_validated, yield_score) for a domain.
    """
    if not domain:
        return {"urls_crawled": 0, "logs_validated": 0, "yield_score": 0.0}
    conn = get_repo_connection()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT urls_crawled, logs_validated, yield_score FROM domain_performance WHERE domain = ?",
        (domain,)
    )
    row = cursor.fetchone()
    conn.close()
    if row:
        return {
            "urls_crawled": row[0] or 0,
            "logs_validated": row[1] or 0,
            "yield_score": row[2] or 0.0
        }
    return {"urls_crawled": 0, "logs_validated": 0, "yield_score": 0.0}


def update_agent_status_field(**kwargs):
    """
    Updates specific fields in the agent_status table.
    """
    conn = None
    try:
        conn = get_repo_connection()
        cursor = conn.cursor()
        updates = []
        params = []
        for k, v in kwargs.items():
            if v is not None:
                updates.append(f"{k} = ?")
                params.append(v)
            else:
                updates.append(f"{k} = NULL")
        if updates:
            cursor.execute(f"UPDATE agent_status SET {', '.join(updates)} WHERE id = 1", params)
            conn.commit()
    except Exception as e:
        print(f"Error updating agent status field: {e}")
    finally:
        if conn:
            conn.close()


def execute_repo_write(query, params=()):
    conn = get_repo_connection()
    try:
        is_postgres = "sqlite3" not in str(type(conn))
        cursor = conn.cursor()
        try:
            if is_postgres:
                query = query.replace("?", "%s")
            cursor.execute(query, params)
            conn.commit()
        finally:
            cursor.close()
    finally:
        conn.close()


def execute_repo_read(query, params=()):
    conn = get_repo_connection()
    try:
        is_postgres = "sqlite3" not in str(type(conn))
        if is_postgres:
            from psycopg2.extras import RealDictCursor
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            try:
                query = query.replace("?", "%s")
                cursor.execute(query, params)
                results = [dict(row) for row in cursor.fetchall()]
            finally:
                cursor.close()
        else:
            cursor = conn.cursor()
            try:
                cursor.execute(query, params)
                results = [dict(row) for row in cursor.fetchall()]
            finally:
                cursor.close()
    finally:
        conn.close()
    return results


def execute_repo_insert(query, params=()):
    conn = get_repo_connection()
    try:
        is_postgres = "sqlite3" not in str(type(conn))
        cursor = conn.cursor()
        try:
            if is_postgres:
                query = query.replace("?", "%s")
                if "INSERT INTO" in query.upper() and "RETURNING" not in query.upper():
                    parts = query.split()
                    table_name = ""
                    try:
                        into_idx = -1
                        for idx, part in enumerate(parts):
                            if part.upper() == "INTO":
                                into_idx = idx
                                break
                        if into_idx != -1 and into_idx + 1 < len(parts):
                            table_name = parts[into_idx + 1].split("(")[0].strip().strip('"').strip('`').strip("'")
                    except Exception:
                        pass
                    
                    if not table_name:
                        try:
                            table_name = parts[2].split("(")[0].strip().strip('"').strip('`').strip("'")
                        except Exception:
                            table_name = ""
                            
                    tables_without_id = {
                        "agent_health_alert_states",
                        "repository_metrics",
                        "technology_coverage",
                        "technology_log_profile",
                        "domain_performance",
                        "technology_aliases"
                    }
                    
                    if table_name and table_name.lower() not in tables_without_id:
                        query = query.rstrip().rstrip(";") + " RETURNING id"
                        cursor.execute(query, params)
                        last_id = cursor.fetchone()[0]
                    else:
                        cursor.execute(query, params)
                        last_id = None
                else:
                    cursor.execute(query, params)
                    last_id = None
            else:
                cursor.execute(query, params)
                last_id = cursor.lastrowid
            conn.commit()
        finally:
            cursor.close()
    finally:
        conn.close()
    return last_id


def table_exists(table_name):
    conn = get_repo_connection()
    try:
        is_postgres = "sqlite3" not in str(type(conn))
        cursor = conn.cursor()
        try:
            if is_postgres:
                cursor.execute("SELECT EXISTS (SELECT FROM information_schema.tables WHERE table_name = %s)", (table_name.lower(),))
                exists = cursor.fetchone()[0]
            else:
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
                exists = cursor.fetchone() is not None
        finally:
            cursor.close()
    finally:
        conn.close()
    return exists


# Helpers for notifications.py
def check_notification_cooldown(event_type, technology=None, subject=None, cooldown_mins=0):
    if cooldown_mins <= 0:
        return False
    from datetime import datetime, timedelta
    lookback_time = (datetime.utcnow() - timedelta(minutes=cooldown_mins)).isoformat() + "Z"
    if technology:
        query = """
            SELECT COUNT(*) FROM notification_history
            WHERE notification_type = ? AND technology = ? AND status = 'Sent' AND timestamp >= ?
        """
        params = (event_type, technology, lookback_time)
    else:
        query = """
            SELECT COUNT(*) FROM notification_history
            WHERE notification_type = ? AND subject = ? AND status = 'Sent' AND timestamp >= ?
        """
        params = (event_type, subject, lookback_time)
    
    rows = execute_repo_read(query, params)
    if rows:
        count = list(rows[0].values())[0]
        return count > 0
    return False


def save_event_notification(event_type, recipient, subject, initial_status, queue_content, is_throttled, severity, job_id, technology, now_str):
    q_query = """
        INSERT INTO notification_queue (created_at, notification_type, status, content, retry_count, attempt_number, next_retry)
        VALUES (?, ?, ?, ?, 0, 0, ?)
    """
    q_params = (now_str, event_type, initial_status, queue_content, now_str)
    queue_id = execute_repo_insert(q_query, q_params)
    
    h_query = """
        INSERT INTO notification_history (timestamp, notification_type, recipient, subject, status, error_message, severity, job_id, technology)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """
    h_params = (now_str, event_type, recipient, subject, initial_status, "Throttled (cooldown active)" if is_throttled else "", severity, job_id, technology)
    history_id = execute_repo_insert(h_query, h_params)
    
    return queue_id, history_id


def update_notification_status(queue_id, history_id, status, error_msg="", next_retry=None):
    from datetime import datetime
    now_str = datetime.utcnow().isoformat() + "Z"
    if status == "Sent":
        q_query = """
            UPDATE notification_queue
            SET status = ?, attempt_number = attempt_number + 1, last_attempt = ?, delivered_at = ?
            WHERE id = ?
        """
        q_params = (status, now_str, now_str, queue_id)
    elif status == "Retrying":
        q_query = """
            UPDATE notification_queue
            SET status = ?, attempt_number = attempt_number + 1, last_attempt = ?, next_retry = ?
            WHERE id = ?
        """
        q_params = (status, now_str, next_retry, queue_id)
    else:
        q_query = """
            UPDATE notification_queue
            SET status = ?, attempt_number = attempt_number + 1, last_attempt = ?
            WHERE id = ?
        """
        q_params = (status, now_str, queue_id)
        
    execute_repo_write(q_query, q_params)
        
    h_query = """
        UPDATE notification_history
        SET status = ?, error_message = ?
        WHERE id = ?
    """
    h_params = (status, error_msg, history_id)
    execute_repo_write(h_query, h_params)


def get_pending_notifications(now_str):
    query = """
        SELECT id, notification_type, content, retry_count, attempt_number FROM notification_queue
        WHERE status IN ('Pending', 'Failed', 'Retrying') AND (next_retry IS NULL OR next_retry <= ?)
    """
    rows = execute_repo_read(query, (now_str,))
    return [(r["id"], r["notification_type"], r["content"], r["retry_count"], r["attempt_number"]) for r in rows]


def mark_notification_sent_and_history(q_id, new_retry_count, new_attempt_number, now_str, n_type, smtp_to, subject, severity, job_id, technology):
    q_query = """
        UPDATE notification_queue
        SET status = 'Sent', retry_count = ?, attempt_number = ?, last_attempt = ?, delivered_at = ?, next_retry = NULL
        WHERE id = ?
    """
    execute_repo_write(q_query, (new_retry_count, new_attempt_number, now_str, now_str, q_id))
    
    h_query = """
        INSERT INTO notification_history (timestamp, notification_type, recipient, subject, status, error_message, severity, job_id, technology)
        VALUES (?, ?, ?, ?, 'Sent', '', ?, ?, ?)
    """
    execute_repo_insert(h_query, (now_str, n_type, smtp_to, f"[Retry Success] {subject}", severity, job_id, technology))


# Helpers for main.py
def get_dashboard_metrics_data(day_ago):
    r_techs = execute_repo_read("SELECT COUNT(*) FROM technology_catalog WHERE accepted = 1")
    techs_tracked = list(r_techs[0].values())[0] if r_techs else 0
    
    db_path = os.getenv("REPO_DB_PATH", "validated_logs.db")
    if not os.path.isabs(db_path):
        base_dir = os.path.dirname(os.path.abspath(__file__))
        db_path = os.path.join(base_dir, db_path)
    size_mb = 0.0
    if os.path.exists(db_path):
        size_mb = round(os.path.getsize(db_path) / (1024 * 1024), 3)
        
    r_logs = execute_repo_read("SELECT COUNT(*) FROM validated_logs WHERE discovered_at >= ?", (day_ago,))
    logs_added_today = list(r_logs[0].values())[0] if r_logs else 0
    
    r_jobs = execute_repo_read("SELECT COUNT(*) FROM agent_job_history WHERE status = 'failed' AND start_time >= ?", (day_ago,))
    failed_jobs = list(r_jobs[0].values())[0] if r_jobs else 0
    
    r_health = execute_repo_read("SELECT COUNT(*) FROM agent_health_history WHERE status != 'healthy' AND timestamp >= ?", (day_ago,))
    failed_health = list(r_health[0].values())[0] if r_health else 0
    
    return {
        "technologies_tracked": techs_tracked,
        "repository_size_mb": size_mb,
        "logs_added_today": logs_added_today,
        "failures_today": failed_jobs + failed_health
    }


def get_agent_status_metrics():
    query = """
        SELECT status, current_job, current_tech,
               technologies_processed, technologies_total,
               current_query, total_queries,
               current_url, total_urls,
               cycle_start_time, cycle_urls_crawled, cycle_pages_classified,
               cycle_logs_extracted, cycle_logs_validated, cycle_logs_inserted,
               cycle_duplicates_skipped, current_phase,
               next_technology_discovery, next_log_discovery, next_health_check, next_daily_report
        FROM agent_status WHERE id = 1
    """
    rows = execute_repo_read(query)
    if rows:
        return rows[0]
    return None


def get_dashboard_technologies():
    query = """
        SELECT technology_name, status, repository_logs, last_discovery_date 
        FROM technology_coverage 
        ORDER BY technology_name ASC
    """
    rows = execute_repo_read(query)
    return [
        {
            "technology": r["technology_name"],
            "status": r["status"],
            "repository_logs": r["repository_logs"],
            "last_discovery_date": r["last_discovery_date"]
        } for r in rows
    ]


def get_dashboard_domains():
    query = """
        SELECT domain, urls_crawled, logs_validated, yield_score 
        FROM domain_performance 
        ORDER BY yield_score DESC, urls_crawled DESC
    """
    rows = execute_repo_read(query)
    return [
        {
            "domain": r["domain"],
            "urls_crawled": r["urls_crawled"],
            "logs_validated": r["logs_validated"],
            "yield_percent": round((r["yield_score"] or 0.0) * 100.0, 2)
        } for r in rows
    ]


def get_dashboard_notifications(severity=None):
    if not table_exists("notification_history"):
        return []
        
    if severity and severity.upper() in ["INFO", "WARNING", "ERROR", "CRITICAL"]:
        query = """
            SELECT id, timestamp, notification_type, recipient, subject, status, error_message, severity, job_id, technology
            FROM notification_history
            WHERE UPPER(severity) = ?
            ORDER BY id DESC LIMIT 20
        """
        params = (severity.upper(),)
    else:
        query = """
            SELECT id, timestamp, notification_type, recipient, subject, status, error_message, severity, job_id, technology
            FROM notification_history
            ORDER BY id DESC LIMIT 20
        """
        params = ()
        
    rows = execute_repo_read(query, params)
    return [
        {
            "id": r["id"],
            "timestamp": r["timestamp"],
            "notification_type": r["notification_type"],
            "recipient": r["recipient"],
            "subject": r["subject"],
            "status": r["status"],
            "error_message": r["error_message"],
            "severity": r["severity"] or "INFO",
            "job_id": r["job_id"] or "N/A",
            "technology": r["technology"] or "N/A"
        } for r in rows
    ]


def control_agent(action):
    from datetime import datetime
    if action == "pause":
        execute_repo_write("UPDATE agent_status SET status = 'paused' WHERE id = 1")
        log_agent_event("command_received", "Agent paused via dashboard command")
    elif action == "resume":
        execute_repo_write("UPDATE agent_status SET status = 'idle' WHERE id = 1")
        log_agent_event("command_received", "Agent resumed via dashboard command")
    elif action in ["run_discovery", "run_health_check"]:
        now_str = datetime.utcnow().isoformat() + "Z"
        execute_repo_insert(
            "INSERT INTO agent_control_queue (command, status, created_at) VALUES (?, ?, ?)",
            (action, "pending", now_str)
        )
        log_agent_event("command_received", f"Queued manual command: {action}")


def has_agent_event_feed_table():
    return table_exists("agent_event_feed")


def get_recent_agent_events(limit=50):
    query = f"SELECT id, timestamp, event_type, message FROM agent_event_feed ORDER BY id DESC LIMIT {limit}"
    rows = execute_repo_read(query)
    return [
        {
            "id": r["id"],
            "timestamp": r["timestamp"],
            "event_type": r["event_type"],
            "message": r["message"]
        } for r in rows
    ]


def get_agent_events_since(last_seen_id):
    query = "SELECT id, timestamp, event_type, message FROM agent_event_feed WHERE id > ? ORDER BY id ASC"
    rows = execute_repo_read(query, (last_seen_id,))
    return [
        {
            "id": r["id"],
            "timestamp": r["timestamp"],
            "event_type": r["event_type"],
            "message": r["message"]
        } for r in rows
    ]


# Helpers for discovery_agent.py
def upsert_technology_in_catalog(tech_name, category, vendor, discovery_source, log_queries, classification=None, confidence=None, accepted=None, reason=None):
    from datetime import datetime
    now = datetime.utcnow().isoformat() + "Z"
    queries_json = json.dumps(log_queries)
    status_str = "discovered" if accepted else "rejected"
    accepted_int = 1 if accepted else 0
    
    row = execute_repo_read("SELECT id FROM technology_catalog WHERE lower(technology_name) = lower(?)", (tech_name,))
    if row:
        tech_id = row[0]["id"]
        query = """
            UPDATE technology_catalog SET 
                last_seen = ?, 
                status = ?, 
                classification = ?, 
                technology_confidence = ?, 
                accepted = ?, 
                rejection_reason = ? 
            WHERE id = ?
        """
        params = (now, status_str, classification, confidence, accepted_int, reason, tech_id)
        execute_repo_write(query, params)
        return "updated"
    else:
        query = """
            INSERT INTO technology_catalog (
                technology_name, category, vendor, discovery_source, first_seen, last_seen, status, log_queries,
                classification, technology_confidence, accepted, rejection_reason
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        params = (tech_name, category, vendor, discovery_source, now, now, status_str, queries_json,
                  classification, confidence, accepted_int, reason)
        execute_repo_insert(query, params)
        return "inserted"


def get_all_catalog_technologies():
    query = "SELECT id, technology_name, category, vendor, discovery_source, log_queries FROM technology_catalog"
    rows = execute_repo_read(query)
    return [
        (r["id"], r["technology_name"], r["category"], r["vendor"], r["discovery_source"], r["log_queries"])
        for r in rows
    ]


# Helpers for batch_processor.py
def get_validated_logs_count_for_platform(platform_lower):
    rows = execute_repo_read("SELECT COUNT(*) FROM validated_logs WHERE lower(platform) = ?", (platform_lower,))
    if rows:
        return list(rows[0].values())[0]
    return 0


def get_accepted_technologies_catalog():
    query = "SELECT technology_name, category, vendor, discovery_source, first_seen, last_seen FROM technology_catalog WHERE accepted = 1 ORDER BY technology_name ASC"
    rows = execute_repo_read(query)
    return [(r["technology_name"], r["category"], r["vendor"], r["discovery_source"], r["first_seen"], r["last_seen"]) for r in rows]


def get_technology_coverage_details():
    query_cov = """
        SELECT technology_name, category, vendor, search_results, urls_crawled, 
               logs_extracted, logs_validated, repository_logs, unique_sources, last_discovery_date 
        FROM technology_coverage 
        ORDER BY technology_name ASC
    """
    rows = execute_repo_read(query_cov)
    res = []
    for r in rows:
        ed_row = execute_repo_read("SELECT error_diversity_score FROM technology_log_profile WHERE technology_name = ?", (r["technology_name"],))
        ed_score = ed_row[0]["error_diversity_score"] if ed_row else 0
        res.append({
            "technology_name": r["technology_name"],
            "category": r["category"],
            "vendor": r["vendor"],
            "search_results": r["search_results"],
            "urls_crawled": r["urls_crawled"],
            "logs_extracted": r["logs_extracted"],
            "logs_validated": r["logs_validated"],
            "repository_logs": r["repository_logs"],
            "unique_sources": r["unique_sources"],
            "last_discovery_date": r["last_discovery_date"],
            "error_diversity_score": ed_score
        })
    return res


def get_technology_log_profiles():
    query = """
        SELECT technology_name, total_logs, unique_sources, unique_error_codes, 
               unique_event_types, unique_components, error_diversity_score 
        FROM technology_log_profile 
        ORDER BY technology_name ASC
    """
    rows = execute_repo_read(query)
    return [
        (r["technology_name"], r["total_logs"], r["unique_sources"], r["unique_error_codes"],
         r["unique_event_types"], r["unique_components"], r["error_diversity_score"])
        for r in rows
    ]


def get_technology_audit_catalog():
    query = "SELECT technology_name, classification, technology_confidence, discovery_source, accepted, rejection_reason FROM technology_catalog ORDER BY technology_name ASC"
    rows = execute_repo_read(query)
    return [(r["technology_name"], r["classification"], r["technology_confidence"], r["discovery_source"], r["accepted"], r["rejection_reason"]) for r in rows]


# Helpers for autonomous_agent.py
def recover_stale_job_locks():
    query = "SELECT id, job_type, start_time FROM agent_job_history WHERE status = 'running'"
    rows = execute_repo_read(query)
    from datetime import datetime
    now = datetime.utcnow()
    recovered_jobs = []
    
    for job in rows:
        job_id, job_type, start_time_str = job["id"], job["job_type"], job["start_time"]
        t_str = start_time_str
        if t_str.endswith("Z"):
            t_str = t_str[:-1]
        try:
            start_time = datetime.fromisoformat(t_str)
            age_minutes = (now - start_time).total_seconds() / 60.0
            if age_minutes > 10.0:
                now_str = datetime.utcnow().isoformat() + "Z"
                execute_repo_write(
                    "UPDATE agent_job_history SET status = 'failed', end_time = ?, errors = 'Auto-recovered stale lock' WHERE id = ?",
                    (now_str, job_id)
                )
                
                status_rows = execute_repo_read("SELECT current_job FROM agent_status WHERE id = 1")
                if status_rows and status_rows[0]["current_job"] == job_type:
                    execute_repo_write("UPDATE agent_status SET status = 'idle', current_job = NULL, current_tech = NULL WHERE id = 1")
                
                recovered_jobs.append((job_id, job_type))
        except Exception as pe:
            print(f"Error checking age of job {job_id}: {pe}")
            
    return recovered_jobs


def acquire_job_lock(job_type):
    recover_stale_job_locks()
    running = execute_repo_read("SELECT id FROM agent_job_history WHERE job_type = ? AND status = 'running'", (job_type,))
    if running:
        return None
        
    from datetime import datetime
    now_str = datetime.utcnow().isoformat() + "Z"
    job_id = execute_repo_insert(
        "INSERT INTO agent_job_history (job_type, start_time, status, records_processed, errors) VALUES (?, ?, 'running', 0, '')",
        (job_type, now_str)
    )
    return job_id


def release_job_lock(job_id, status, records_processed=0, errors=""):
    from datetime import datetime
    now_str = datetime.utcnow().isoformat() + "Z"
    execute_repo_write(
        "UPDATE agent_job_history SET status = ?, end_time = ?, records_processed = ?, errors = ? WHERE id = ?",
        (status, now_str, records_processed, errors, job_id)
    )


def log_health_status(component, status, details):
    from datetime import datetime
    now_str = datetime.utcnow().isoformat() + "Z"
    execute_repo_insert(
        "INSERT INTO agent_health_history (timestamp, component, status, details) VALUES (?, ?, ?, ?)",
        (now_str, component, status, details)
    )


def is_agent_paused():
    rows = execute_repo_read("SELECT status FROM agent_status WHERE id = 1")
    if rows and rows[0]["status"] == "paused":
        return True
    return False


def update_agent_status_db(status=None, current_job=None, current_tech=None):
    from datetime import datetime
    now_str = datetime.utcnow().isoformat() + "Z"
    rows = execute_repo_read("SELECT COUNT(*) FROM agent_status WHERE id = 1")
    exists = list(rows[0].values())[0] > 0 if rows else False
    
    if exists:
        updates = []
        params = []
        if status is not None:
            updates.append("status = ?")
            params.append(status)
        if current_job is not None:
            if current_job == "":
                updates.append("current_job = NULL")
            else:
                updates.append("current_job = ?")
                params.append(current_job)
        if current_tech is not None:
            if current_tech == "":
                updates.append("current_tech = NULL")
            else:
                updates.append("current_tech = ?")
                params.append(current_tech)
        updates.append("last_active = ?")
        params.append(now_str)
        
        query = f"UPDATE agent_status SET {', '.join(updates)} WHERE id = 1"
        execute_repo_write(query, params)
    else:
        execute_repo_insert(
            "INSERT INTO agent_status (id, status, current_job, current_tech, last_active) VALUES (1, ?, ?, ?, ?)",
            (status or 'idle', current_job or None, current_tech or None, now_str)
        )


def update_scheduler_next_run_times(next_tech, next_log, next_health, next_daily):
    query = """
        UPDATE agent_status
        SET next_technology_discovery = ?,
            next_log_discovery = ?,
            next_health_check = ?,
            next_daily_report = ?
        WHERE id = 1
    """
    execute_repo_write(query, (next_tech, next_log, next_health, next_daily))


def set_agent_idle():
    rows = execute_repo_read("SELECT status FROM agent_status WHERE id = 1")
    current_status = rows[0]["status"] if rows else "idle"
    new_status = 'paused' if current_status == 'paused' else 'idle'
    update_agent_status_db(status=new_status, current_job="", current_tech="")
    update_agent_status_field(
        technologies_processed=0,
        technologies_total=0,
        current_query=0,
        total_queries=0,
        current_url=0,
        total_urls=0,
        cycle_urls_crawled=0,
        cycle_pages_classified=0,
        cycle_logs_extracted=0,
        cycle_logs_validated=0,
        cycle_logs_inserted=0,
        cycle_duplicates_skipped=0
    )


def log_runtime_metrics(tech_processed, urls_crawled, logs_extracted, logs_validated, logs_inserted, failures, urls_skipped, pages_log_rich, pages_low_value, insert_yield_pct):
    from datetime import datetime
    now_str = datetime.utcnow().isoformat() + "Z"
    query = """
        INSERT INTO agent_runtime_metrics (
            timestamp, technologies_processed, urls_crawled, logs_extracted, logs_validated, logs_inserted, failures,
            urls_skipped, pages_log_rich, pages_low_value, insert_yield_pct
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """
    execute_repo_insert(query, (now_str, tech_processed, urls_crawled, logs_extracted, logs_validated, logs_inserted, failures,
                               urls_skipped, pages_log_rich, pages_low_value, insert_yield_pct))


def enqueue_notification(notification_type, content):
    from datetime import datetime
    now_str = datetime.utcnow().isoformat() + "Z"
    query = "INSERT INTO notification_queue (created_at, notification_type, status, content) VALUES (?, ?, ?, ?)"
    return execute_repo_insert(query, (now_str, notification_type, "pending", content))


def mark_notification_sent(queue_id):
    execute_repo_write("UPDATE notification_queue SET status = 'sent' WHERE id = ?", (queue_id,))


def get_accepted_catalog_count():
    rows = execute_repo_read("SELECT COUNT(*) FROM technology_catalog WHERE accepted = 1")
    return list(rows[0].values())[0] if rows else 0


def get_new_accepted_technologies(start_utc):
    query = "SELECT technology_name, category, vendor FROM technology_catalog WHERE accepted = 1 AND first_seen >= ?"
    rows = execute_repo_read(query, (start_utc,))
    return [(r["technology_name"], r["category"], r["vendor"]) for r in rows]


def get_technology_discovery_details(t_name):
    query = "SELECT discovery_source, log_queries FROM technology_catalog WHERE technology_name = ?"
    rows = execute_repo_read(query, (t_name,))
    if rows:
        return rows[0]["discovery_source"], rows[0]["log_queries"]
    return None


def get_validated_logs_count():
    rows = execute_repo_read("SELECT COUNT(*) FROM validated_logs")
    return list(rows[0].values())[0] if rows else 0


def get_last_job_executions(job_type, limit=3):
    query = """
        SELECT id, status, start_time, end_time, errors FROM agent_job_history 
        WHERE job_type = ? AND status IN ('success', 'failed', 'warning') 
        ORDER BY id DESC LIMIT ?
    """
    rows = execute_repo_read(query, (job_type, limit))
    return [dict(r) for r in rows]


def get_agent_health_alert_state(job_type):
    query = "SELECT last_state, downtime_start FROM agent_health_alert_states WHERE job_type = ?"
    rows = execute_repo_read(query, (job_type,))
    if rows:
        return rows[0]
    return None


def insert_initial_health_alert_state(job_type, last_state, downtime_start):
    existing = execute_repo_read("SELECT 1 FROM agent_health_alert_states WHERE job_type = ?", (job_type,))
    if not existing:
        execute_repo_insert("INSERT INTO agent_health_alert_states (job_type, last_state, downtime_start) VALUES (?, ?, ?)",
                            (job_type, last_state, downtime_start))


def update_agent_health_alert_state(job_type, last_state, downtime_start):
    query = "UPDATE agent_health_alert_states SET last_state = ?, downtime_start = ? WHERE job_type = ?"
    execute_repo_write(query, (last_state, downtime_start, job_type))


def run_database_integrity_check():
    conn = get_repo_connection()
    is_postgres = "sqlite3" not in str(type(conn))
    cursor = conn.cursor()
    if is_postgres:
        try:
            cursor.execute("SELECT 1")
            res = "ok"
        except Exception as e:
            res = str(e)
    else:
        cursor.execute("PRAGMA integrity_check")
        res = cursor.fetchone()[0]
    cursor.close()
    conn.close()
    return res


def get_next_pending_agent_command():
    conn = get_repo_connection()
    is_postgres = "sqlite3" not in str(type(conn))
    cursor = conn.cursor()
    
    q_select = "SELECT id, command FROM agent_control_queue WHERE status = 'pending' ORDER BY id ASC LIMIT 1"
    q_update = "UPDATE agent_control_queue SET status = 'processing' WHERE id = ?"
    
    if is_postgres:
        q_update = q_update.replace("?", "%s")
        
    cursor.execute(q_select)
    row = cursor.fetchone()
    if row:
        cmd_id, command = row[0], row[1]
        cursor.execute(q_update, (cmd_id,))
        conn.commit()
        cursor.close()
        conn.close()
        return cmd_id, command
    else:
        cursor.close()
        conn.close()
        return None


def complete_agent_command(cmd_id):
    execute_repo_write("UPDATE agent_control_queue SET status = 'completed' WHERE id = ?", (cmd_id,))


def fail_agent_command(cmd_id):
    execute_repo_write("UPDATE agent_control_queue SET status = 'failed' WHERE id = ?", (cmd_id,))


def get_daily_report_data(yesterday_str):
    r_new_tech = execute_repo_read("SELECT COUNT(*) FROM technology_catalog WHERE first_seen >= ? AND accepted = 1", (yesterday_str,))
    new_techs = list(r_new_tech[0].values())[0] if r_new_tech else 0
    
    r_total_tech = execute_repo_read("SELECT COUNT(*) FROM technology_catalog WHERE accepted = 1")
    total_techs = list(r_total_tech[0].values())[0] if r_total_tech else 0
    
    r_new_logs = execute_repo_read("SELECT COUNT(*) FROM validated_logs WHERE discovered_at >= ?", (yesterday_str,))
    new_logs = list(r_new_logs[0].values())[0] if r_new_logs else 0
    
    r_total_logs = execute_repo_read("SELECT COUNT(*) FROM validated_logs")
    total_logs = list(r_total_logs[0].values())[0] if r_total_logs else 0
    
    job_stats_rows = execute_repo_read("SELECT status, COUNT(*) as cnt, job_type FROM agent_job_history WHERE start_time >= ? GROUP BY status, job_type", (yesterday_str,))
    job_stats = [(r["status"], r["cnt"], r["job_type"]) for r in job_stats_rows]
    
    failed_jobs_rows = execute_repo_read("SELECT job_type, errors, start_time FROM agent_job_history WHERE status = 'failed' AND start_time >= ? ORDER BY id DESC LIMIT 5", (yesterday_str,))
    failed_jobs_list = [f"- **{r['job_type']}** (at {r['start_time']}): {r['errors'][:150]}..." for r in failed_jobs_rows]
    
    unhealthy_alerts_rows = execute_repo_read("SELECT component, status, details, timestamp FROM agent_health_history WHERE status != 'healthy' AND timestamp >= ? ORDER BY id DESC LIMIT 10", (yesterday_str,))
    alert_lines = [f"- [{r['timestamp']}] Component: `{r['component']}` | Status: `{r['status']}` | Detail: {r['details']}" for r in unhealthy_alerts_rows]
    
    return {
        "new_techs": new_techs,
        "total_techs": total_techs,
        "new_logs": new_logs,
        "total_logs": total_logs,
        "job_stats": job_stats,
        "failed_jobs_list": failed_jobs_list,
        "alert_lines": alert_lines
    }


def get_crash_diagnostic_stats():
    status_row = execute_repo_read("SELECT status, current_job, current_tech FROM agent_status WHERE id = 1")
    last_running_job = "None"
    current_tech = "None"
    if status_row:
        last_running_job = status_row[0]["current_job"] or "None"
        current_tech = status_row[0]["current_tech"] or "None"
        
    succ_row = execute_repo_read("SELECT job_type FROM agent_job_history WHERE status = 'success' ORDER BY id DESC LIMIT 1")
    last_successful_job = "None"
    if succ_row:
        last_successful_job = succ_row[0]["job_type"]
        
    r_count = execute_repo_read("SELECT COUNT(*) FROM validated_logs")
    repo_count = list(r_count[0].values())[0] if r_count else 0
    
    return last_running_job, current_tech, last_successful_job, repo_count


class DatabaseError(Exception):
    pass


def insert_notification_history(timestamp, notification_type, recipient, subject, status, error_message, severity, job_id, technology):
    query = """
        INSERT INTO notification_history (timestamp, notification_type, recipient, subject, status, error_message, severity, job_id, technology)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """
    execute_repo_insert(query, (timestamp, notification_type, recipient, subject, status, error_message, severity, job_id, technology))


def insert_notification_queue(created_at, notification_type, status, content, next_retry):
    query = """
        INSERT INTO notification_queue (created_at, notification_type, status, content, retry_count, attempt_number, next_retry)
        VALUES (?, ?, ?, ?, 0, 0, ?)
    """
    execute_repo_insert(query, (created_at, notification_type, status, content, next_retry))


def mark_notification_failed_and_history(q_id, new_retry_count, new_attempt_number, now_str, next_retry_time, n_type, smtp_to, subject, err_msg, severity, job_id, technology):
    q_query = """
        UPDATE notification_queue
        SET status = 'Retrying', retry_count = ?, attempt_number = ?, last_attempt = ?, next_retry = ?
        WHERE id = ?
    """
    execute_repo_write(q_query, (new_retry_count, new_attempt_number, now_str, next_retry_time, q_id))
    
    h_query = """
        INSERT INTO notification_history (timestamp, notification_type, recipient, subject, status, error_message, severity, job_id, technology)
        VALUES (?, ?, ?, ?, 'Retrying', ?, ?, ?, ?)
    """
    execute_repo_insert(h_query, (now_str, n_type, smtp_to, f"[Retry Failed] {subject}", f"Attempt failed: {err_msg}", severity, job_id, technology))


def get_dashboard_stats_data():
    status_row = get_agent_status_metrics()
    if status_row:
        agent_status = status_row["status"]
        current_job = status_row["current_job"]
        current_tech = status_row["current_tech"]
        technologies_processed = status_row["technologies_processed"]
        technologies_total = status_row["technologies_total"]
        current_query = status_row["current_query"]
        total_queries = status_row["total_queries"]
        current_url = status_row["current_url"]
        total_urls = status_row["total_urls"]
        cycle_start_time = status_row["cycle_start_time"]
        cycle_urls_crawled = status_row["cycle_urls_crawled"]
        cycle_pages_classified = status_row["cycle_pages_classified"]
        cycle_logs_extracted = status_row["cycle_logs_extracted"]
        cycle_logs_validated = status_row["cycle_logs_validated"]
        cycle_logs_inserted = status_row["cycle_logs_inserted"]
        cycle_duplicates_skipped = status_row["cycle_duplicates_skipped"]
        current_phase = status_row["current_phase"]
        next_technology_discovery = status_row["next_technology_discovery"]
        next_log_discovery = status_row["next_log_discovery"]
        next_health_check = status_row["next_health_check"]
        next_daily_report = status_row["next_daily_report"]
    else:
        agent_status = "idle"
        current_job = None
        current_tech = None
        technologies_processed = 0
        technologies_total = 0
        current_query = 0
        total_queries = 0
        current_url = 0
        total_urls = 0
        cycle_start_time = None
        cycle_urls_crawled = 0
        cycle_pages_classified = 0
        cycle_logs_extracted = 0
        cycle_logs_validated = 0
        cycle_logs_inserted = 0
        cycle_duplicates_skipped = 0
        current_phase = "Idle"
        next_technology_discovery = None
        next_log_discovery = None
        next_health_check = None
        next_daily_report = None
        
    r_count = execute_repo_read("SELECT COUNT(*) FROM validated_logs")
    repo_log_count = list(r_count[0].values())[0] if r_count else 0
    
    r_techs = execute_repo_read("SELECT COUNT(*) FROM technology_catalog WHERE accepted = 1")
    techs_tracked = list(r_techs[0].values())[0] if r_techs else 0
    
    r_techs_with_logs = execute_repo_read("""
        SELECT COUNT(DISTINCT tc.technology_name)
        FROM technology_catalog tc
        JOIN validated_logs vl ON LOWER(tc.technology_name) = LOWER(vl.product_name)
        WHERE tc.accepted = 1
    """)
    techs_with_logs = list(r_techs_with_logs[0].values())[0] if r_techs_with_logs else 0
    
    techs_without_logs = max(0, techs_tracked - techs_with_logs)
    
    from datetime import datetime
    today_start = datetime.utcnow().date().isoformat() + "T00:00:00Z"
    r_metrics = execute_repo_read("""
        SELECT SUM(urls_crawled) as s_urls, SUM(logs_extracted) as s_ext, SUM(logs_validated) as s_val, SUM(logs_inserted) as s_ins
        FROM agent_runtime_metrics
        WHERE timestamp >= ?
    """, (today_start,))
    
    urls_crawled_today = 0
    logs_extracted_today = 0
    logs_validated_today = 0
    logs_inserted_today = 0
    if r_metrics and r_metrics[0]:
        urls_crawled_today = r_metrics[0].get("s_urls") or 0
        logs_extracted_today = r_metrics[0].get("s_ext") or 0
        logs_validated_today = r_metrics[0].get("s_val") or 0
        logs_inserted_today = r_metrics[0].get("s_ins") or 0
        
    insert_yield_pct = round(logs_inserted_today * 100.0 / urls_crawled_today, 2) if urls_crawled_today > 0 else 0.0
    
    r_dup = execute_repo_read("SELECT value FROM repository_metrics WHERE key = 'duplicates_skipped'")
    duplicates_skipped = list(r_dup[0].values())[0] if r_dup else 0
    
    return {
        "agent_status": agent_status,
        "current_job": current_job,
        "current_tech": current_tech,
        "repository_log_count": repo_log_count,
        "technologies_tracked": techs_tracked,
        "technologies_with_logs": techs_with_logs,
        "technologies_without_logs": techs_without_logs,
        "urls_crawled_today": urls_crawled_today,
        "logs_extracted_today": logs_extracted_today,
        "logs_validated_today": logs_validated_today,
        "logs_inserted_today": logs_inserted_today,
        "insert_yield_pct": insert_yield_pct,
        "duplicates_skipped": duplicates_skipped,
        "technologies_processed": technologies_processed,
        "technologies_total": technologies_total,
        "current_query": current_query,
        "total_queries": total_queries,
        "current_url": current_url,
        "total_urls": total_urls,
        "cycle_start_time": cycle_start_time,
        "cycle_urls_crawled": cycle_urls_crawled,
        "cycle_pages_classified": cycle_pages_classified,
        "cycle_logs_extracted": cycle_logs_extracted,
        "cycle_logs_validated": cycle_logs_validated,
        "cycle_logs_inserted": cycle_logs_inserted,
        "cycle_duplicates_skipped": cycle_duplicates_skipped,
        "current_phase": current_phase,
        "next_technology_discovery": next_technology_discovery,
        "next_log_discovery": next_log_discovery,
        "next_health_check": next_health_check,
        "next_daily_report": next_daily_report
    }





