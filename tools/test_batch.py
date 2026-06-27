import sys
import os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
import os
import csv
import zipfile
import tempfile
import unittest
from unittest.mock import patch, MagicMock
import openpyxl

# Force memory SQLite for tests
os.environ["SQLITE_DB_PATH"] = ":memory:"

from backend import db_manager
from backend import batch_processor

class TestBatchLogCollection(unittest.TestCase):
    def setUp(self):
        os.environ["SQLITE_DB_PATH"] = ":memory:"
        db_manager.init_db()

    @patch('backend.batch_processor.collect_logs_from_web')
    @patch('backend.batch_processor.parse_logs_with_llm')
    @patch('backend.batch_processor.parse_logs_with_regex')
    @patch('backend.batch_processor.validate_logs_with_claude')
    @patch('backend.batch_processor.vector_store')
    def test_invalid_excel_rows(self, mock_vs, mock_val, mock_regex, mock_llm, mock_web):
        # Create a mock CSV file with invalid entries (missing platform or version)
        fd, path = tempfile.mkstemp(suffix=".csv")
        os.close(fd)
        try:
            with open(path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["Platform", "Version", "Service", "Max Logs"])
                writer.writerow(["", "1.0", "svc", "5"]) # Missing platform
                writer.writerow(["Nginx", "", "svc", "5"]) # Missing version
                writer.writerow(["Apache", "2.4", "web", "10"]) # Valid
            
            rows = batch_processor.parse_uploaded_file(path)
            self.assertEqual(len(rows), 3)
            
            job_id = "test_job_invalid"
            db_manager.create_job(job_id, len(rows))
            for r in rows:
                db_manager.add_job_row(job_id, r.get("platform", ""), r.get("version", ""), r.get("service", ""), 5)
                
            mock_web.return_value = "raw text"
            mock_llm.return_value = [{"message": "msg1", "severity": "ERROR", "original_log": "log1"}]
            mock_regex.return_value = []
            mock_val.return_value = [{"message": "msg1", "severity": "ERROR", "original_log": "log1", "validation": {"valid": True, "reason": "ok"}}]
            
            batch_processor.process_batch_job(job_id)
            
            # Verify DB job stats
            job = db_manager.get_job(job_id)
            self.assertEqual(job["completed_rows"], 1)
            self.assertEqual(job["skipped_rows"], 2)
            self.assertEqual(job["failed_rows"], 0)
            self.assertEqual(job["remaining_rows"], 0)
            
            # Verify ZIP file contents
            zip_path = job["zip_path"]
            self.assertTrue(os.path.exists(zip_path))
            
            with zipfile.ZipFile(zip_path, "r") as z:
                namelist = z.namelist()
                self.assertIn("apache_web.txt", namelist)
                self.assertIn("Sources/apache_web_sources.txt", namelist)
                self.assertIn("diagnostic_report.xlsx", namelist)
                self.assertNotIn("collection_summary.txt", namelist)
                self.assertNotIn("not_found_report.txt", namelist)
        finally:
            if os.path.exists(path):
                os.remove(path)

    @patch('backend.batch_processor.collect_logs_from_web')
    @patch('backend.batch_processor.parse_logs_with_llm')
    @patch('backend.batch_processor.parse_logs_with_regex')
    @patch('backend.batch_processor.validate_logs_with_claude')
    def test_zip_generation(self, mock_val, mock_regex, mock_llm, mock_web):
        # Verify structure and existence of files in ZIP output
        job_id = "test_zip_job"
        db_manager.create_job(job_id, 2)
        db_manager.add_job_row(job_id, "Nginx", "1.25", "http", 5)
        db_manager.add_job_row(job_id, "Docker", "25", "daemon", 5)
        
        mock_web.return_value = "raw text"
        mock_llm.return_value = [{"message": "log line", "severity": "INFO", "original_log": "nginx_log_val"}]
        mock_regex.return_value = []
        mock_val.return_value = [{"message": "log line", "severity": "INFO", "original_log": "nginx_log_val", "validation": {"valid": True, "reason": "ok"}}]
        
        batch_processor.process_batch_job(job_id)
        job = db_manager.get_job(job_id)
        self.assertEqual(job["status"], "completed")
        self.assertTrue(os.path.exists(job["zip_path"]))
        
        with zipfile.ZipFile(job["zip_path"], "r") as z:
            namelist = z.namelist()
            self.assertIn("nginx_http.txt", namelist)
            self.assertIn("Sources/nginx_http_sources.txt", namelist)
            self.assertIn("docker_daemon.txt", namelist)
            self.assertIn("Sources/docker_daemon_sources.txt", namelist)
            self.assertIn("diagnostic_report.xlsx", namelist)
            self.assertNotIn("collection_summary.txt", namelist)
            self.assertNotIn("not_found_report.txt", namelist)
        

    @patch('backend.batch_processor.collect_logs_from_web')
    @patch('backend.batch_processor.parse_logs_with_llm')
    @patch('backend.batch_processor.parse_logs_with_regex')
    @patch('backend.batch_processor.validate_logs_with_claude')
    def test_restart_recovery(self, mock_val, mock_regex, mock_llm, mock_web):
        # Test restart recovery by simulating interrupted jobs in DB
        job_id = "interrupted_job"
        db_manager.create_job(job_id, 2)
        db_manager.add_job_row(job_id, "Ubuntu", "20.04", "syslog", 10)
        db_manager.add_job_row(job_id, "CentOS", "7", "audit", 5)
        
        db_manager.update_job_status(job_id, status="pending")
        
        mock_web.return_value = "some text"
        mock_llm.return_value = [{"message": "recovered log", "severity": "WARN", "original_log": "recovered_log_text"}]
        mock_regex.return_value = []
        mock_val.return_value = [{"message": "recovered log", "severity": "WARN", "original_log": "recovered_log_text", "validation": {"valid": True, "reason": "ok"}}]
        
        with patch('threading.Thread') as mock_thread:
            def run_sync(target, args):
                target(*args)
            mock_thread.side_effect = lambda target, args: MagicMock(start=lambda: run_sync(target, args))
            
            batch_processor.resume_unfinished_jobs()
            
        job = db_manager.get_job(job_id)
        self.assertEqual(job["status"], "completed")
        self.assertEqual(job["completed_rows"], 2)

    @patch('backend.batch_processor.collect_logs_from_web')
    @patch('backend.batch_processor.parse_logs_with_llm')
    @patch('backend.batch_processor.parse_logs_with_regex')
    @patch('backend.batch_processor.validate_logs_with_claude')
    def test_duplicate_platform_rows(self, mock_val, mock_regex, mock_llm, mock_web):
        # Verify rows with duplicate platform/version combinations aggregate correctly.
        job_id = "test_duplicates"
        db_manager.create_job(job_id, 2)
        db_manager.add_job_row(job_id, "Nginx", "1.25", "upstream", 5)
        db_manager.add_job_row(job_id, "Nginx", "1.25", "upstream", 5)
        
        mock_web.return_value = "log text"
        mock_llm.side_effect = [
            [{"message": "log msg 1", "severity": "INFO", "original_log": "nginx_msg_1"}],
            [{"message": "log msg 2", "severity": "INFO", "original_log": "nginx_msg_2"}]
        ]
        mock_regex.return_value = []
        mock_val.side_effect = [
            [{"message": "log msg 1", "severity": "INFO", "original_log": "nginx_msg_1", "validation": {"valid": True, "reason": "ok"}}],
            [{"message": "log msg 2", "severity": "INFO", "original_log": "nginx_msg_2", "validation": {"valid": True, "reason": "ok"}}]
        ]
        
        batch_processor.process_batch_job(job_id)
        job = db_manager.get_job(job_id)
        
        with zipfile.ZipFile(job["zip_path"], "r") as z:
            namelist = z.namelist()
            self.assertIn("nginx_upstream.txt", namelist)
            self.assertIn("nginx_upstream_2.txt", namelist)
            self.assertIn("Sources/nginx_upstream_sources.txt", namelist)
            self.assertIn("Sources/nginx_upstream_2_sources.txt", namelist)
            
            self.assertEqual(z.read("nginx_upstream.txt").decode("utf-8").strip(), "nginx_msg_1")
            self.assertEqual(z.read("nginx_upstream_2.txt").decode("utf-8").strip(), "nginx_msg_2")

    def test_large_excel_files(self):
        # Test parsing of a large excel file (50+ rows)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Platform", "Version", "Service", "Max Logs"])
        for i in range(50):
            ws.append([f"Platform_{i}", f"v{i}", f"service_{i}", 10])
            
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            wb.save(path)
            rows = batch_processor.parse_uploaded_file(path)
            self.assertEqual(len(rows), 50)
            self.assertEqual(rows[0]["platform"], "Platform_0")
            self.assertEqual(rows[49]["version"], "v49")
        finally:
            if os.path.exists(path):
                os.remove(path)

    def test_product_name_extraction(self):
        # 1. Year range pattern at the end
        p, v = batch_processor.extract_product_and_version("Amazon SNS (2010-present)")
        self.assertEqual(p, "Amazon SNS")
        self.assertEqual(v, "2010-present")
        
        p, v = batch_processor.extract_product_and_version("Amazon SQS (2006-present)")
        self.assertEqual(p, "Amazon SQS")
        self.assertEqual(v, "2006-present")

        p, v = batch_processor.extract_product_and_version("API Gateway REST (2015-present)")
        self.assertEqual(p, "API Gateway REST")
        self.assertEqual(v, "2015-present")

        # 2. Year range with fixed years
        p, v = batch_processor.extract_product_and_version("Some Product (2006-2012)")
        self.assertEqual(p, "Some Product")
        self.assertEqual(v, "2006-2012")

        # 3. No year range - fallback to full string as version
        p, v = batch_processor.extract_product_and_version("1.25")
        self.assertEqual(p, "")
        self.assertEqual(v, "1.25")

        p, v = batch_processor.extract_product_and_version("v2.4-alpine")
        self.assertEqual(p, "")
        self.assertEqual(v, "v2.4-alpine")

    @patch('backend.crawler.search_log_sources')
    def test_scraped_search_queries_with_product_name(self, mock_search):
        mock_search.return_value = []
        
        # 1. With product_name
        from backend.crawler import collect_logs_from_web
        collect_logs_from_web("AWS", "2010-present", "Delivery Log", count=1, product_name="Amazon SNS")
        
        # AWS + Amazon SNS has 2 base terms. 3 stages: (3 + 4 + 5) * 2 = 24 queries + 3 site queries = 27
        self.assertEqual(mock_search.call_count, 27)
        queries = [call[0][0] for call in mock_search.call_args_list]
        
        # Verify some representative queries are generated
        self.assertIn('AWS SNS 2010-present Delivery Log "actual log" site:github.com/issues', queries)
        self.assertIn('Amazon SNS 2010-present Delivery Log "actual log" site:github.com/issues', queries)
        self.assertIn('AWS SNS 2010-present Delivery Log "stack trace" site:github.com', queries)
        self.assertIn('Amazon SNS 2010-present Delivery Log "bug report" log', queries)
        
        mock_search.reset_mock()
        
        # 2. Without product_name
        collect_logs_from_web("nginx", "1.25", "upstream", count=1)
        # Nginx has 1 base term. 3 stages: 3 + 4 + 5 = 12 queries + 2 site queries = 14
        self.assertEqual(mock_search.call_count, 14)
        queries_no_prod = [call[0][0] for call in mock_search.call_args_list]
        self.assertIn('nginx 1.25 upstream "actual log" site:github.com/issues', queries_no_prod)

    def test_sanitize_filename_aws_sns(self):
        fn = batch_processor.sanitize_filename("AWS", "Delivery Log", product_name="Amazon SNS")
        self.assertEqual(fn, "aws_amazon_sns_delivery_log.txt")
        
        # Verify extraction and integration
        prod_name, version = batch_processor.extract_product_and_version("Amazon SNS (2010-present)")
        self.assertEqual(prod_name, "Amazon SNS")
        self.assertEqual(version, "2010-present")
        
        fn_full = batch_processor.sanitize_filename("AWS", "Delivery Log", product_name=prod_name)
        self.assertEqual(fn_full, "aws_amazon_sns_delivery_log.txt")

    @patch('backend.batch_processor.collect_logs_from_web')
    @patch('backend.batch_processor.parse_logs_with_llm')
    @patch('backend.batch_processor.parse_logs_with_regex')
    @patch('backend.batch_processor.validate_logs_with_claude')
    def test_aws_sns_integration(self, mock_val, mock_regex, mock_llm, mock_web):
        job_id = "test_aws_sns_job"
        db_manager.create_job(job_id, 1)
        db_manager.add_job_row(job_id, "AWS", "2010-present", "Delivery Log", 5, product_name="Amazon SNS")
        
        mock_web.return_value = "log text"
        mock_llm.return_value = [{"message": "aws log line", "severity": "INFO", "original_log": "aws_log_text"}]
        mock_regex.return_value = []
        mock_val.return_value = [{"message": "aws log line", "severity": "INFO", "original_log": "aws_log_text", "validation": {"valid": True, "reason": "ok"}}]
        
        batch_processor.process_batch_job(job_id)
        job = db_manager.get_job(job_id)
        self.assertEqual(job["status"], "completed")
        
        with zipfile.ZipFile(job["zip_path"], "r") as z:
            namelist = z.namelist()
            self.assertIn("aws_amazon_sns_delivery_log.txt", namelist)
            self.assertIn("Sources/aws_amazon_sns_delivery_log_sources.txt", namelist)

    @patch('backend.batch_processor.collect_logs_from_web')
    @patch('backend.batch_processor.parse_logs_with_llm')
    @patch('backend.batch_processor.parse_logs_with_regex')
    @patch('backend.batch_processor.validate_logs_with_claude')
    def test_50_rows_generation(self, mock_val, mock_regex, mock_llm, mock_web):
        job_id = "test_50_rows_job"
        db_manager.create_job(job_id, 50)
        
        for i in range(50):
            db_manager.add_job_row(job_id, "Nginx", "1.25", "upstream", 5)
            
        mock_web.return_value = "log text"
        mock_regex.return_value = []
        
        llm_outputs = []
        val_outputs = []
        for i in range(50):
            if i < 30:
                llm_outputs.append([{"message": f"log msg {i}", "severity": "INFO", "original_log": f"log_content_{i}"}])
                val_outputs.append([{"message": f"log msg {i}", "severity": "INFO", "original_log": f"log_content_{i}", "validation": {"valid": True, "reason": "ok"}}])
            else:
                llm_outputs.append([])
                val_outputs.append([])
                
        mock_llm.side_effect = llm_outputs
        mock_val.side_effect = val_outputs
        
        batch_processor.process_batch_job(job_id)
        job = db_manager.get_job(job_id)
        self.assertEqual(job["status"], "completed")
        self.assertEqual(job["completed_rows"], 50)
        
        with zipfile.ZipFile(job["zip_path"], "r") as z:
            namelist = z.namelist()
            
            self.assertNotIn("collection_summary.txt", namelist)
            self.assertNotIn("not_found_report.txt", namelist)
            self.assertIn("nginx_upstream.txt", namelist)
            self.assertIn("Sources/nginx_upstream_sources.txt", namelist)
            
            for i in range(2, 31):
                self.assertIn(f"nginx_upstream_{i}.txt", namelist)
                self.assertIn(f"Sources/nginx_upstream_{i}_sources.txt", namelist)
                
            for i in range(31, 51):
                self.assertNotIn(f"nginx_upstream_{i}.txt", namelist)
                
            import tempfile
            import shutil
            temp_dir = tempfile.mkdtemp()
            xlsx_path = z.extract("diagnostic_report.xlsx", temp_dir)
            try:
                wb = openpyxl.load_workbook(xlsx_path)
                self.assertIn("SUMMARY", wb.sheetnames)
                ws_summary = wb["SUMMARY"]
                rows = list(ws_summary.iter_rows(values_only=True))
                # 1 header + 50 data rows = 51 rows total
                self.assertEqual(len(rows), 51)
                
                # First 30 rows (index 1 to 30) have Validated Logs Count = 1
                for r in rows[1:31]:
                    self.assertEqual(r[6], 1)
                    
                # Last 20 rows (index 31 to 50) have Validated Logs Count = 0
                for r in rows[31:51]:
                    self.assertEqual(r[6], 0)
                    self.assertEqual(r[8], "No search results")
            finally:
                shutil.rmtree(temp_dir)

    @patch('backend.batch_processor.validate_logs_with_claude')
    @patch('backend.batch_processor.parse_logs_with_regex')
    @patch('backend.batch_processor.parse_logs_with_llm')
    @patch('backend.batch_processor.collect_logs_from_web')
    def test_diagnostic_mode(self, mock_collect, mock_llm, mock_regex, mock_val):
        from backend.crawler import ScrapedText
        
        # Prepare mock ScrapedText results
        # Row 1: Validated log found
        res1 = ScrapedText("2026-06-17 12:00:00 [INFO] app started")
        res1.queries = ["query1"]
        res1.raw_results_count = 1
        res1.visited_urls = ["https://github.com/nginx"]
        res1.search_results_ranked = [{"rank": 1, "url": "https://github.com/nginx", "domain": "github.com"}]
        res1.url_info_map = {
            "https://github.com/nginx": {
                "Platform": "Nginx",
                "Product": "",
                "Log Type": "upstream",
                "Search Query Used": "query1",
                "Search Position": 1,
                "Source Type": "GITHUB_ISSUE",
                "URL": "https://github.com/nginx",
                "Search Rank": 1,
                "Crawled": "Yes",
                "Logs Extracted": 1,
                "Logs Validated": 1,
                "Logs Rejected": 0,
                "Status": "validated",
                "Reason": "",
                "Failure Stage": ""
            }
        }
        
        # Row 2: No search results
        res2 = ScrapedText("")
        res2.queries = ["query2"]
        res2.raw_results_count = 0
        res2.visited_urls = []
        res2.search_results_ranked = []
        res2.url_info_map = {}
        
        # Row 3: Blocklisted/Source filtered
        res3 = ScrapedText("")
        res3.queries = ["query3"]
        res3.raw_results_count = 2
        res3.visited_urls = ["https://medium.com/foo", "https://twitter.com/bar"]
        res3.search_results_ranked = [
            {"rank": 1, "url": "https://medium.com/foo", "domain": "medium.com"},
            {"rank": 2, "url": "https://twitter.com/bar", "domain": "twitter.com"}
        ]
        res3.url_info_map = {
            "https://medium.com/foo": {
                "Platform": "Docker",
                "Product": "",
                "Log Type": "daemon",
                "Search Query Used": "query3",
                "Search Position": 1,
                "Source Type": "BLOG",
                "URL": "https://medium.com/foo",
                "Search Rank": 1,
                "Crawled": "No",
                "Logs Extracted": 0,
                "Logs Validated": 0,
                "Logs Rejected": 0,
                "Status": "access_denied",
                "Reason": "Domain blocklisted",
                "Failure Stage": "search_filtering"
            },
            "https://twitter.com/bar": {
                "Platform": "Docker",
                "Product": "",
                "Log Type": "daemon",
                "Search Query Used": "query3",
                "Search Position": 2,
                "Source Type": "BLOG",
                "URL": "https://twitter.com/bar",
                "Search Rank": 2,
                "Crawled": "No",
                "Logs Extracted": 0,
                "Logs Validated": 0,
                "Logs Rejected": 0,
                "Status": "access_denied",
                "Reason": "Domain blocklisted",
                "Failure Stage": "search_filtering"
            }
        }
        
        # Row 4: Logs extracted but validation rejected all
        res4 = ScrapedText("some garbage text that looks like logs")
        res4.queries = ["query4"]
        res4.raw_results_count = 1
        res4.visited_urls = ["https://github.com/garbage"]
        res4.search_results_ranked = [{"rank": 1, "url": "https://github.com/garbage", "domain": "github.com"}]
        res4.url_info_map = {
            "https://github.com/garbage": {
                "Platform": "Ubuntu",
                "Product": "",
                "Log Type": "syslog",
                "Search Query Used": "query4",
                "Search Position": 1,
                "Source Type": "GITHUB_ISSUE",
                "URL": "https://github.com/garbage",
                "Search Rank": 1,
                "Crawled": "Yes",
                "Logs Extracted": 1,
                "Logs Validated": 0,
                "Logs Rejected": 1,
                "Status": "no_logs_found",
                "Reason": "Logs extracted but validation rejected all",
                "Failure Stage": "validation"
            }
        }
        
        mock_collect.side_effect = [res1, res2, res3, res4]
        
        # LLM parsing
        mock_llm.side_effect = [
            [{"message": "app started", "severity": "INFO", "original_log": "2026-06-17 12:00:00 [INFO] app started"}], # Row 1
            [{"message": "human explanation description", "severity": "INFO", "original_log": "human explanation description"}] # Row 4
        ]
        mock_regex.return_value = []
        
        # Validation
        mock_val.side_effect = [
            # Row 1 (valid)
            [{"message": "app started", "severity": "INFO", "original_log": "2026-06-17 12:00:00 [INFO] app started", "validation": {"valid": True, "reason": "conforms to Path A"}}],
            # Row 4 (invalid)
            [{"message": "human explanation description", "severity": "INFO", "original_log": "human explanation description", "validation": {"valid": False, "reason": "Rejected prose sentence"}}]
        ]
        
        # Create a job
        job_id = "test_diag_job"
        db_manager.create_job(job_id, 4)
        
        # Row 1 (succeeds)
        db_manager.add_job_row(job_id, "Nginx", "1.25", "upstream", product_name="", max_logs=1)
        # Row 2 (no search results)
        db_manager.add_job_row(job_id, "Apache", "2.4", "web", product_name="", max_logs=1)
        # Row 3 (source filtered)
        db_manager.add_job_row(job_id, "Docker", "25", "daemon", product_name="", max_logs=1)
        # Row 4 (validation rejected all)
        db_manager.add_job_row(job_id, "Ubuntu", "20.04", "syslog", product_name="", max_logs=1)
        
        # Process job
        batch_processor.process_batch_job(job_id)
        
        job = db_manager.get_job(job_id)
        self.assertEqual(job["status"], "completed")
        
        # Verify ZIP contains reports
        with zipfile.ZipFile(job["zip_path"], "r") as z:
            namelist = z.namelist()
            self.assertNotIn("diagnostic_report.csv", namelist)
            self.assertIn("diagnostic_report.xlsx", namelist)
            self.assertNotIn("diagnostic_report.txt", namelist)
            
            # Read CSV report from disk
            job_work_dir = os.path.join(batch_processor.BASE_OUTPUT_DIR, f"job_{job_id}")
            csv_path = os.path.join(job_work_dir, "diagnostic_report.csv")
            with open(csv_path, "r", encoding="utf-8") as f:
                csv_content = f.read()
            reader = csv.reader(csv_content.splitlines())
            rows = list(reader)
            
            self.assertEqual(len(rows), 5) # header + 4 rows
            
            # Row 1: Validated Logs Count = 1, Failure Reason = ""
            self.assertEqual(rows[1][0], "Nginx")
            self.assertEqual(rows[1][6], "1") # Validated Logs Count
            self.assertEqual(rows[1][8], "") # Failure Reason
            
            # Verify Excel Report
            import tempfile
            import shutil
            temp_dir = tempfile.mkdtemp()
            xlsx_path = z.extract("diagnostic_report.xlsx", temp_dir)
            try:
                wb = openpyxl.load_workbook(xlsx_path)
                self.assertIn("SUMMARY", wb.sheetnames)
                self.assertIn("UNPROCESSED_SOURCES", wb.sheetnames)
                self.assertIn("SOURCE_AUDIT", wb.sheetnames)
                
                # Check UNPROCESSED_SOURCES columns
                ws_unprocessed = wb["UNPROCESSED_SOURCES"]
                unprocessed_rows = list(ws_unprocessed.iter_rows(values_only=True))
                self.assertEqual(unprocessed_rows[0][0], "Platform")
                self.assertEqual(unprocessed_rows[0][3], "Search Query Used")
                self.assertEqual(unprocessed_rows[0][4], "URL")
                self.assertEqual(unprocessed_rows[0][5], "Title")
                self.assertEqual(unprocessed_rows[0][6], "Source Type")
                self.assertEqual(unprocessed_rows[0][9], "Failure Stage")
                
                unprocessed_platforms = [r[0] for r in unprocessed_rows[1:]]
                self.assertIn("Docker", unprocessed_platforms)
                self.assertIn("Ubuntu", unprocessed_platforms)
                
                # Check SOURCE_AUDIT columns
                ws_audit = wb["SOURCE_AUDIT"]
                audit_rows = list(ws_audit.iter_rows(values_only=True))
                self.assertEqual(audit_rows[0][0], "Platform")
                self.assertEqual(audit_rows[0][3], "Search Query Used")
                self.assertEqual(audit_rows[0][4], "URL")
                self.assertEqual(audit_rows[0][5], "Search Position")
                self.assertEqual(audit_rows[0][6], "Source Type")
                self.assertEqual(audit_rows[0][11], "Logs Rejected")
                self.assertEqual(audit_rows[0][12], "Final Status")
                
                audit_platforms = [r[0] for r in audit_rows[1:]]
                self.assertIn("Nginx", audit_platforms)
                self.assertIn("Docker", audit_platforms)
                self.assertIn("Ubuntu", audit_platforms)
                
            finally:
                shutil.rmtree(temp_dir)
            
            # Read text report from disk and verify sample rejected logs and search result ranking
            txt_path = os.path.join(job_work_dir, "diagnostic_report.txt")
            with open(txt_path, "r", encoding="utf-8") as f:
                txt_content = f.read()
            self.assertIn("Sample Rejected Logs:", txt_content)
            self.assertIn("Search Result Ranking:", txt_content)
            self.assertIn("Rank 1 | github.com | https://github.com/garbage", txt_content)
            self.assertIn("Reason: Rejected prose sentence", txt_content)
            self.assertIn("Validation Format Detected:", txt_content)
            
            # Row 2: No search results
            self.assertEqual(rows[2][0], "Apache")
            self.assertEqual(rows[2][6], "0")
            self.assertEqual(rows[2][8], "No search results")
            
            # Row 3: Source filtering
            self.assertEqual(rows[3][0], "Docker")
            self.assertEqual(rows[3][6], "0")
            self.assertEqual(rows[3][8], "Source filtering removed all candidates")
            
            # Row 4: Validation rejected all
            self.assertEqual(rows[4][0], "Ubuntu")
            self.assertEqual(rows[4][6], "0")
            self.assertEqual(rows[4][8], "Logs extracted but validation rejected all")

    def test_dict_message_handling(self):
        # Input has dict message and dict original_log
        from backend.extractor import check_log_nature_detail
        dict_log = {
            "timestamp": "2026-06-19T10:00:00Z",
            "severity": "ERROR",
            "message": {"eventName": "CreateUser"},
            "original_log": {"eventSource": "iam.amazonaws.com"},
            "source_url": "https://example.com"
        }
        
        # Verify check_log_nature_detail doesn't crash on dict
        is_genuine, reason, indicators = check_log_nature_detail(dict_log["message"])
        self.assertFalse(is_genuine) # A dict eventName alone shouldn't pass log validation but should not throw exception
        
        # Verify deterministic validation doesn't crash
        from backend.validator import deterministic_validation
        is_valid, reason, confidence = deterministic_validation(dict_log, "some text", "aws")
        self.assertFalse(is_valid)

    def test_cloudtrail_json_log(self):
        # Input is a valid CloudTrail log record represented as dicts
        from backend.extractor import check_log_nature_detail
        
        cloudtrail_event = {
            "eventVersion": "1.08",
            "userIdentity": {"type": "IAMUser", "userName": "Alice"},
            "eventTime": "2026-06-19T10:00:00Z",
            "eventSource": "iam.amazonaws.com",
            "eventName": "CreateUser",
            "awsRegion": "us-east-1",
            "sourceIPAddress": "192.0.2.1",
            "userAgent": "console.amazonaws.com"
        }
        
        # Verify it processes safely (doesn't raise strip AttributeError)
        is_genuine, reason, indicators = check_log_nature_detail(cloudtrail_event)
        self.assertIsNotNone(is_genuine)

    @patch('backend.batch_processor.collect_logs_from_web')
    def test_reporting_survives_validation_failure(self, mock_web):
        # Force extraction failure on row 1, but confirm UNPROCESSED_SOURCES and SOURCE_AUDIT are populated
        from backend.crawler import ScrapedText
        from backend import db_manager, batch_processor
        import tempfile
        import shutil
        import zipfile
        
        job_id = "test_survive_fail"
        db_manager.create_job(job_id, 1)
        db_manager.add_job_row(job_id, "Nginx", "1.25", "upstream", product_name="", max_logs=1)
        
        # Create a mock scraped_text object that has extraction_error set, but has populated url_info_map
        scraped_text = ScrapedText("some raw text")
        scraped_text.queries = ["query1"]
        scraped_text.raw_results_count = 1
        scraped_text.visited_urls = ["https://example.com/failed"]
        scraped_text.search_results_ranked = [{"rank": 1, "url": "https://example.com/failed", "domain": "example.com", "source_rank": 4, "title": "failed source"}]
        scraped_text.extracted_logs = []
        scraped_text.extraction_error = "Mocked LLM parsing failure (e.g. Rate Limit exceeded)"
        scraped_text.url_info_map = {
            "https://example.com/failed": {
                "Platform": "Nginx",
                "Product": "",
                "Log Type": "upstream",
                "Search Query Used": "query1",
                "URL": "https://example.com/failed",
                "Title": "failed source",
                "Source Type": "BLOG",
                "Source Rank": 4,
                "Search Rank": 1,
                "Search Position": 1,
                "Crawled": "Yes",
                "Logs Extracted": 0,
                "Logs Validated": 0,
                "Logs Rejected": 0,
                "Status": "no_logs_found",
                "Reason": "Mocked LLM parsing failure (e.g. Rate Limit exceeded)",
                "Failure Stage": "extraction"
            }
        }
        
        mock_web.return_value = scraped_text
        
        # Process the batch job. This should complete but the row will fail due to extraction_error
        batch_processor.process_batch_job(job_id)
        
        job = db_manager.get_job(job_id)
        self.assertEqual(job["status"], "completed")
        self.assertEqual(job["failed_rows"], 1)
        
        # Verify ZIP contains the XLSX report with secondary worksheets populated
        with zipfile.ZipFile(job["zip_path"], "r") as z:
            namelist = z.namelist()
            self.assertIn("diagnostic_report.xlsx", namelist)
            
            temp_dir = tempfile.mkdtemp()
            xlsx_path = z.extract("diagnostic_report.xlsx", temp_dir)
            try:
                wb = openpyxl.load_workbook(xlsx_path)
                self.assertIn("UNPROCESSED_SOURCES", wb.sheetnames)
                self.assertIn("SOURCE_AUDIT", wb.sheetnames)
                
                # Verify UNPROCESSED_SOURCES has the URL row
                ws_unprocessed = wb["UNPROCESSED_SOURCES"]
                rows = list(ws_unprocessed.iter_rows(values_only=True))
                self.assertEqual(len(rows), 2) # header + 1 row
                self.assertEqual(rows[1][4], "https://example.com/failed")
                self.assertEqual(rows[1][9], "extraction")
                self.assertEqual(rows[1][10], "Mocked LLM parsing failure (e.g. Rate Limit exceeded)")
                
                # Verify SOURCE_AUDIT has the URL row
                ws_audit = wb["SOURCE_AUDIT"]
                audit_rows = list(ws_audit.iter_rows(values_only=True))
                self.assertEqual(len(audit_rows), 2) # header + 1 row
                self.assertEqual(audit_rows[1][4], "https://example.com/failed")
                
            finally:
                shutil.rmtree(temp_dir)

    def test_duckduckgo_search_provider(self):
        from backend.search_providers import DuckDuckGoProvider
        import backend.search_providers as sp_module

        fake_ddgs_results = [
            {"title": "Official documentation logs",
             "href": "https://example.com/logs/doc",
             "body": "This is an example log snippet."},
        ]

        class FakeDDGS:
            def __init__(self, timeout=None): pass
            def __enter__(self): return self
            def __exit__(self, *a): pass
            def text(self, query, max_results=7):
                return fake_ddgs_results[:max_results]

        original_ddgs = sp_module.DDGS
        original_available = sp_module._DDGS_AVAILABLE
        try:
            sp_module.DDGS = FakeDDGS
            sp_module._DDGS_AVAILABLE = True

            provider = DuckDuckGoProvider()
            res = provider.search("some query", max_results=1)

            self.assertEqual(res["status_code"], 200)
            self.assertEqual(res["error"], "success")
            self.assertEqual(len(res["results"]), 1)
            self.assertEqual(res["results"][0]["title"], "Official documentation logs")
            self.assertEqual(res["results"][0]["url"], "https://example.com/logs/doc")
            self.assertEqual(res["results"][0]["snippet"], "This is an example log snippet.")
        finally:
            sp_module.DDGS = original_ddgs
            sp_module._DDGS_AVAILABLE = original_available

        # DDGS returning empty should skip HTML fallback and return fast
        class EmptyDDGS:
            def __init__(self, timeout=None): pass
            def __enter__(self): return self
            def __exit__(self, *a): pass
            def text(self, query, max_results=7): return []

        try:
            sp_module.DDGS = EmptyDDGS
            sp_module._DDGS_AVAILABLE = True

            provider2 = DuckDuckGoProvider()
            res2 = provider2.search("some obscure query", max_results=5)

            self.assertEqual(res2["results"], [])
            self.assertIn("DDGS returned empty", res2["error"])
            self.assertLess(res2["duration"], 2.0,
                            "Should return fast without hitting HTML scraper")
        finally:
            sp_module.DDGS = original_ddgs
            sp_module._DDGS_AVAILABLE = original_available

    @patch('backend.search_providers.DuckDuckGoProvider.search')
    @patch('backend.search_providers.BingProvider.search')
    @patch('backend.search_providers.BraveProvider.search')
    @patch('backend.search_providers.YahooProvider.search')
    @patch('backend.search_providers.AOLProvider.search')
    def test_provider_fallback_mechanism(self, mock_aol_search, mock_yahoo_search, mock_brave_search, mock_bing_search, mock_ddg_search):
        from backend.crawler import search_log_sources, clear_thread_diagnostics, get_thread_diagnostics
        
        clear_thread_diagnostics()
        
        # DDG, Bing, Brave, Yahoo fail (0 results), AOL succeeds (1 result)
        mock_ddg_search.return_value = {"results": [], "status_code": 500, "error": "Blocked", "duration": 0.1}
        mock_bing_search.return_value = {"results": [], "status_code": 200, "error": "Captcha", "duration": 0.1}
        mock_brave_search.return_value = {"results": [], "status_code": 429, "error": "Rate limited", "duration": 0.1}
        mock_yahoo_search.return_value = {"results": [], "status_code": 500, "error": "Blocked", "duration": 0.1}
        mock_aol_search.return_value = {
            "results": [{"title": "aol title", "url": "https://example.com/aol", "snippet": "aol snippet"}],
            "status_code": 200,
            "error": "success",
            "duration": 0.1
        }
        
        results = search_log_sources("some query", max_results=5)
        self.assertEqual(len(results), 1)
        self.assertEqual(results[0]["url"], "https://example.com/aol")
        
        # Check that diagnostics are recorded for all providers tried
        diags = get_thread_diagnostics()
        self.assertEqual(len(diags), 5)
        self.assertEqual(diags[0]["Provider"], "DuckDuckGo")
        self.assertEqual(diags[0]["HTTP Status"], 500)
        self.assertEqual(diags[4]["Provider"], "AOL")
        self.assertEqual(diags[4]["HTTP Status"], 200)

    @patch('backend.crawler.search_log_sources')
    @patch('backend.crawler.scrape_url_content')
    def test_direct_source_discovery_fallback(self, mock_scrape, mock_search):
        from backend.crawler import collect_logs_from_web
        from backend.crawler import ScrapedContent
        
        # Search returns 0 results for both standard and site queries
        mock_search.return_value = []
        mock_scrape.return_value = ScrapedContent("some direct content log lines")
        
        res = collect_logs_from_web("Docker", version="25", service="daemon", count=1)
        
        # Verify it fell back to direct URLs
        self.assertTrue(len(res.visited_urls) > 0)
        self.assertTrue(any("docker" in u or "github" in u or "stackoverflow" in u for u in res.visited_urls))
        self.assertEqual(res.search_diagnostics[-1]["Provider"], "DirectDiscovery")

    @patch('backend.batch_processor.collect_logs_from_web')
    def test_search_diagnostics_report(self, mock_web):
        from backend.crawler import ScrapedText
        from backend import db_manager, batch_processor
        import tempfile
        import shutil
        import zipfile
        import openpyxl
        
        job_id = "test_search_diag_job"
        db_manager.create_job(job_id, 1)
        db_manager.add_job_row(job_id, "Nginx", "1.25", "upstream", product_name="", max_logs=1)
        
        scraped_text = ScrapedText("raw log text")
        scraped_text.queries = ["query1"]
        scraped_text.raw_results_count = 1
        scraped_text.visited_urls = ["https://example.com/diag"]
        scraped_text.search_results_ranked = [{"rank": 1, "url": "https://example.com/diag", "domain": "example.com", "source_rank": 3, "title": "diag source"}]
        scraped_text.url_info_map = {
            "https://example.com/diag": {
                "Platform": "Nginx",
                "Product": "",
                "Log Type": "upstream",
                "Search Query Used": "query1",
                "URL": "https://example.com/diag",
                "Title": "diag source",
                "Source Type": "DOCUMENT",
                "Source Rank": 3,
                "Search Rank": 1,
                "Search Position": 1,
                "Crawled": "Yes",
                "Logs Extracted": 1,
                "Logs Validated": 1,
                "Logs Rejected": 0,
                "Status": "validated",
                "Reason": "",
                "Failure Stage": ""
            }
        }
        scraped_text.search_diagnostics = [{
            "Platform": "Nginx",
            "Product": "",
            "Log Type": "upstream",
            "Query": "query1",
            "Provider": "DuckDuckGo",
            "HTTP Status": 200,
            "Results Parsed": 1,
            "URLs Returned": 1,
            "URLs Crawled": 1,
            "URLs Rejected": 0,
            "Duration": 0.15,
            "Failure Reason": "success"
        }]
        
        mock_web.return_value = scraped_text
        
        batch_processor.process_batch_job(job_id)
        job = db_manager.get_job(job_id)
        
        with zipfile.ZipFile(job["zip_path"], "r") as z:
            temp_dir = tempfile.mkdtemp()
            xlsx_path = z.extract("diagnostic_report.xlsx", temp_dir)
            try:
                wb = openpyxl.load_workbook(xlsx_path)
                self.assertIn("SEARCH_DIAGNOSTICS", wb.sheetnames)
                ws = wb["SEARCH_DIAGNOSTICS"]
                rows = list(ws.iter_rows(values_only=True))
                
                # Check headers
                self.assertEqual(rows[0][0], "Platform")
                self.assertEqual(rows[0][4], "Provider")
                self.assertEqual(rows[0][7], "URLs Returned")
                
                # Check data row
                self.assertEqual(rows[1][0], "Nginx")
                self.assertEqual(rows[1][4], "DuckDuckGo")
                self.assertEqual(rows[1][5], 200)
                self.assertEqual(rows[1][7], 1)
                
            finally:
                shutil.rmtree(temp_dir)

    @patch('backend.batch_processor.collect_logs_from_web')
    def test_low_value_url_reconciliation(self, mock_web):
        # Test that low-value URLs (marked as pre_crawl_filter status) reconcile correctly
        import tempfile
        import shutil
        import zipfile
        import openpyxl
        from backend.crawler import ScrapedText
        
        job_id = "test_low_value_job"
        db_manager.create_job(job_id, 1)
        db_manager.add_job_row(job_id, "Nginx", "1.25", "upstream", product_name="", max_logs=1)
        
        scraped_text = ScrapedText("raw log text")
        scraped_text.queries = ["query1"]
        scraped_text.raw_results_count = 2
        scraped_text.visited_urls = ["https://example.com/ok"]
        scraped_text.search_results_ranked = [
            {"rank": 1, "url": "https://example.com/ok", "domain": "example.com", "source_rank": 3, "title": "good source"},
            {"rank": 2, "url": "https://gist.github.com/search", "domain": "gist.github.com", "source_rank": 1, "title": "low value page"}
        ]
        scraped_text.url_info_map = {
            "https://example.com/ok": {
                "Platform": "Nginx",
                "Product": "",
                "Log Type": "upstream",
                "Search Query Used": "query1",
                "URL": "https://example.com/ok",
                "Title": "good source",
                "Source Type": "DOCUMENT",
                "Source Rank": 3,
                "Search Rank": 1,
                "Search Position": 1,
                "Crawled": "Yes",
                "Logs Extracted": 1,
                "Logs Validated": 1,
                "Logs Rejected": 0,
                "Status": "validated",
                "Reason": "",
                "Failure Stage": ""
            },
            "https://gist.github.com/search": {
                "Platform": "Nginx",
                "Product": "",
                "Log Type": "upstream",
                "Search Query Used": "query1",
                "URL": "https://gist.github.com/search",
                "Title": "low value page",
                "Source Type": "BLOG",
                "Source Rank": 1,
                "Search Rank": 2,
                "Search Position": 2,
                "Crawled": "No",
                "Logs Extracted": 0,
                "Logs Validated": 0,
                "Logs Rejected": 0,
                "Status": "low_value_url",
                "Reason": "Navigation page / Homepage / Search page",
                "Failure Stage": "pre_crawl_filter"
            }
        }
        scraped_text.search_diagnostics = []
        
        mock_web.return_value = scraped_text
        
        batch_processor.process_batch_job(job_id)
        job = db_manager.get_job(job_id)
        
        # Verify job completed successfully (no reconciliation errors)
        self.assertEqual(job["status"], "completed")
        self.assertEqual(job["completed_rows"], 1)
        
        # Verify ZIP contains correct sheets and they contain the low-value URL
        with zipfile.ZipFile(job["zip_path"], "r") as z:
            temp_dir = tempfile.mkdtemp()
            xlsx_path = z.extract("diagnostic_report.xlsx", temp_dir)
            try:
                wb = openpyxl.load_workbook(xlsx_path)
                self.assertIn("SOURCE_AUDIT", wb.sheetnames)
                self.assertIn("UNPROCESSED_SOURCES", wb.sheetnames)
                
                ws_audit = wb["SOURCE_AUDIT"]
                audit_rows = list(ws_audit.iter_rows(values_only=True))
                # 1 header + 2 data rows = 3 rows total
                self.assertEqual(len(audit_rows), 3)
                
                # Check low value row status in SOURCE_AUDIT
                self.assertEqual(audit_rows[2][4], "https://gist.github.com/search")
                self.assertEqual(audit_rows[2][12], "low_value_url") # Final Status
                
                ws_unprocessed = wb["UNPROCESSED_SOURCES"]
                unp_rows = list(ws_unprocessed.iter_rows(values_only=True))
                # 1 header + 1 data row (gist.github.com/search because logs_validated == 0) = 2 rows
                self.assertEqual(len(unp_rows), 2)
                self.assertEqual(unp_rows[1][4], "https://gist.github.com/search")
                self.assertEqual(unp_rows[1][8], "low_value_url") # Status
                self.assertEqual(unp_rows[1][9], "pre_crawl_filter") # Failure Stage
                
            finally:
                shutil.rmtree(temp_dir)

    def test_validated_logs_repository(self):
        # Create a temporary database file
        fd, temp_db_path = tempfile.mkstemp(suffix=".db")
        os.close(fd)
        
        # Patch the environment variable REPO_DB_PATH
        old_repo_db_path = os.environ.get("REPO_DB_PATH")
        os.environ["REPO_DB_PATH"] = temp_db_path
        db_manager.init_repo_db()
        
        try:
            # 1. Verify schema is initialized
            conn = db_manager.get_repo_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tables = [r[0] for r in cursor.fetchall()]
            self.assertIn("validated_logs", tables)
            self.assertIn("repository_metrics", tables)
            self.assertIn("repository_discovery_history", tables)
            
            # 2. Insert valid logs
            mock_logs = [
                {
                    "original_log": "2026-06-23T10:00:00.123Z [ERROR] nginx_upstream: upstream timed out while connecting to backend server",
                    "message": "upstream timed out while connecting to backend server",
                    "severity": "ERROR",
                    "source_url": "https://github.com/nginx/issues/123",
                    "source_title": "Nginx Upstream Timeout issue",
                    "query_used": "nginx upstream error logs",
                    "validation": {
                        "valid": True,
                        "confidence": 95,
                        "source_type": "GITHUB_ISSUE",
                        "source_rank": 1
                    }
                }
            ]
            
            inserted, duplicates = db_manager.insert_validated_logs(
                mock_logs,
                job_platform="Nginx",
                job_product_name="Web Server",
                job_log_type="upstream"
            )
            self.assertEqual(inserted, 1)
            self.assertEqual(duplicates, 0)
            
            # Verify observability/RCA fields
            cursor.execute("SELECT platform, vendor, platform_category, source_domain, log_severity, process_name FROM validated_logs")
            row = cursor.fetchone()
            self.assertEqual(row["platform"], "Nginx")
            self.assertEqual(row["vendor"], "NGINX / F5")
            self.assertEqual(row["platform_category"], "Web Server")
            self.assertEqual(row["source_domain"], "github.com")
            self.assertEqual(row["log_severity"], "ERROR")
            self.assertEqual(row["process_name"], "nginx_upstream")
            
            # 3. Duplicate insert
            inserted2, duplicates2 = db_manager.insert_validated_logs(
                mock_logs,
                job_platform="Nginx",
                job_product_name="Web Server",
                job_log_type="upstream"
            )
            self.assertEqual(inserted2, 0)
            self.assertEqual(duplicates2, 1)
            
            # Verify repository metrics
            cursor.execute("SELECT value FROM repository_metrics WHERE key = 'duplicates_skipped'")
            val = cursor.fetchone()[0]
            self.assertEqual(val, 1)
            
            # Verify discovery history has 2 records
            cursor.execute("SELECT status FROM repository_discovery_history ORDER BY id ASC")
            hist = [r[0] for r in cursor.fetchall()]
            self.assertEqual(hist, ["inserted", "duplicate_skipped"])
            
            # 4. Health data
            health = db_manager.get_repository_health_data()
            self.assertEqual(health["total_logs"], 1)
            self.assertEqual(health["duplicates_skipped"], 1)
            self.assertEqual(health["unique_sources"], 1)
            self.assertEqual(health["unique_platforms"], 1)
            
            # 5. Stats data
            stats = db_manager.get_repository_stats_for_sheet()
            self.assertEqual(len(stats), 1)
            self.assertEqual(stats[0]["platform"], "Nginx")
            self.assertEqual(stats[0]["total_logs"], 1)
            
            conn.close()
        finally:
            # Restore environment variable and clean up file
            if old_repo_db_path is not None:
                os.environ["REPO_DB_PATH"] = old_repo_db_path
            else:
                os.environ.pop("REPO_DB_PATH", None)
            
            if os.path.exists(temp_db_path):
                # SQLite sometimes locks the file momentarily, so try deleting it
                for _ in range(5):
                    try:
                        os.remove(temp_db_path)
                        break
                    except Exception:
                        import time
                        time.sleep(0.1)

if __name__ == '__main__':
    unittest.main()
