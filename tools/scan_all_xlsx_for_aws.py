import sys
import os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
import glob
import openpyxl
import os

xlsx_files = glob.glob('backend/outputs/**/*.xlsx', recursive=True)
for f in xlsx_files:
    try:
        wb = openpyxl.load_workbook(f, read_only=True)
        print(f"\n==================== File: {f} ====================")
        print("Sheets:", wb.sheetnames)
        if "SUMMARY" in wb.sheetnames:
            ws = wb["SUMMARY"]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                header = rows[0]
                for r_idx, r in enumerate(rows[1:], 1):
                    r_str = " | ".join([str(x) for x in r if x is not None])
                    if ("aws" in r_str.lower() or "amazon" in r_str.lower()) and len(r) > 4:
                        try:
                            val = int(r[4])
                        except (ValueError, TypeError):
                            val = 0
                        if val > 30:
                            print(f"SUMMARY Row {r_idx} matches (Search Results Count={val}):")
                            print("  ", r)
        # Check source audit or others
        for sheet in ["SOURCE_AUDIT", "UNPROCESSED_SOURCES", "SEARCH_DIAGNOSTICS"]:
            if sheet in wb.sheetnames:
                ws = wb[sheet]
                rows = list(ws.iter_rows(values_only=True))
                print(f"  {sheet}: {len(rows)} total rows")
    except Exception as e:
        print(f"Error reading {f}: {e}")
